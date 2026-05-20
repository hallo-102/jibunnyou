# =========================
# keibayosou_pipeline.py
# =========================
# penalties（新規ファイル）を呼び出す形に整理した完全版。
# 旧：pipeline内に _calc_extra_penalty / _calc_rest_dist_risk を直書き
# 新：keibayosou_penalties.py に分離し、ここは「流れ」に集中
#
# 今回の追加修正:
# - レース登録馬の過去走情報が不足しているレースを予想対象から除外
# - 判定基準:
#   レースIDごとに
#     1) 今走レース情報シートの頭数
#     2) feat_df 側で過去走特徴量を作れた馬名ユニーク数
#   を比較し、
#     過去走特徴量を作れた馬名ユニーク数 < 頭数
#   のレースは予想除外する
#
# 例:
#   頭数=16、過去走あり馬数=15 → 初出走馬などがいるとみなし、そのレースは除外
#
# ※ 既存ロジックはなるべくそのまま維持しています。

# -*- coding: utf-8 -*-
"""パイプライン全体の実行フローをまとめたモジュール。"""

from __future__ import annotations

import os
import re
import shutil
from datetime import datetime
from typing import Dict, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from keibayosou_config import (
    TARGET_SHEET,
    NOW_SHEET,
    ALPHA,
    EXTRA_ALPHA,
    DL_PROB_BLEND,
    DL_RANK_BLEND,
    DL_SCORE_BONUS,
    RACE_LEVEL_XLSX,
    BASE_TIME_XLSX,
    ODDS_CSV,
    SUCCESS_REPORT,
    FEAT_COLS,
    JAPANESE_FEATURE_NAMES,
)
from keibayosou_features import (
    _normalize_rid_series,
    _normalize_umaban_series,
    apply_weights,
    build_calc_favorite_risk,
    build_features_from_excel,
    normalize_score,
    score_sum,
)
from keibayosou_loaders import load_base_time, load_odds_csv, load_race_levels
from keibayosou_penalties import calc_extra_penalty, calc_rest_dist_risk
from keibayosou_utils import (
    _build_feature_sheet_for_export,
    _normalize_place,
    _normalize_surface,
    _to_int,
)


def compute_scores_with_pipeline_logic(
    feat_df: pd.DataFrame,
    place_map: Dict[str, str],
    surface_map: Dict[str, str],
    calc_fav_risk,
    alpha: float = ALPHA,
    extra_alpha: float = EXTRA_ALPHA,
) -> pd.DataFrame:
    """pipeline 本番と同じ式で total / score / rank を計算する。"""
    out = feat_df.copy()

    out["total_raw"] = out.apply(
        lambda r: score_sum(
            apply_weights(
                {k: r.get(k) for k in FEAT_COLS},
                place=_normalize_place(place_map.get(str(r.get("rid_str", "")))),
                surface=_normalize_surface(surface_map.get(str(r.get("rid_str", "")))),
            )
        ),
        axis=1,
    )

    out["favorite_risk"] = out.apply(calc_fav_risk, axis=1)
    out["rest_dist_risk"] = out.apply(calc_rest_dist_risk, axis=1)
    out["extra_penalty"] = out.apply(
        lambda r: calc_extra_penalty(r, rest_dist_risk=r.get("rest_dist_risk")),
        axis=1,
    )

    # dl_score は 0.5 を中立点として total に反映する。
    # 1回目は dl 系列が無いので 0.5 扱いとなり、2回目だけ順位へ効く。
    if "dl_score" not in out.columns:
        out["dl_score"] = 0.5

    out["dl_bonus"] = (pd.to_numeric(out["dl_score"], errors="coerce").fillna(0.5) - 0.5) * DL_SCORE_BONUS
    out["total"] = (
        out["total_raw"]
        + out["dl_bonus"]
        - alpha * out["favorite_risk"]
        - extra_alpha * out["extra_penalty"]
    )
    out["score"] = out.groupby("rid_str")["total"].transform(normalize_score).round(2)
    out["rank"] = out.groupby("rid_str")["score"].rank("dense", ascending=False).astype(int)
    return out


# ================================================================
# 今回追加：過去走不足レースを除外するための補助関数
# ================================================================
def _pick_first_existing_col(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """候補の中から、最初に存在する列名を返す。"""
    for col in candidates:
        if col in df.columns:
            return col
    return None


def _exclude_races_with_missing_history(
    merged: pd.DataFrame,
    feat_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    レース登録頭数に対して、実際に過去走特徴量を作れた馬数が不足しているレースを除外する。

    判定:
      レースIDごとに
        実際に過去走特徴量あり馬数 < 登録頭数
      なら、そのレースは予想対象外

    ここでいう「実際に過去走特徴量あり馬」とは、
    feat_df に行があるだけではなく、
    ta_n（または同等列）が 1 以上ある馬を指す。
    つまり、行だけ存在して主要特徴量が実質作れていない馬は数えない。

    merged:
      build_features_from_excel の戻り値（今走ベース）
    feat_df:
      build_features_from_excel の戻り値（過去走から特徴量を作れた馬だけ入る想定）

    戻り値:
      (filtered_merged, filtered_feat_df)
    """
    if merged is None or merged.empty:
        return merged, feat_df

    if feat_df is None or feat_df.empty:
        print("[WARN] feat_df が空のため、全レースを予想除外します")
        return merged.iloc[0:0].copy(), feat_df.iloc[0:0].copy()

    work_merged = merged.copy()
    work_feat = feat_df.copy()

    if "rid_str" not in work_merged.columns:
        if "レースID" in work_merged.columns:
            work_merged["rid_str"] = work_merged["レースID"]
        else:
            print("[WARN] merged に rid_str/レースID が無いため、過去走不足レース除外をスキップします")
            return merged, feat_df

    if "rid_str" not in work_feat.columns:
        print("[WARN] feat_df に rid_str が無いため、過去走不足レース除外をスキップします")
        return merged, feat_df

    work_merged["rid_str"] = _normalize_rid_series(work_merged["rid_str"])
    work_feat["rid_str"] = _normalize_rid_series(work_feat["rid_str"])

    # merged 側の馬名列候補
    merged_name_col = _pick_first_existing_col(work_merged, ["馬名", "horse_name", "name"])
    feat_name_col = _pick_first_existing_col(work_feat, ["馬名", "horse_name", "name"])

    if feat_name_col is None:
        print("[WARN] feat_df に馬名列が無いため、過去走不足レース除外をスキップします")
        return merged, feat_df

    # 登録頭数
    field_col = _pick_first_existing_col(work_merged, ["頭数", "頭 数", "field_size"])

    if field_col is not None:
        race_field_df = (
            work_merged[["rid_str", field_col]]
            .copy()
            .assign(**{field_col: pd.to_numeric(work_merged[field_col], errors="coerce")})
            .groupby("rid_str", as_index=False)[field_col]
            .first()
            .rename(columns={field_col: "registered_field_size"})
        )
    else:
        # 念のためのフォールバック
        if merged_name_col is None:
            print("[WARN] merged に頭数列も馬名列も無いため、過去走不足レース除外をスキップします")
            return merged, feat_df

        race_field_df = (
            work_merged[["rid_str", merged_name_col]]
            .dropna(subset=[merged_name_col])
            .copy()
        )
        race_field_df[merged_name_col] = race_field_df[merged_name_col].astype(str).str.strip()
        race_field_df = (
            race_field_df.groupby("rid_str", as_index=False)[merged_name_col]
            .nunique()
            .rename(columns={merged_name_col: "registered_field_size"})
        )

    # 実際に過去走特徴量を作れた馬数
    # 重要:
    # feat_df に行があるだけでは数えず、ta_n（同等列）が 1 以上ある馬だけ数える。
    feat_horse_df = work_feat[["rid_str", feat_name_col]].dropna(subset=[feat_name_col]).copy()
    feat_horse_df[feat_name_col] = feat_horse_df[feat_name_col].astype(str).str.strip()
    feat_horse_df = feat_horse_df[feat_horse_df[feat_name_col] != ""]

    ta_n_col = _pick_first_existing_col(work_feat, ["ta_n", "f_race_count", "レース数"])
    if ta_n_col is not None:
        feat_horse_df[ta_n_col] = pd.to_numeric(work_feat.loc[feat_horse_df.index, ta_n_col], errors="coerce")
        feat_horse_df = feat_horse_df[feat_horse_df[ta_n_col].fillna(0) > 0]
    else:
        # ta_n 相当列が無い場合は、既存挙動に近い保険として全件を数える
        print("[WARN] feat_df に ta_n/f_race_count/レース数 列が無いため、行ベースで過去走あり馬数を数えます")

    history_count_df = (
        feat_horse_df.groupby("rid_str", as_index=False)[feat_name_col]
        .nunique()
        .rename(columns={feat_name_col: "history_horse_count"})
    )

    audit_df = pd.merge(
        race_field_df,
        history_count_df,
        on="rid_str",
        how="left",
    )

    audit_df["history_horse_count"] = pd.to_numeric(audit_df["history_horse_count"], errors="coerce").fillna(0).astype(int)
    audit_df["registered_field_size"] = pd.to_numeric(audit_df["registered_field_size"], errors="coerce")

    exclude_rids = audit_df.loc[
        audit_df["registered_field_size"].notna()
        & (audit_df["history_horse_count"] < audit_df["registered_field_size"]),
        "rid_str",
    ].astype(str).tolist()

    if exclude_rids:
        preview_df = audit_df[audit_df["rid_str"].isin(exclude_rids)].copy()
        preview_df = preview_df.sort_values(["rid_str"], kind="mergesort")

        print(
            f"[INFO] 過去走不足レースを除外します: {len(exclude_rids)}レース "
            f"(全{len(audit_df)}レース中)"
        )
        for _, r in preview_df.iterrows():
            print(
                f"[INFO] 除外 rid={r['rid_str']} "
                f"登録頭数={int(r['registered_field_size']) if pd.notna(r['registered_field_size']) else 'NaN'} "
                f"過去走あり馬数={int(r['history_horse_count'])}"
            )
    else:
        print("[INFO] 過去走不足による除外レースはありません")

    filtered_merged = work_merged.loc[~work_merged["rid_str"].astype(str).isin(exclude_rids)].copy()
    filtered_feat = work_feat.loc[~work_feat["rid_str"].astype(str).isin(exclude_rids)].copy()

    return filtered_merged, filtered_feat



# ================================================================
# 追加で作るシート名（過去の出力と互換）
# ================================================================
BET_SHEET = "買い目_レース別1行"
ROI_FOCUS_BET_SHEET = "回収率重視_買い目候補"

ROI_FOCUS_BET_COLUMNS = [
    "レースID",
    "レース名",
    "発走時刻",
    "場所",
    "コース",
    "馬場",
    "頭数",
    "クラス",
    "score1",
    "score2",
    "gap12",
    "dango_2_5",
    "1位馬番",
    "2位馬番",
    "3位馬番",
    "4位馬番",
    "5位馬番",
    "購入判定",
    "購入理由",
    "3連複1点目_馬番1",
    "3連複1点目_馬番2",
    "3連複1点目_馬番3",
    "3連複2点目_馬番1",
    "3連複2点目_馬番2",
    "3連複2点目_馬番3",
    "3連複3点目_馬番1",
    "3連複3点目_馬番2",
    "3連複3点目_馬番3",
    "3連複_点数",
    "3連複_金額",
    "馬連1点目_馬番1",
    "馬連1点目_馬番2",
    "馬連2点目_馬番1",
    "馬連2点目_馬番2",
    "馬連_点数",
    "馬連_金額",
    "ワイド1点目_馬番1",
    "ワイド1点目_馬番2",
    "ワイド2点目_馬番1",
    "ワイド2点目_馬番2",
    "ワイド_点数",
    "ワイド_金額",
    "合計購入金額",
]


# ================================================================
# 買い目シート作成（過去版の出力互換）
# ================================================================
def _to_int_safe(x: object) -> Optional[int]:
    try:
        if pd.isna(x):
            return None
        if isinstance(x, str) and x.strip() == "":
            return None
        return int(float(x))
    except Exception:
        return None


def _to_float_safe(x: object) -> Optional[float]:
    try:
        if pd.isna(x):
            return None
        if isinstance(x, str) and x.strip() == "":
            return None
        return float(x)
    except Exception:
        return None


def _normalize_race_key(x: object) -> Optional[str]:
    """Excel読み戻し時の 123.0 のような表記ゆれをレースID比較用にそろえる。"""
    if pd.isna(x):
        return None
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        if np.isfinite(x) and float(x).is_integer():
            return str(int(x))
        return str(x).strip()

    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return None
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return m.group(1)
    return s


def _to_umaban_int_for_roi(x: object) -> Optional[int]:
    """買い目用の馬番を整数化する。欠損・非整数・0以下は異常値として扱う。"""
    if pd.isna(x):
        return None
    if isinstance(x, str) and x.strip() == "":
        return None

    try:
        v = float(str(x).strip())
    except Exception:
        return None

    if not np.isfinite(v) or not v.is_integer():
        return None

    n = int(v)
    if n <= 0:
        return None
    return n


def _class_map_from_now_df(now_df: Optional[pd.DataFrame]) -> Dict[str, object]:
    """今走シートから レースID -> クラス の対応を作る。"""
    if now_df is None or not isinstance(now_df, pd.DataFrame) or now_df.empty:
        return {}

    now = now_df.copy()
    race_col = _pick_first_existing_col(now, ["レースID", "rid_str", "race_id"])
    class_col = _pick_first_existing_col(now, ["クラス", "レースクラス", "条件クラス"])
    if race_col is None or class_col is None:
        return {}

    now["_race_key_for_roi"] = now[race_col].map(_normalize_race_key)
    now = now.dropna(subset=["_race_key_for_roi"])
    if now.empty:
        return {}

    return now.groupby("_race_key_for_roi")[class_col].first().to_dict()


def build_roi_focus_bet_sheet(
    bet_df: pd.DataFrame,
    now_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    回収率検証で採用した条件に合うレースだけを、購入候補シート用に1レース1行へ整形する。

    注意:
    - 着順、払戻、的中判定などの結果情報は使わない。
    - score1/score2/gap12/dango_2_5/1位馬番〜5位馬番は既存の買い目シート値を使う。
    """
    if bet_df is None or not isinstance(bet_df, pd.DataFrame) or bet_df.empty:
        return pd.DataFrame(columns=ROI_FOCUS_BET_COLUMNS)

    work = bet_df.copy()
    class_map = _class_map_from_now_df(now_df)

    if "クラス" not in work.columns:
        work["クラス"] = pd.NA

    if class_map:
        race_col = _pick_first_existing_col(work, ["レースID", "rid_str", "race_id"])
        if race_col is not None:
            race_key = work[race_col].map(_normalize_race_key)
            mapped_class = race_key.map(class_map)
            work["クラス"] = work["クラス"].combine_first(mapped_class)

    rows: list[dict] = []

    for _, row in work.iterrows():
        score1 = _to_float_safe(row.get("score1"))
        score2 = _to_float_safe(row.get("score2"))
        gap12 = _to_float_safe(row.get("gap12"))
        dango_2_5 = _to_float_safe(row.get("dango_2_5"))

        if score1 is None or gap12 is None:
            continue
        if not (gap12 >= 2.0 and score1 >= 65.0):
            continue

        race_id = row.get("レースID", row.get("rid_str", pd.NA))
        ranks = [_to_umaban_int_for_roi(row.get(f"{i}位馬番")) for i in range(1, 6)]
        if any(v is None for v in ranks) or len(set(ranks)) != 5:
            print(
                f"[WARN] '{ROI_FOCUS_BET_SHEET}' 出力除外: "
                f"レースID={race_id} 1位〜5位馬番に欠損/重複/変換不可があります"
            )
            continue

        rank1, rank2, rank3, rank4, rank5 = [int(v) for v in ranks]

        rows.append(
            {
                "レースID": race_id,
                "レース名": row.get("レース名", pd.NA),
                "発走時刻": row.get("発走時刻", pd.NA),
                "場所": row.get("場所", pd.NA),
                "コース": row.get("コース", pd.NA),
                "馬場": row.get("馬場", pd.NA),
                "頭数": row.get("頭数", pd.NA),
                "クラス": row.get("クラス", pd.NA),
                "score1": score1,
                "score2": score2,
                "gap12": gap12,
                "dango_2_5": dango_2_5,
                "1位馬番": rank1,
                "2位馬番": rank2,
                "3位馬番": rank3,
                "4位馬番": rank4,
                "5位馬番": rank5,
                "購入判定": "購入",
                "購入理由": "gap12>=2 かつ score1>=65",
                "3連複1点目_馬番1": rank1,
                "3連複1点目_馬番2": rank3,
                "3連複1点目_馬番3": rank2,
                "3連複2点目_馬番1": rank1,
                "3連複2点目_馬番2": rank3,
                "3連複2点目_馬番3": rank4,
                "3連複3点目_馬番1": rank1,
                "3連複3点目_馬番2": rank3,
                "3連複3点目_馬番3": rank5,
                "3連複_点数": 3,
                "3連複_金額": 300,
                "馬連1点目_馬番1": rank1,
                "馬連1点目_馬番2": rank3,
                "馬連2点目_馬番1": rank2,
                "馬連2点目_馬番2": rank3,
                "馬連_点数": 2,
                "馬連_金額": 200,
                "ワイド1点目_馬番1": rank1,
                "ワイド1点目_馬番2": rank3,
                "ワイド2点目_馬番1": rank2,
                "ワイド2点目_馬番2": rank3,
                "ワイド_点数": 2,
                "ワイド_金額": 200,
                "合計購入金額": 700,
            }
        )

    out = pd.DataFrame(rows)
    for c in ROI_FOCUS_BET_COLUMNS:
        if c not in out.columns:
            out[c] = pd.NA
    out = out[ROI_FOCUS_BET_COLUMNS]

    int_cols = [
        c
        for c in ROI_FOCUS_BET_COLUMNS
        if "馬番" in c or c.endswith("_点数") or c.endswith("_金額") or c == "合計購入金額"
    ]
    if not out.empty:
        for c in int_cols:
            out[c] = out[c].astype("int64")

    return out


def _delete_excel_sheet_if_exists(excel_path: str, sheet_name: str) -> bool:
    """同名シートをいったん削除して、次の書き込みで末尾に追加できる状態にする。"""
    if not os.path.exists(excel_path):
        return False

    wb = None
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            return False
        if len(wb.sheetnames) <= 1:
            return False

        del wb[sheet_name]
        wb.save(excel_path)
        return True
    finally:
        if wb is not None:
            wb.close()


def _format_roi_focus_bet_worksheet(excel_path: str) -> None:
    """回収率重視_買い目候補 シートだけ最低限見やすく整える。"""
    wb = None
    try:
        wb = load_workbook(excel_path)
        if ROI_FOCUS_BET_SHEET not in wb.sheetnames:
            return

        ws = wb[ROI_FOCUS_BET_SHEET]
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        sample_max_row = min(ws.max_row, 200)
        for col_idx in range(1, ws.max_column + 1):
            values = [
                ws.cell(row=row_idx, column=col_idx).value
                for row_idx in range(1, sample_max_row + 1)
            ]
            max_len = max((len(str(v)) for v in values if v is not None), default=0)
            width = min(max(max_len + 2, 10), 32)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。整形をスキップします: {excel_path}")
    except Exception as e:
        print(f"[WARN] '{ROI_FOCUS_BET_SHEET}' の整形に失敗しました: {e}")
    finally:
        if wb is not None:
            wb.close()


def append_roi_focus_bet_sheet_to_excel(out_excel_path: str) -> int:
    """
    既存の 買い目_レース別1行 から、回収率重視_買い目候補 をブック末尾に作り直す。
    """
    if not os.path.exists(out_excel_path):
        print(f"[WARN] 出力Excelが見つからないため、'{ROI_FOCUS_BET_SHEET}' 作成をスキップ: {out_excel_path}")
        return 0

    xls = None
    try:
        xls = pd.ExcelFile(out_excel_path, engine="openpyxl")
        if BET_SHEET not in xls.sheet_names:
            print(f"[WARN] '{BET_SHEET}' シートが無いため、'{ROI_FOCUS_BET_SHEET}' 作成をスキップします")
            return 0

        bet_df = pd.read_excel(out_excel_path, sheet_name=BET_SHEET, engine="openpyxl")
        now_df = (
            pd.read_excel(out_excel_path, sheet_name=NOW_SHEET, engine="openpyxl")
            if NOW_SHEET in xls.sheet_names
            else pd.DataFrame()
        )
    except Exception as e:
        print(f"[WARN] '{ROI_FOCUS_BET_SHEET}' 用のExcel読み込みに失敗しました: {e}")
        return 0
    finally:
        if xls is not None:
            xls.close()

    roi_focus_bet_df = build_roi_focus_bet_sheet(bet_df=bet_df, now_df=now_df)

    try:
        _delete_excel_sheet_if_exists(out_excel_path, ROI_FOCUS_BET_SHEET)
        with pd.ExcelWriter(out_excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            roi_focus_bet_df.to_excel(writer, sheet_name=ROI_FOCUS_BET_SHEET, index=False)
        _format_roi_focus_bet_worksheet(out_excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。Excelを閉じてから再実行してください: {out_excel_path}")
        return 0
    except Exception as e:
        print(f"[WARN] '{ROI_FOCUS_BET_SHEET}' のExcel書き込みに失敗しました: {e}")
        return 0

    print(f"[INFO] '{ROI_FOCUS_BET_SHEET}' 作成完了: {len(roi_focus_bet_df)}レース -> {out_excel_path}")
    return int(len(roi_focus_bet_df))


def _build_bet_sheet(
    feat_export: pd.DataFrame,
    now_export: pd.DataFrame,
    odds_df: Optional[pd.DataFrame] = None,
    gap_min: float = 3.5,
    extra_th: float = 0.8,
    rest_th: float = 0.4,
) -> pd.DataFrame:
    """TARGET（feat_export）と今走（now_export）から
    買い目_レース別1行 を作る。

    ここでのポイント（過去出力の互換）：
    - score1/score2/gap12 は「上位2頭（同点含む）」で計算
    - dango_2_5 は「rank2 と rank5 の score差」で計算（rank不足なら 999）
    """

    # ----------------------------
    # オッズ（rid_str, umaban -> tansho）マップ
    # ----------------------------
    odds_map: Dict[Tuple[str, int], float] = {}
    if odds_df is not None and isinstance(odds_df, pd.DataFrame) and not odds_df.empty:
        od = odds_df.copy()
        if "rid_str" in od.columns:
            od["rid_str"] = od["rid_str"].astype(str)
        if "umaban" in od.columns:
            od["umaban"] = pd.to_numeric(od["umaban"], errors="coerce").astype("Int64")
        if "tansho" in od.columns:
            od["tansho"] = pd.to_numeric(od["tansho"], errors="coerce")

        for r, u, t in zip(od.get("rid_str", []), od.get("umaban", []), od.get("tansho", [])):
            if pd.isna(r) or pd.isna(u) or pd.isna(t):
                continue
            odds_map[(str(r), int(u))] = float(t)

    # ----------------------------
    # 今走：レース情報（1レース1行）
    # ----------------------------
    now = now_export.copy()
    if "rid_str" not in now.columns:
        if "レースID" in now.columns:
            now["rid_str"] = now["レースID"].astype(str)
        else:
            now["rid_str"] = pd.NA
    now["rid_str"] = now["rid_str"].astype(str)

    race_info_cols = ["レースID", "レース名", "発走時刻", "場所", "コース", "馬場", "頭数", "レース種別", "クラス"]
    for c in race_info_cols:
        if c not in now.columns:
            now[c] = pd.NA

    race_info = now.groupby("rid_str", as_index=False)[race_info_cols].first()

    # ----------------------------
    # TARGET：rid_str+馬番+score+rank を前提
    # ----------------------------
    ft = feat_export.copy()
    if "rid_str" not in ft.columns:
        raise ValueError("TARGET（feat_export）に rid_str 列がありません")

    ft["rid_str"] = ft["rid_str"].astype(str)

    if "馬番" not in ft.columns:
        for cand in ["umaban", "馬 番", "馬番 "]:
            if cand in ft.columns:
                ft["馬番"] = ft[cand]
                break

    ft["馬番"] = pd.to_numeric(ft["馬番"], errors="coerce").astype("Int64")
    ft["score"] = pd.to_numeric(ft.get("score", pd.Series([pd.NA] * len(ft))), errors="coerce")
    ft["rank"] = pd.to_numeric(ft.get("rank", pd.Series([pd.NA] * len(ft))), errors="coerce").astype("Int64")

    # 休養×距離差リスク列名のゆらぎ吸収
    rest_col = (
        "休養×距離差リスク"
        if "休養×距離差リスク" in ft.columns
        else ("rest_dist_risk" if "rest_dist_risk" in ft.columns else None)
    )
    if rest_col is None:
        ft["休養×距離差リスク"] = 0.0
        rest_col = "休養×距離差リスク"

    bet_rows: list[dict] = []

    for rid, sub in ft.groupby("rid_str", sort=True):
        # 上位馬番（同点は馬番昇順）を取る
        sub2 = sub.sort_values(["score", "馬番"], ascending=[False, True], kind="mergesort")
        top7 = sub2.head(7)

        horses7 = [_to_int_safe(v) for v in top7["馬番"].tolist()]
        horses6 = horses7[:6] + [None] * (6 - len(horses7[:6]))
        horses7 = horses7[:7] + [None] * (7 - len(horses7[:7]))

        # score1/score2/gap12（上位2頭）
        top_scores = [_to_float_safe(v) for v in top7["score"].tolist()]
        score1 = top_scores[0] if len(top_scores) > 0 else None
        score2 = top_scores[1] if len(top_scores) > 1 else score1
        gap12 = round(float(score1 - score2), 2) if (score1 is not None and score2 is not None) else 0.0

        # dango_2_5：rank2とrank5のscore差（rank不足なら999）
        rank_score = sub2.dropna(subset=["rank", "score"]).groupby("rank")["score"].max().sort_index()
        if 2 in rank_score.index and 5 in rank_score.index:
            dango_2_5 = round(float(rank_score.loc[2] - rank_score.loc[5]), 2)
        else:
            dango_2_5 = 999.0

        # 1位馬（先頭行）のリスク
        top1 = sub2.iloc[0] if len(sub2) > 0 else pd.Series()
        fav = _to_float_safe(top1.get("favorite_risk", 0.0)) or 0.0
        extra = _to_float_safe(top1.get("extra_penalty", 0.0)) or 0.0
        rest = _to_float_safe(top1.get(rest_col, 0.0)) or 0.0

        risk_high = (extra >= extra_th) or (rest >= rest_th)

        # レース情報（無ければ最低限）
        info = race_info[race_info["rid_str"] == rid]
        info_row = info.iloc[0].to_dict() if not info.empty else {c: pd.NA for c in race_info_cols}
        if pd.isna(info_row.get("レースID")):
            info_row["レースID"] = rid

        # 単勝オッズ（上位1頭）
        odds_top1 = odds_map.get((str(rid), horses6[0])) if horses6[0] is not None else None

        # 判定ロジック（過去シート互換）
        if gap12 >= gap_min:
            if risk_high:
                rank_label = "A"
                judge = "1頭軸（相手2～6位）＋保険BOX（上位5）"
                reason = f"gap12={gap12:.2f}だが1位リスク高（fav={fav:.2f},extra={extra:.2f},rest={rest:.2f}）→保険"
                axis_umaban = horses6[0]
                axis_opp = ",".join(str(x) for x in horses6[1:6] if x is not None)
                axis_yen = 100.0
                box_list = [x for x in horses7[:5] if x is not None]
                box_umaban = ",".join(str(x) for x in box_list) if box_list else np.nan
                box_yen = 100.0
            else:
                rank_label = "S"
                judge = "1頭軸（相手2～6位）"
                reason = f"gap12={gap12:.2f}で1位が強い + リスク低（fav={fav:.2f},extra={extra:.2f},rest={rest:.2f}）"
                axis_umaban = horses6[0]
                axis_opp = ",".join(str(x) for x in horses6[1:6] if x is not None)
                axis_yen = 100.0
                box_umaban = np.nan
                box_yen = np.nan
        else:
            rank_label = "-"
            judge = "見送り"
            reason = f"gap12={gap12:.2f}小 + 団子でもない（2～5位幅={dango_2_5:.2f}）"
            axis_umaban = np.nan
            axis_opp = np.nan
            axis_yen = np.nan
            box_umaban = np.nan
            box_yen = np.nan

        bet_rows.append(
            {
                "レースID": info_row.get("レースID", rid),
                "レース名": info_row.get("レース名", pd.NA),
                "発走時刻": info_row.get("発走時刻", pd.NA),
                "場所": info_row.get("場所", pd.NA),
                "コース": info_row.get("コース", pd.NA),
                "馬場": info_row.get("馬場", pd.NA),
                "頭数": info_row.get("頭数", pd.NA),
                "score1": score1,
                "score2": score2,
                "gap12": gap12,
                "dango_2_5": dango_2_5,
                "1位馬番": horses6[0],
                "2位馬番": horses6[1],
                "3位馬番": horses6[2],
                "4位馬番": horses6[3],
                "5位馬番": horses6[4],
                "6位馬番": horses6[5],
                "ランク(S/A/B)": rank_label,
                "判定": judge,
                "理由": reason,
                "1頭軸_馬番": axis_umaban,
                "1頭軸_相手": axis_opp,
                "1頭軸_金額": axis_yen,
                "保険BOX_馬番": box_umaban,
                "保険BOX_金額": box_yen,
                "単勝オッズ_1位": odds_top1,
            }
        )

    bet_df = pd.DataFrame(bet_rows)
    bet_cols = [
        "レースID",
        "レース名",
        "発走時刻",
        "場所",
        "コース",
        "馬場",
        "頭数",
        "score1",
        "score2",
        "gap12",
        "dango_2_5",
        "1位馬番",
        "2位馬番",
        "3位馬番",
        "4位馬番",
        "5位馬番",
        "6位馬番",
        "ランク(S/A/B)",
        "判定",
        "理由",
        "1頭軸_馬番",
        "1頭軸_相手",
        "1頭軸_金額",
        "保険BOX_馬番",
        "保険BOX_金額",
        "単勝オッズ_1位",
    ]
    for c in bet_cols:
        if c not in bet_df.columns:
            bet_df[c] = pd.NA
    bet_df = bet_df[bet_cols]

    return bet_df


# ================================================================
# Excel 出力処理
# ================================================================
def write_features_to_excel(
    src_excel: str,
    out_excel: str,
    feat_df: pd.DataFrame,
    now_df: pd.DataFrame,
    odds_df: Optional[pd.DataFrame] = None,
) -> None:
    """もとの EXCEL をコピーし、TARGET シートと今走シートを上書き。
    さらに過去版互換で 買い目_レース別1行 も作成する。
    """
    print(f"[INFO] 特徴量を {out_excel} に出力します")

    src_abs = os.path.normcase(os.path.abspath(src_excel))
    out_abs = os.path.normcase(os.path.abspath(out_excel))
    if src_abs == out_abs:
        print("[INFO] 入力Excelと出力Excelが同じため、コピーせず既存ブックを更新します")
    else:
        try:
            shutil.copy2(src_excel, out_excel)
        except PermissionError:
            stem, ext = os.path.splitext(out_excel)
            alt = f"{stem}_{datetime.now().strftime('%H%M%S')}{ext}"
            print(f"[WARN] 出力先ファイルに書き込めません（Excelで開いている可能性）: {out_excel} -> {alt}")
            shutil.copy2(src_excel, alt)
            out_excel = alt

    feat_export = _build_feature_sheet_for_export(feat_df, FEAT_COLS, JAPANESE_FEATURE_NAMES)

    # TARGETは rid_str ごとに rank 昇順（上位=1が先）に並べる
    if {"rid_str", "rank"}.issubset(feat_export.columns):
        feat_export = feat_export.copy()
        feat_export["_rid_sort"] = feat_export["rid_str"].astype(str).str.replace(r"\D", "", regex=True)
        feat_export["_rank_sort"] = pd.to_numeric(feat_export["rank"], errors="coerce")
        feat_export = feat_export.sort_values(
            ["_rid_sort", "_rank_sort"],
            ascending=[True, True],
            kind="mergesort",
            na_position="last",
        ).drop(columns=["_rid_sort", "_rank_sort"])

    now_export = now_df.copy()

    # ★追加：買い目シートを生成
    try:
        bet_df = _build_bet_sheet(feat_export=feat_export, now_export=now_export, odds_df=odds_df)
    except Exception as e:
        print(f"[WARN] '{BET_SHEET}' の作成に失敗したためスキップします: {e}")
        bet_df = pd.DataFrame()

    with pd.ExcelWriter(out_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        feat_export.to_excel(writer, sheet_name=TARGET_SHEET, index=False)
        now_export.to_excel(writer, sheet_name=NOW_SHEET, index=False)

        if not bet_df.empty:
            bet_df.to_excel(writer, sheet_name=BET_SHEET, index=False)

    if not bet_df.empty:
        append_roi_focus_bet_sheet_to_excel(out_excel)


def append_success_report(df: pd.DataFrame, report_path: str) -> None:
    """success_report.xlsx に簡易集計を追記。"""
    n_races = df["rid_str"].nunique() if "rid_str" in df.columns else 0
    n_horses = len(df)

    row = {
        "日付": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "レース数": n_races,
        "頭数": n_horses,
    }

    if os.path.exists(report_path):
        rep = pd.read_excel(report_path, engine="openpyxl")
        rep = pd.concat([rep, pd.DataFrame([row])], ignore_index=True)
    else:
        rep = pd.DataFrame([row])

    rep.to_excel(report_path, index=False)
    print(f"[INFO] success_report.xlsx を更新しました: {report_path}")


def _merge_dl_rank_override(merged: pd.DataFrame, dl_rank_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """メモリ上で作成した dl_rank / dl_prob を今走データへ結合する。"""
    if dl_rank_df is None or dl_rank_df.empty:
        return merged
    if merged.empty:
        return merged
    if not {"rid_str", "馬番"}.issubset(merged.columns):
        raise RuntimeError("今走データに rid_str / 馬番 列が無く、DL順位データを結合できません")

    required_cols = {"rid_str", "馬番"}
    if not required_cols.issubset(dl_rank_df.columns):
        missing = ", ".join(sorted(required_cols - set(dl_rank_df.columns)))
        raise RuntimeError(f"DL順位データに必要列がありません: {missing}")

    value_cols = [c for c in ["dl_rank", "dl_prob"] if c in dl_rank_df.columns]
    if not value_cols:
        return merged

    work = dl_rank_df[["rid_str", "馬番", *value_cols]].copy()
    work["rid_str"] = _normalize_rid_series(work["rid_str"])
    work["馬番"] = _normalize_umaban_series(work["馬番"])
    if "dl_rank" in work.columns:
        work["dl_rank"] = pd.to_numeric(work["dl_rank"], errors="coerce")
    if "dl_prob" in work.columns:
        work["dl_prob"] = pd.to_numeric(work["dl_prob"], errors="coerce")
    work = work.drop_duplicates(subset=["rid_str", "馬番"], keep="last")

    out = merged.copy()
    out["rid_str"] = _normalize_rid_series(out["rid_str"])
    out["馬番"] = _normalize_umaban_series(out["馬番"])
    out = out.drop(columns=[c for c in value_cols if c in out.columns], errors="ignore")
    out = pd.merge(out, work, on=["rid_str", "馬番"], how="left")
    print(f"[INFO] DL順位データをメモリ上で結合しました: {len(work)}行")
    return out


# ================================================================
# メイン処理
# ================================================================
def run_pipeline(
    SRC_EXCEL: str,
    OUT_EXCEL: str,
    LEVELS_XL: str = str(RACE_LEVEL_XLSX),
    BASE_TIME: str = str(BASE_TIME_XLSX),
    ODDS_CSV_PATH: str = str(ODDS_CSV),
    RACEDAY: str | None = None,
    DL_RANK_DF: Optional[pd.DataFrame] = None,
) -> None:
    # 各種マスタ読み込み
    levels_df = load_race_levels(LEVELS_XL)
    base_time_df = load_base_time(BASE_TIME)
    odds_df = load_odds_csv(ODDS_CSV_PATH, raceday=RACEDAY)

    # 特徴量構築
    merged, feat_df = build_features_from_excel(
        SRC_EXCEL,
        levels_df,
        base_time_df,
        odds_df,
        raceday=RACEDAY,
    )

    # ============================================================
    # 今回追加:
    # レース登録馬の過去走情報が足りないレースを予想対象から除外
    # ============================================================
    merged, feat_df = _exclude_races_with_missing_history(merged, feat_df)
    merged = _merge_dl_rank_override(merged, DL_RANK_DF)

    # 除外後に空になった場合
    if merged.empty or feat_df.empty:
        print("[WARN] 過去走不足レース除外後、予想対象がありませんでした")
        out_df = merged.copy()

        # score系の列が無いと後続や出力で困るので、念のため空列を作る
        for c in ["score", "rank", "favorite_risk", "extra_penalty", "rest_dist_risk", "dl_rank_score"]:
            if c not in out_df.columns:
                out_df[c] = pd.NA

        write_features_to_excel(
            src_excel=SRC_EXCEL,
            out_excel=OUT_EXCEL,
            feat_df=feat_df,
            now_df=out_df,
            odds_df=odds_df,
        )
        append_success_report(out_df, str(SUCCESS_REPORT))
        return

    # 場所・馬場を rid_str ごとに取得
    place_map: Dict[str, str] = {}
    surface_map: Dict[str, str] = {}

    if "場所" in merged.columns:
        place_map = merged.groupby("rid_str")["場所"].first().to_dict()
    if "芝・ダ" in merged.columns:
        surface_map = merged.groupby("rid_str")["芝・ダ"].first().to_dict()
    elif "芝ダ" in merged.columns:
        surface_map = merged.groupby("rid_str")["芝ダ"].first().to_dict()

    # 距離マップ
    dist_map = {}
    for col in ["距離", "距離(m)", "距離 ", "Distance"]:
        if col in merged.columns:
            dist_map = (
                merged.groupby("rid_str")[col]
                .first()
                .apply(
                    lambda v: _to_int(re.search(r"(\d+)", str(v)).group(1))
                    if pd.notna(v) and re.search(r"(\d+)", str(v))
                    else None
                )
                .to_dict()
            )
            break

    # 頭数マップ
    field_size_map = {}
    for col in ["頭数", "頭 数", "field_size"]:
        if col in merged.columns:
            field_size_map = merged.groupby("rid_str")[col].first().apply(lambda v: _to_int(v)).to_dict()
            break

    # 馬場マップ
    baba_map = {}
    for col in ["馬場状態", "馬場", "馬 場"]:
        if col in merged.columns:
            baba_map = merged.groupby("rid_str")[col].first().to_dict()
            break

    # 人気マップ（rid_str, 馬番 -> 人気）
    pop_map = {}
    pop_col = None
    for col in merged.columns:
        if "人気" in str(col):
            pop_col = col
            break
    if pop_col:
        pop_series = pd.to_numeric(merged[pop_col], errors="coerce")
        pop_map = {
            (str(rid), _to_int(uma)): _to_int(pop)
            for rid, uma, pop in zip(merged.get("rid_str"), merged.get("馬番"), pop_series)
            if _to_int(uma) is not None and _to_int(pop) is not None
        }

    calc_fav_risk = build_calc_favorite_risk(place_map, surface_map, dist_map, field_size_map, pop_map, baba_map)

    # dl_rank を feat_df に付与（rid_str+馬番で結合）
    dl_join = merged[["rid_str", "馬番"]].copy()
    dl_join["dl_rank"] = merged["dl_rank"] if "dl_rank" in merged.columns else pd.NA
    dl_join["dl_prob"] = merged["dl_prob"] if "dl_prob" in merged.columns else pd.NA
    dl_join["頭数"] = merged["頭数"] if "頭数" in merged.columns else pd.NA
    dl_join["rid_str"] = _normalize_rid_series(dl_join["rid_str"])
    dl_join["馬番"] = _normalize_umaban_series(dl_join["馬番"])
    dl_join["dl_rank"] = pd.to_numeric(dl_join["dl_rank"], errors="coerce")
    dl_join["dl_prob"] = pd.to_numeric(dl_join["dl_prob"], errors="coerce")
    dl_join["頭数"] = pd.to_numeric(dl_join["頭数"], errors="coerce")

    feat_df["rid_str"] = _normalize_rid_series(feat_df["rid_str"])
    feat_df["馬番"] = _normalize_umaban_series(feat_df["馬番"])
    feat_df = pd.merge(feat_df, dl_join, on=["rid_str", "馬番"], how="left")

    # dl_rank_score の計算（事故防止の例外ルール付き）
    def _calc_dl_rank_score(row: pd.Series) -> float:
        r = row.get("dl_rank")
        n = row.get("頭数")
        if pd.isna(r) or pd.isna(n):
            return 0.5
        try:
            r_f = float(r)
            n_f = float(n)
        except Exception:
            return 0.5
        if n_f <= 1 or r_f < 1 or r_f > n_f:
            return 0.5
        return (n_f - r_f) / (n_f - 1.0)

    feat_df["dl_rank_score"] = feat_df.apply(_calc_dl_rank_score, axis=1)

    # dl_prob をレース内 0-1 に正規化して、dl_rank_score と混合する。
    # 確率差の情報を残しつつ、順位情報も少し残す。
    feat_df["dl_prob_score"] = pd.to_numeric(feat_df.get("dl_prob"), errors="coerce")

    def _normalize_prob_within_race(s: pd.Series) -> pd.Series:
        x = pd.to_numeric(s, errors="coerce")
        if x.notna().sum() == 0:
            return pd.Series([0.5] * len(s), index=s.index, dtype=float)
        mn = x.min(skipna=True)
        mx = x.max(skipna=True)
        if pd.isna(mn) or pd.isna(mx) or mx == mn:
            return pd.Series([0.5] * len(s), index=s.index, dtype=float)
        return ((x - mn) / (mx - mn)).fillna(0.5)

    feat_df["dl_prob_score"] = feat_df.groupby("rid_str")["dl_prob_score"].transform(_normalize_prob_within_race)
    feat_df["dl_score"] = (
        pd.to_numeric(feat_df["dl_prob_score"], errors="coerce").fillna(0.5) * DL_PROB_BLEND
        + pd.to_numeric(feat_df["dl_rank_score"], errors="coerce").fillna(0.5) * DL_RANK_BLEND
    )
    feat_df = feat_df.drop(columns=["頭数"], errors="ignore")

    feat_df = compute_scores_with_pipeline_logic(
        feat_df,
        place_map=place_map,
        surface_map=surface_map,
        calc_fav_risk=calc_fav_risk,
    )

    # 既に同名列が merged 側に入っている場合（過去の with_feat を入力にした等）
    # pandas merge が suffix（_x/_y）を付ける際に列名が衝突して MergeError になることがあります。
    # ここでは「今回あらためて計算した列」で上書きする前提で、古い同名列を削除してから結合します。
    _cols_to_add = ["score", "rank", "favorite_risk", "extra_penalty", "rest_dist_risk", "dl_rank_score"]
    _cols_to_drop = []
    for _c in _cols_to_add:
        _cols_to_drop.extend([_c, f"{_c}_x", f"{_c}_y"])
    merged = merged.drop(columns=[c for c in _cols_to_drop if c in merged.columns], errors="ignore")

    # 今走情報へ結合（rest_dist_risk も出力する）
    out_df = pd.merge(
        merged,
        feat_df[["rid_str", "馬番", "score", "rank", "favorite_risk", "extra_penalty", "rest_dist_risk", "dl_rank_score"]],
        on=["rid_str", "馬番"],
        how="left",
    )

    # Excel 出力
    write_features_to_excel(
        src_excel=SRC_EXCEL,
        out_excel=OUT_EXCEL,
        feat_df=feat_df,
        now_df=out_df,
        odds_df=odds_df,  # ★追加：買い目シートの単勝オッズ反映に使う
    )

    # 集計
    append_success_report(out_df, str(SUCCESS_REPORT))
