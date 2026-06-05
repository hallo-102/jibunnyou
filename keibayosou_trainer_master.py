# -*- coding: utf-8 -*-
"""
keibayosou_trainer_master.py

目的:
- keibayosou_training_v9_self_trainer_pattern.py などで作成した
  「調教明細CSV」または「調教明細シート」から、調教師ごとの調教パターンDBを作成・更新する。
- DB保存先は data/master を想定。
- 既存の調教スコアExcelに「調教スコア_DB反映」シートを追加し、
  蓄積DBから見た調教師パターンスコアを反映する。

想定コマンド例:
python -u ".\\keibayosou_trainer_master.py" ^
  --score-xlsx ".\\data\\output\\馬の競走成績_20260531_training_v9.xlsx" ^
  --raw-csv ".\\data\\output\\training_raw_20260531_v9.csv" ^
  --master-dir ".\\data\\master" ^
  --out-xlsx ".\\data\\output\\馬の競走成績_20260531_training_v9_trainerDB.xlsx"
"""

from __future__ import annotations

import argparse
import os
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 基本設定
# ============================================================
SCORE_SHEET = "調教スコア"
RAW_SHEET = "調教明細"
OUT_SCORE_DB_SHEET = "調教スコア_DB反映"

MASTER_XLSX_NAME = "trainer_training_pattern_master.xlsx"
MASTER_HISTORY_CSV_NAME = "trainer_training_history_master.csv"

GOOD_FINISH_MAX = 3
BAD_FINISH_MIN = 6
MIN_TRAINER_SAMPLE_ROWS = 10

TRAINER_MASTER_WEIGHT = 0.20


# ============================================================
# 小さな共通関数
# ============================================================
def _safe_str(x: Any) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def _safe_float(x: Any, default: Optional[float] = np.nan) -> float:
    try:
        if x is None:
            return default
        if pd.isna(x):
            return default
        s = str(x).replace(",", "").strip()
        if s == "" or s.lower() == "nan":
            return default
        m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
        if not m:
            return default
        return float(m.group(0))
    except Exception:
        return default


def _safe_int(x: Any, default: Optional[int] = None) -> Optional[int]:
    try:
        v = _safe_float(x, default=np.nan)
        if pd.isna(v):
            return default
        return int(v)
    except Exception:
        return default


def _normalize_trainer_name(x: Any) -> str:
    s = _safe_str(x)
    s = s.replace("　", " ")
    s = re.sub(r"\s+", "", s)
    return s


def _make_trainer_key(row: pd.Series) -> str:
    code = _safe_str(row.get("調教師コード"))
    name = _normalize_trainer_name(row.get("調教師"))
    if code and code.lower() != "nan":
        # Excel由来の .0 対策
        code2 = re.sub(r"\.0$", "", code)
        return f"code:{code2}"
    if name:
        return f"name:{name}"
    return "unknown"


def _mode_value(s: pd.Series) -> str:
    vals = s.dropna().astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return ""
    return vals.value_counts().index[0]


def _mean_num(s: pd.Series) -> float:
    return pd.to_numeric(s, errors="coerce").dropna().mean()


def _count_num(s: pd.Series) -> int:
    return int(pd.to_numeric(s, errors="coerce").notna().sum())


def _clip(v: float, low: float = 0.0, high: float = 100.0) -> float:
    try:
        if pd.isna(v):
            return 50.0
        return float(max(low, min(high, v)))
    except Exception:
        return 50.0


def _judge_from_score(score: float) -> str:
    score = _safe_float(score, 50.0)
    if score >= 80:
        return "調教師DBパターン合致"
    if score >= 65:
        return "やや合致"
    if score >= 45:
        return "普通"
    if score >= 30:
        return "やや不一致"
    return "不一致気味"


def _final_judge(score: float) -> str:
    score = _safe_float(score, 50.0)
    if score >= 85:
        return "かなり良い"
    if score >= 75:
        return "良い"
    if score >= 60:
        return "普通"
    if score >= 45:
        return "やや不安"
    return "不安"


def _excel_autofit(path: Path, sheet_names: Optional[list[str]] = None) -> None:
    try:
        wb = load_workbook(path)
        targets = sheet_names or wb.sheetnames
        for ws_name in targets:
            if ws_name not in wb.sheetnames:
                continue
            ws = wb[ws_name]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")
            for col_cells in ws.columns:
                col_letter = get_column_letter(col_cells[0].column)
                max_len = 8
                for cell in col_cells[:200]:
                    try:
                        val = "" if cell.value is None else str(cell.value)
                        max_len = max(max_len, min(len(val) + 2, 40))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_len
        wb.save(path)
    except Exception as e:
        print(f"[WARN] Excel整形をスキップしました: {e}")


# ============================================================
# 読み込み
# ============================================================
def load_score_and_raw(score_xlsx: Path, raw_csv: Optional[Path] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    if not score_xlsx.exists():
        raise FileNotFoundError(f"score-xlsx が見つかりません: {score_xlsx}")

    xls = pd.ExcelFile(score_xlsx, engine="openpyxl")
    if SCORE_SHEET not in xls.sheet_names:
        raise RuntimeError(f"{SCORE_SHEET} シートが見つかりません: {score_xlsx}")

    score_df = pd.read_excel(score_xlsx, sheet_name=SCORE_SHEET, engine="openpyxl")

    if raw_csv is not None and raw_csv.exists():
        raw_df = pd.read_csv(raw_csv, encoding="utf-8-sig", low_memory=False)
    else:
        if RAW_SHEET not in xls.sheet_names:
            raise RuntimeError(f"raw-csv が無く、{RAW_SHEET} シートも見つかりません")
        raw_df = pd.read_excel(score_xlsx, sheet_name=RAW_SHEET, engine="openpyxl")

    return score_df, raw_df


def prepare_raw_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    df = raw_df.copy()

    for col in ["time_1f", "time_2f", "time_3f", "time_4f", "time_5f", "time_6f", "race_finish_order"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "trainer_key" not in df.columns:
        df["trainer_key"] = df.apply(_make_trainer_key, axis=1)
    else:
        # 空だけ補完
        empty = df["trainer_key"].astype(str).str.strip().isin(["", "nan", "None"])
        if empty.any():
            df.loc[empty, "trainer_key"] = df.loc[empty].apply(_make_trainer_key, axis=1)

    if "調教師" not in df.columns:
        df["調教師"] = ""
    if "調教師コード" not in df.columns:
        df["調教師コード"] = ""

    df["is_good_run"] = pd.to_numeric(df.get("race_finish_order"), errors="coerce").between(1, GOOD_FINISH_MAX).astype(int)
    df["is_bad_run"] = (pd.to_numeric(df.get("race_finish_order"), errors="coerce") >= BAD_FINISH_MIN).astype(int)

    # 重複排除用キー
    key_cols = [c for c in ["horse_id", "race_id", "training_date", "course", "time_6f", "time_5f", "time_4f", "time_3f", "time_2f", "time_1f", "trainer_key"] if c in df.columns]
    if key_cols:
        df = df.drop_duplicates(subset=key_cols, keep="last").copy()
    else:
        df = df.drop_duplicates().copy()

    return df


# ============================================================
# マスター更新
# ============================================================
def update_history_master(raw_df: pd.DataFrame, master_dir: Path) -> pd.DataFrame:
    master_dir.mkdir(parents=True, exist_ok=True)
    history_path = master_dir / MASTER_HISTORY_CSV_NAME

    new_df = prepare_raw_df(raw_df)
    new_df["master_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if history_path.exists():
        old_df = pd.read_csv(history_path, encoding="utf-8-sig", low_memory=False)
        hist = pd.concat([old_df, new_df], ignore_index=True, sort=False)
    else:
        hist = new_df.copy()

    hist = prepare_raw_df(hist)
    hist["master_updated_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    hist.to_csv(history_path, index=False, encoding="utf-8-sig")
    print(f"[INFO] 調教師履歴CSVを保存しました: {history_path} rows={len(hist):,}")
    return hist


# ============================================================
# パターン集計
# ============================================================
def build_trainer_summary(hist: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for trainer_key, g in hist.groupby("trainer_key", dropna=False):
        good = g[g["is_good_run"] == 1].copy()
        bad = g[g["is_bad_run"] == 1].copy()
        rows.append({
            "trainer_key": trainer_key,
            "調教師": _mode_value(g.get("調教師", pd.Series(dtype=object))),
            "調教師コード": _mode_value(g.get("調教師コード", pd.Series(dtype=object))),
            "sample_training_rows": len(g),
            "good_training_rows": len(good),
            "bad_training_rows": len(bad),
            "good_course_main": _mode_value(good.get("course", pd.Series(dtype=object))),
            "bad_course_main": _mode_value(bad.get("course", pd.Series(dtype=object))),
            "good_avg_1f": _mean_num(good.get("time_1f", pd.Series(dtype=float))),
            "good_avg_4f": _mean_num(good.get("time_4f", pd.Series(dtype=float))),
            "good_avg_5f": _mean_num(good.get("time_5f", pd.Series(dtype=float))),
            "bad_avg_1f": _mean_num(bad.get("time_1f", pd.Series(dtype=float))),
            "bad_avg_4f": _mean_num(bad.get("time_4f", pd.Series(dtype=float))),
            "bad_avg_5f": _mean_num(bad.get("time_5f", pd.Series(dtype=float))),
            "good_rate_by_rows": len(good) / len(g) if len(g) else np.nan,
            "avg_1f_gap_good_minus_bad": _mean_num(good.get("time_1f", pd.Series(dtype=float))) - _mean_num(bad.get("time_1f", pd.Series(dtype=float))),
            "avg_4f_gap_good_minus_bad": _mean_num(good.get("time_4f", pd.Series(dtype=float))) - _mean_num(bad.get("time_4f", pd.Series(dtype=float))),
            "avg_5f_gap_good_minus_bad": _mean_num(good.get("time_5f", pd.Series(dtype=float))) - _mean_num(bad.get("time_5f", pd.Series(dtype=float))),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["good_training_rows", "sample_training_rows"], ascending=False, kind="mergesort")
    return out


def build_trainer_course_pattern(hist: pd.DataFrame) -> pd.DataFrame:
    if "course" not in hist.columns:
        return pd.DataFrame()
    rows = []
    for (trainer_key, course), g in hist.groupby(["trainer_key", "course"], dropna=False):
        good = g[g["is_good_run"] == 1]
        bad = g[g["is_bad_run"] == 1]
        rows.append({
            "trainer_key": trainer_key,
            "調教師": _mode_value(g.get("調教師", pd.Series(dtype=object))),
            "調教師コード": _mode_value(g.get("調教師コード", pd.Series(dtype=object))),
            "course": course,
            "sample_training_rows": len(g),
            "good_training_rows": len(good),
            "bad_training_rows": len(bad),
            "good_rate_by_rows": len(good) / len(g) if len(g) else np.nan,
            "good_avg_1f": _mean_num(good.get("time_1f", pd.Series(dtype=float))),
            "good_avg_4f": _mean_num(good.get("time_4f", pd.Series(dtype=float))),
            "good_avg_5f": _mean_num(good.get("time_5f", pd.Series(dtype=float))),
            "bad_avg_1f": _mean_num(bad.get("time_1f", pd.Series(dtype=float))),
            "bad_avg_4f": _mean_num(bad.get("time_4f", pd.Series(dtype=float))),
            "bad_avg_5f": _mean_num(bad.get("time_5f", pd.Series(dtype=float))),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["trainer_key", "good_training_rows", "sample_training_rows"], ascending=[True, False, False], kind="mergesort")
    return out


def build_trainer_eval_pattern(hist: pd.DataFrame) -> pd.DataFrame:
    if "eval_grade" not in hist.columns:
        return pd.DataFrame()
    rows = []
    for (trainer_key, eval_grade), g in hist.groupby(["trainer_key", "eval_grade"], dropna=False):
        good = g[g["is_good_run"] == 1]
        bad = g[g["is_bad_run"] == 1]
        rows.append({
            "trainer_key": trainer_key,
            "調教師": _mode_value(g.get("調教師", pd.Series(dtype=object))),
            "調教師コード": _mode_value(g.get("調教師コード", pd.Series(dtype=object))),
            "eval_grade": eval_grade,
            "sample_training_rows": len(g),
            "good_training_rows": len(good),
            "bad_training_rows": len(bad),
            "good_rate_by_rows": len(good) / len(g) if len(g) else np.nan,
            "good_avg_1f": _mean_num(good.get("time_1f", pd.Series(dtype=float))),
            "good_avg_4f": _mean_num(good.get("time_4f", pd.Series(dtype=float))),
            "good_avg_5f": _mean_num(good.get("time_5f", pd.Series(dtype=float))),
        })
    out = pd.DataFrame(rows)
    if not out.empty:
        out = out.sort_values(["trainer_key", "good_training_rows", "sample_training_rows"], ascending=[True, False, False], kind="mergesort")
    return out


# ============================================================
# DB反映スコア
# ============================================================
def _score_trainer_master(row: pd.Series, summary_map: dict[str, dict[str, Any]]) -> Tuple[float, str]:
    trainer_key = _safe_str(row.get("trainer_key"))
    m = summary_map.get(trainer_key)
    if not m:
        return 50.0, "調教師DBなし"

    sample_n = _safe_int(m.get("sample_training_rows"), 0) or 0
    good_n = _safe_int(m.get("good_training_rows"), 0) or 0
    if sample_n < MIN_TRAINER_SAMPLE_ROWS or good_n <= 0:
        return 50.0, f"調教師DBサンプル不足 sample={sample_n} good={good_n}"

    score = 50.0
    reasons = []

    current_course = _safe_str(row.get("current_course_main")) or _safe_str(row.get("latest_course"))
    good_course = _safe_str(m.get("good_course_main"))
    if current_course and good_course:
        if current_course == good_course:
            score += 12
            reasons.append(f"主コース一致({current_course})")
        else:
            score -= 4
            reasons.append(f"主コース不一致(今回{current_course}/DB好走{good_course})")

    current_1f = _safe_float(row.get("current_avg_1f"), np.nan)
    current_4f = _safe_float(row.get("current_avg_4f"), np.nan)
    current_5f = _safe_float(row.get("current_avg_5f"), np.nan)

    good_1f = _safe_float(m.get("good_avg_1f"), np.nan)
    good_4f = _safe_float(m.get("good_avg_4f"), np.nan)
    good_5f = _safe_float(m.get("good_avg_5f"), np.nan)

    def _diff_score(label: str, cur: float, good: float, strong: float, mild: float):
        nonlocal score
        if pd.isna(cur) or pd.isna(good):
            return
        diff = cur - good  # マイナスが良い
        if diff <= -strong:
            score += 8
            reasons.append(f"{label}がDB好走平均より速い({diff:.2f})")
        elif diff <= mild:
            score += 4
            reasons.append(f"{label}がDB好走平均水準({diff:.2f})")
        elif diff >= strong:
            score -= 8
            reasons.append(f"{label}がDB好走平均より遅い({diff:.2f})")
        else:
            score -= 2
            reasons.append(f"{label}がやや物足りない({diff:.2f})")

    _diff_score("1F", current_1f, good_1f, strong=0.4, mild=0.2)
    _diff_score("4F", current_4f, good_4f, strong=1.5, mild=0.8)
    _diff_score("5F", current_5f, good_5f, strong=2.0, mild=1.0)

    # DB信頼度を少しだけ反映
    if sample_n >= 100 and good_n >= 30:
        score += 4
        reasons.append("DBサンプル十分")
    elif sample_n >= 50 and good_n >= 15:
        score += 2
        reasons.append("DBサンプル中程度")

    return _clip(score), " / ".join(reasons) if reasons else "調教師DB比較材料少"


def apply_trainer_master_to_score(score_df: pd.DataFrame, trainer_summary: pd.DataFrame) -> pd.DataFrame:
    out = score_df.copy()

    if "trainer_key" not in out.columns:
        out["trainer_key"] = out.apply(_make_trainer_key, axis=1)

    summary_map = {
        _safe_str(r["trainer_key"]): r.to_dict()
        for _, r in trainer_summary.iterrows()
        if _safe_str(r.get("trainer_key"))
    }

    scores = []
    reasons = []
    for _, row in out.iterrows():
        s, reason = _score_trainer_master(row, summary_map)
        scores.append(s)
        reasons.append(reason)

    out["trainer_master_score"] = scores
    out["trainer_master_judge"] = out["trainer_master_score"].map(_judge_from_score)
    out["trainer_master_reason"] = reasons

    # DB側の代表値を横持ちで付与
    summary_cols = [
        "trainer_key", "good_course_main", "good_avg_1f", "good_avg_4f", "good_avg_5f",
        "bad_course_main", "bad_avg_1f", "bad_avg_4f", "bad_avg_5f",
        "sample_training_rows", "good_training_rows", "bad_training_rows",
    ]
    add = trainer_summary[[c for c in summary_cols if c in trainer_summary.columns]].copy()
    add = add.rename(columns={
        "good_course_main": "trainer_master_good_course_main",
        "good_avg_1f": "trainer_master_good_avg_1f",
        "good_avg_4f": "trainer_master_good_avg_4f",
        "good_avg_5f": "trainer_master_good_avg_5f",
        "bad_course_main": "trainer_master_bad_course_main",
        "bad_avg_1f": "trainer_master_bad_avg_1f",
        "bad_avg_4f": "trainer_master_bad_avg_4f",
        "bad_avg_5f": "trainer_master_bad_avg_5f",
        "sample_training_rows": "trainer_master_sample_training_rows",
        "good_training_rows": "trainer_master_good_training_rows",
        "bad_training_rows": "trainer_master_bad_training_rows",
    })
    out = out.merge(add, on="trainer_key", how="left")

    out["trainer_master_course_match"] = (
        out.get("current_course_main", pd.Series([""] * len(out))).astype(str).str.strip()
        == out.get("trainer_master_good_course_main", pd.Series([""] * len(out))).astype(str).str.strip()
    ).astype(int)

    base = pd.to_numeric(out.get("training_score_final", out.get("training_score", 50)), errors="coerce").fillna(50.0)
    master = pd.to_numeric(out["trainer_master_score"], errors="coerce").fillna(50.0)
    out["training_score_final_with_master"] = (base + (master - 50.0) * TRAINER_MASTER_WEIGHT).clip(0, 100).round(2)
    out["training_judge_final_with_master"] = out["training_score_final_with_master"].map(_final_judge)

    base_reason = out.get("training_reason_final", out.get("training_reason", "")).fillna("").astype(str)
    out["training_reason_final_with_master"] = (
        base_reason
        + " / 調教師DB=" + out["trainer_master_judge"].astype(str)
        + "(" + out["trainer_master_reason"].astype(str) + ")"
    )

    return out


# ============================================================
# 保存
# ============================================================
def save_master_xlsx(master_dir: Path, hist: pd.DataFrame, trainer_summary: pd.DataFrame, course_pattern: pd.DataFrame, eval_pattern: pd.DataFrame) -> Path:
    master_dir.mkdir(parents=True, exist_ok=True)
    path = master_dir / MASTER_XLSX_NAME

    metadata = pd.DataFrame([
        {"key": "updated_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"key": "history_rows", "value": len(hist)},
        {"key": "trainer_count", "value": trainer_summary["trainer_key"].nunique() if not trainer_summary.empty else 0},
        {"key": "good_finish_max", "value": GOOD_FINISH_MAX},
        {"key": "bad_finish_min", "value": BAD_FINISH_MIN},
        {"key": "trainer_master_weight", "value": TRAINER_MASTER_WEIGHT},
    ])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        metadata.to_excel(writer, sheet_name="metadata", index=False)
        trainer_summary.to_excel(writer, sheet_name="trainer_summary", index=False)
        course_pattern.to_excel(writer, sheet_name="trainer_course_pattern", index=False)
        eval_pattern.to_excel(writer, sheet_name="trainer_eval_pattern", index=False)
        # Excel肥大化対策で履歴は最大100,000行まで
        hist_tail = hist.tail(100000).copy()
        hist_tail.to_excel(writer, sheet_name="trainer_training_history", index=False)

    _excel_autofit(path)
    print(f"[INFO] 調教師パターンExcelを保存しました: {path}")
    return path


def save_output_score_xlsx(score_xlsx: Path, out_xlsx: Path, db_score_df: pd.DataFrame) -> Path:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    if score_xlsx.resolve() != out_xlsx.resolve():
        shutil.copy2(score_xlsx, out_xlsx)

    mode = "a" if out_xlsx.exists() else "w"
    if mode == "a":
        with pd.ExcelWriter(out_xlsx, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            db_score_df.to_excel(writer, sheet_name=OUT_SCORE_DB_SHEET, index=False)
    else:
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            db_score_df.to_excel(writer, sheet_name=OUT_SCORE_DB_SHEET, index=False)

    _excel_autofit(out_xlsx, sheet_names=[OUT_SCORE_DB_SHEET])
    print(f"[INFO] DB反映済みExcelを保存しました: {out_xlsx} / sheet={OUT_SCORE_DB_SHEET}")
    return out_xlsx


# ============================================================
# メイン
# ============================================================
def main() -> None:
    parser = argparse.ArgumentParser(description="調教師の調教パターンDBを作成・更新し、調教スコアへ反映します。")
    parser.add_argument("--score-xlsx", required=True, help="v9等で作成した調教スコアExcel")
    parser.add_argument("--raw-csv", default="", help="v9等で作成した調教明細CSV。省略時はscore-xlsxの調教明細シートを読む")
    parser.add_argument("--master-dir", default=r".\data\master", help="調教師DBを保存するフォルダ")
    parser.add_argument("--out-xlsx", default="", help="DB反映済みExcelの出力先。省略時は score-xlsx の末尾に _trainerDB を付ける")
    args = parser.parse_args()

    score_xlsx = Path(args.score_xlsx)
    raw_csv = Path(args.raw_csv) if args.raw_csv else None
    master_dir = Path(args.master_dir)

    if args.out_xlsx:
        out_xlsx = Path(args.out_xlsx)
    else:
        out_xlsx = score_xlsx.with_name(score_xlsx.stem + "_trainerDB" + score_xlsx.suffix)

    print(f"[INFO] score_xlsx: {score_xlsx}")
    print(f"[INFO] raw_csv: {raw_csv if raw_csv else '(Excelの調教明細シートを使用)'}")
    print(f"[INFO] master_dir: {master_dir}")

    score_df, raw_df = load_score_and_raw(score_xlsx, raw_csv)
    print(f"[INFO] 調教スコア読み込み: {len(score_df):,} rows")
    print(f"[INFO] 調教明細読み込み: {len(raw_df):,} rows")

    hist = update_history_master(raw_df, master_dir)
    trainer_summary = build_trainer_summary(hist)
    course_pattern = build_trainer_course_pattern(hist)
    eval_pattern = build_trainer_eval_pattern(hist)

    save_master_xlsx(master_dir, hist, trainer_summary, course_pattern, eval_pattern)

    db_score_df = apply_trainer_master_to_score(score_df, trainer_summary)
    save_output_score_xlsx(score_xlsx, out_xlsx, db_score_df)

    print("[INFO] 完了")
    print(f"[INFO] 履歴件数: {len(hist):,}")
    print(f"[INFO] 調教師件数: {trainer_summary['trainer_key'].nunique() if not trainer_summary.empty else 0:,}")
    print(f"[INFO] 出力Excel: {out_xlsx}")


if __name__ == "__main__":
    main()
