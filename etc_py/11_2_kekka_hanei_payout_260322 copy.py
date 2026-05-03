# -*- coding: utf-8 -*-
"""
目的：
  racedata_results.xlsx（シート=開催日YYYYMMDD）を元にして、
  「馬の競走成績_with_feat_YYYYMMDD_with_dl.xlsx」の TARGET シートへ

  1) Ag列に「結果_着順」を埋める
  2) Ah列以降に「払戻情報」を埋める（レース単位の払戻を同一 rid_str の全行へ同じ値で入れる）

  ★追加済み機能
  3) 「買い目_レース別1行」シートの
     - AA列：L〜Q列（1位〜6位馬番）の中に、実際の1〜3着の馬番が“全部”含まれていれば「〇」
     - AB列：そのレースの3連複の払戻金（円）を入れる（不的中は 0）
     - AC列〜AE列：結果の1〜3着の馬番を入れる（AC=1着, AD=2着, AE=3着）
     - AF列：L列の「ランキング1位の馬番」が実際に3着以内なら、その馬の「複勝払戻金」を入れる
              （3着以内でなければ 0、結果が無ければ空欄）

入出力（同じフォルダ想定）：
  入力(結果) : racedata_results.xlsx
  入力(TARGET): 馬の競走成績_with_feat_{RACE_DATE}.xlsx
  出力       : 入力ファイル名の末尾に _with_result を付与

使い方（例）：
  python 11_1_kekka_hanei_payout_260112.py

注意：
  - このスクリプトは「RACE_DATE のシート名」にある結果を使います
  - 3連複払戻は、結果側の「払戻種別/組番/払戻金」から拾います
  - 複勝払戻も、結果側の「払戻種別=複勝」「組番」「払戻金」から拾います
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


# ================================================================
# 設定
# ================================================================
RACE_DATE = "20260503"         # 対象日（シート名）※必要に応じて変更してください.
TARGET_SHEET_NAME = "TARGET"   # 反映先シート名（TARGET）

# パス設定（keiba_yosou_2026 を基準）
BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), os.pardir))
DATA_DIR = os.path.join(BASE_DIR, "data")
MASTER_DIR = os.path.join(DATA_DIR, "master")
OUTPUT_DIR = os.path.join(DATA_DIR, "output")

# 着順を書き込む列
WRITE_RANK_COL_LETTER = "Ag"

# 払戻を書き込む開始列
WRITE_PAYOUT_START_COL_LETTER = "Ah"

# 払戻列の並び（左から順に）
PAYOUT_KINDS_ORDER = ["単勝", "複勝", "枠連", "馬連", "ワイド", "馬単", "3連複", "3連単"]


# ================================================================
# 小ユーティリティ
# ================================================================
def _script_dir() -> str:
    return os.path.dirname(os.path.abspath(__file__))


def _norm_text(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _norm_colname(c: Any) -> str:
    s = _norm_text(c)
    s = s.replace(" ", "")
    s = s.lower()
    return s


def _to_int(x: Any) -> Optional[int]:
    if x is None:
        return None
    try:
        if isinstance(x, int):
            return int(x)
        if isinstance(x, float):
            if pd.isna(x):
                return None
            return int(x)
        s = str(x)
        s = re.sub(r"[^\d\-]", "", s)
        if s == "" or s == "-":
            return None
        return int(s)
    except Exception:
        return None


def _rid_to_str(x: Any) -> str:
    if x is None:
        return ""
    s = str(x).strip()
    s = re.sub(r"\D", "", s)
    return s


def _norm_name(name: Any) -> str:
    s = _norm_text(name)
    s = s.replace(" ", "")
    return s


def _normalize_combo(combo: Any) -> str:
    """
    組番（例: "1-2-3", "1 2 3", "1,2,3"）を "1-2-3" 形式に揃える
    単勝・複勝のような 1頭だけの組番なら "5" のように返る
    """
    if combo is None:
        return ""
    s = str(combo).strip()
    if not s:
        return ""
    nums = re.findall(r"\d+", s)
    if not nums:
        return ""
    arr = sorted(int(n) for n in nums)
    return "-".join(str(n) for n in arr)


def _excel_col_letter_to_index(col_letter: str) -> int:
    """
    "A"->1, "Z"->26, "AA"->27 ...
    """
    col_letter = col_letter.upper().strip()
    n = 0
    for ch in col_letter:
        if not ("A" <= ch <= "Z"):
            continue
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def _excel_index_to_col_letter(n: int) -> str:
    """
    1->"A", 26->"Z", 27->"AA" ...
    """
    if n <= 0:
        return ""
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(ord("A") + r) + s
    return s


# ================================================================
# 買い目_レース別1行 シート用ユーティリティ
# ================================================================
def _parse_payout_text(payout_text: Any) -> Dict[str, int]:
    """
    payout_text 例:
      "1-2-3=12340 / 2-3-5=6780"
      "5=150 / 8=120 / 11=180"
    を {"1-2-3": 12340, ...} や {"5": 150, ...} にする。
    """
    out: Dict[str, int] = {}
    if payout_text is None:
        return out
    s = str(payout_text).strip()
    if not s:
        return out

    parts = [p.strip() for p in s.split("/") if p.strip()]
    for p in parts:
        if "=" not in p:
            continue
        combo, yen = p.split("=", 1)
        combo_norm = _normalize_combo(combo)
        yen_int = _to_int(yen)
        if combo_norm and yen_int is not None:
            out[combo_norm] = int(yen_int)
    return out


def _build_rid_to_top3_umaban(map_rid_umaban_to_rank: Dict[Tuple[str, int], int]) -> Dict[str, List[int]]:
    """
    (rid, 馬番) -> 着順 のマップから、rid -> [1着馬番,2着馬番,3着馬番] を作る
    """
    tmp: Dict[str, Dict[int, int]] = {}
    for (rid, umaban), rank in map_rid_umaban_to_rank.items():
        if rid is None:
            continue
        try:
            r = int(rank)
        except Exception:
            continue
        if r not in (1, 2, 3):
            continue
        if umaban is None:
            continue
        rid = str(rid)
        tmp.setdefault(rid, {})[r] = int(umaban)

    out: Dict[str, List[int]] = {}
    for rid, d in tmp.items():
        if all(k in d for k in (1, 2, 3)):
            out[rid] = [d[1], d[2], d[3]]
    return out


def _update_buysheet_hit_and_payout(
    wb,
    rid_to_top3: Dict[str, List[int]],
    payout_map: Dict[str, Dict[str, str]],
    sheet_name: str = "買い目_レース別1行",
) -> Tuple[int, int]:
    """
    買い目_レース別1行 シートに
      - AA列: 予想馬番(L-Q列=1位〜6位) に 1〜3着が全て入っていれば「〇」
      - AB列: 的中した場合、3連複の払戻金（円）
      - AC列〜AE列: 結果の1〜3着の馬番（AC=1着, AD=2着, AE=3着）
      - AF列: L列の「ランキング1位の馬番」が3着以内なら、その馬の複勝払戻金

    を書き込む。

    戻り値: (hit_count, miss_count)
    """
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] '{sheet_name}' シートが見つからないためスキップします（既存: {wb.sheetnames}）")
        return 0, 0

    ws = wb[sheet_name]

    # 列位置
    RID_COL = 1           # A列: レースID
    PRED_START_COL = 12   # L列: 1位馬番
    PRED_END_COL = 17     # Q列: 6位馬番

    HIT_COL = 27          # AA列
    PAY_COL = 28          # AB列（3連複払戻）

    RES_START_COL = 29    # AC列: 1着馬番
    RES_COLS = 3          # AC=1着, AD=2着, AE=3着

    FUKUSHO_TOP1_COL = 32 # AF列: ランキング1位馬番の複勝払戻

    # ヘッダ
    if ws.cell(row=1, column=HIT_COL).value in (None, ""):
        ws.cell(row=1, column=HIT_COL).value = "的中"
    if ws.cell(row=1, column=PAY_COL).value in (None, ""):
        ws.cell(row=1, column=PAY_COL).value = "3連複払戻"

    res_headers = ["1着馬番", "2着馬番", "3着馬番"]
    for i in range(RES_COLS):
        c = RES_START_COL + i
        if ws.cell(row=1, column=c).value in (None, ""):
            ws.cell(row=1, column=c).value = res_headers[i]

    if ws.cell(row=1, column=FUKUSHO_TOP1_COL).value in (None, ""):
        ws.cell(row=1, column=FUKUSHO_TOP1_COL).value = "1位馬番_複勝払戻"

    hit_count = 0
    miss_count = 0

    for row in range(2, ws.max_row + 1):
        rid = _rid_to_str(ws.cell(row=row, column=RID_COL).value)
        if not rid:
            continue

        top3 = rid_to_top3.get(rid)
        if not top3 or len(top3) < 3:
            ws.cell(row=row, column=HIT_COL).value = ""
            ws.cell(row=row, column=PAY_COL).value = ""
            for i in range(RES_COLS):
                ws.cell(row=row, column=RES_START_COL + i).value = ""
            ws.cell(row=row, column=FUKUSHO_TOP1_COL).value = ""
            continue

        # 結果馬番（1〜3着）を書き込む
        for i in range(RES_COLS):
            ws.cell(row=row, column=RES_START_COL + i).value = int(top3[i])

        # 予想馬番（L-Q）
        pred_nums: List[int] = []
        for c in range(PRED_START_COL, PRED_END_COL + 1):
            v = _to_int(ws.cell(row=row, column=c).value)
            if v is not None:
                pred_nums.append(int(v))
        pred_set = set(pred_nums)

        # L列のランキング1位馬番
        rank1_umaban = _to_int(ws.cell(row=row, column=PRED_START_COL).value)

        # 的中判定（1〜3着が全部 L-Q に含まれるか）
        ok = all(int(n) in pred_set for n in top3)

        if ok:
            ws.cell(row=row, column=HIT_COL).value = "〇"
            hit_count += 1

            pm = payout_map.get(rid, {})
            tri_text = pm.get("3連複") if pm else None
            tri_dict = _parse_payout_text(tri_text)

            combo = _normalize_combo("-".join(str(n) for n in top3))
            pay = tri_dict.get(combo, 0)

            ws.cell(row=row, column=PAY_COL).value = int(pay) if pay else 0
        else:
            ws.cell(row=row, column=HIT_COL).value = ""
            ws.cell(row=row, column=PAY_COL).value = 0
            miss_count += 1

        # ★追加：ランキング1位馬番の複勝払戻（AF列）
        # 条件：L列の馬番が、実際に1〜3着のどれかに入っていること
        if rank1_umaban is not None and int(rank1_umaban) in set(top3):
            pm = payout_map.get(rid, {})
            fuku_text = pm.get("複勝") if pm else None
            fuku_dict = _parse_payout_text(fuku_text)
            fuku_key = _normalize_combo(str(rank1_umaban))
            fuku_pay = fuku_dict.get(fuku_key, 0)
            ws.cell(row=row, column=FUKUSHO_TOP1_COL).value = int(fuku_pay) if fuku_pay else 0
        else:
            ws.cell(row=row, column=FUKUSHO_TOP1_COL).value = 0

    return hit_count, miss_count


# ================================================================
# 結果Excelから「着順」「払戻」を作る
# ================================================================
def _load_results_maps(results_xlsx: str, sheet_name: str):
    """
    戻り値：
      map_rid_umaban_to_rank : Dict[(rid_str, umaban_int)] = rank_int
      map_rid_name_to_rank   : Dict[(rid_str, name_norm)] = rank_int
      payout_map             : Dict[rid_str][kind] = "組番=払戻金 / 組番=払戻金 ..."
    """
    if not os.path.exists(results_xlsx):
        raise FileNotFoundError(f"結果ファイルが見つかりません: {results_xlsx}")

    df = pd.read_excel(results_xlsx, sheet_name=sheet_name, engine="openpyxl")
    if df is None or df.empty:
        raise ValueError(f"結果シートが空です: {results_xlsx} sheet={sheet_name}")

    col_map = {_norm_colname(c): c for c in df.columns}

    # 必須（レースID / 着順 / 馬名）
    rid_col = None
    for key in ["レースid", "raceid", "race_id"]:
        if key in col_map:
            rid_col = col_map[key]
            break
    if rid_col is None:
        for k_norm, orig in col_map.items():
            if "レースid" in k_norm or "raceid" in k_norm:
                rid_col = orig
                break
    if rid_col is None:
        raise ValueError("結果側に レースID 列が見つかりません")

    rank_col = None
    for key in ["着順", "着順num", "順位"]:
        if key in col_map:
            rank_col = col_map[key]
            break
    if rank_col is None:
        for k_norm, orig in col_map.items():
            if "着順" in k_norm or "順位" in k_norm:
                rank_col = orig
                break
    if rank_col is None:
        raise ValueError("結果側に 着順 列が見つかりません")

    name_col = None
    for key in ["馬名"]:
        if key in col_map:
            name_col = col_map[key]
            break
    if name_col is None:
        for k_norm, orig in col_map.items():
            if "馬名" in k_norm:
                name_col = orig
                break
    if name_col is None:
        raise ValueError("結果側に 馬名 列が見つかりません")

    # 任意（馬番）
    umaban_col = None
    for key in ["馬番", "馬番int", "馬番num", "馬番番号", "馬番 "]:
        if key in col_map:
            umaban_col = col_map[key]
            break
    if umaban_col is None:
        for k_norm, orig in col_map.items():
            if "馬番" in k_norm:
                umaban_col = orig
                break

    # 払戻列
    pay_type_col = None
    for key in ["払戻種別", "券種", "券種別"]:
        if key in col_map:
            pay_type_col = col_map[key]
            break
    if pay_type_col is None:
        for k_norm, orig in col_map.items():
            if "払戻種別" in k_norm or "券種" in k_norm:
                pay_type_col = orig
                break

    combo_col = None
    for key in ["組番", "組番号"]:
        if key in col_map:
            combo_col = col_map[key]
            break
    if combo_col is None:
        for k_norm, orig in col_map.items():
            if k_norm == "組番":
                combo_col = orig
                break

    payout_col = None
    for key in ["払戻金", "払戻金額", "払戻金払戻金"]:
        if key in col_map:
            payout_col = col_map[key]
            break
    if payout_col is None:
        for k_norm, orig in col_map.items():
            if "払戻" in k_norm and ("金" in k_norm or "額" in k_norm):
                payout_col = orig
                break

    # --- 着順マップ作成 ---
    map_rid_umaban_to_rank: Dict[Tuple[str, int], int] = {}
    map_rid_name_to_rank: Dict[Tuple[str, str], int] = {}

    for _, r in df.iterrows():
        rid = _rid_to_str(r.get(rid_col))
        if not rid:
            continue

        rank = _to_int(r.get(rank_col))
        if rank is None:
            continue

        if umaban_col is not None:
            umaban = _to_int(r.get(umaban_col))
            if umaban is not None:
                map_rid_umaban_to_rank[(rid, int(umaban))] = int(rank)

        nm = _norm_name(r.get(name_col))
        if nm:
            map_rid_name_to_rank[(rid, nm)] = int(rank)

    # --- 払戻マップ作成 ---
    payout_map: Dict[str, Dict[str, str]] = {}
    if pay_type_col is not None and combo_col is not None and payout_col is not None:
        tmp = df[[rid_col, pay_type_col, combo_col, payout_col]].copy()
        tmp.columns = ["レースID", "払戻種別", "組番", "払戻金"]
        tmp["rid_str"] = tmp["レースID"].map(_rid_to_str)
        tmp["kind"] = tmp["払戻種別"].map(_norm_text)
        tmp["combo_norm"] = tmp["組番"].map(_normalize_combo)
        tmp["yen_int"] = tmp["払戻金"].map(_to_int)

        tmp = tmp.dropna(subset=["rid_str", "kind"])
        tmp = tmp[tmp["rid_str"].astype(str) != ""]
        tmp = tmp[~tmp["combo_norm"].astype(str).eq("")]

        for rid, g in tmp.groupby("rid_str"):
            payout_map.setdefault(rid, {})
            for kind, g2 in g.groupby("kind"):
                rows = []
                for _, rr in g2.iterrows():
                    cn = rr["combo_norm"]
                    y = rr["yen_int"]
                    if cn and (y is not None):
                        rows.append(f"{cn}={int(y)}")
                if rows:
                    payout_map[rid][kind] = " / ".join(rows)

    return map_rid_umaban_to_rank, map_rid_name_to_rank, payout_map


def _find_header_cols(ws):
    """
    1行目のヘッダ行を読み取り、列名→列番号 を返す
    """
    header_cols: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        kn = _norm_colname(v)
        if kn:
            header_cols[kn] = c
    return header_cols


def _ensure_header(ws, header_cols: Dict[str, int], wanted_header: str, fixed_col_letter: str) -> int:
    """
    header_cols に無い場合でも、固定列へヘッダを書き込んで列番号を返す。
    """
    wanted_norm = _norm_colname(wanted_header)
    if wanted_norm in header_cols:
        return header_cols[wanted_norm]

    col_idx = _excel_col_letter_to_index(fixed_col_letter)
    ws.cell(row=1, column=col_idx).value = wanted_header
    header_cols[wanted_norm] = col_idx
    return col_idx


def _choose_target_xlsx(base_dir: str, race_date: str) -> str:
    cand1 = os.path.join(base_dir, f"馬の競走成績_with_feat_{race_date}.xlsx")
    cand2 = os.path.join(base_dir, f"馬の競走成績_with_feat_{race_date}_with_dl.xlsx")

    if os.path.exists(cand1):
        return cand1
    if os.path.exists(cand2):
        return cand2

    d0 = _script_dir()
    cand3 = os.path.join(d0, f"馬の競走成績_with_feat_{race_date}_with_dl.xlsx")
    cand4 = os.path.join(d0, f"馬の競走成績_with_feat_{race_date}.xlsx")
    if os.path.exists(cand3):
        return cand3
    if os.path.exists(cand4):
        return cand4

    raise FileNotFoundError(
        f"TARGET Excel が見つかりません。\n"
        f"探した: {cand1} / {cand2} / {cand3} / {cand4}"
    )


def _make_out_path(target_xlsx: str) -> str:
    p = Path(target_xlsx)
    stem = p.stem
    out = p.with_name(stem + "_with_result" + p.suffix)
    return str(out)


def main():
    RESULTS_XLSX = os.path.join(MASTER_DIR, "racedata_results.xlsx")
    TARGET_XLSX = _choose_target_xlsx(OUTPUT_DIR, RACE_DATE)
    OUT_XLSX = _make_out_path(TARGET_XLSX)

    print("[INFO] 入力(結果):", RESULTS_XLSX, "sheet=", RACE_DATE)
    print("[INFO] 入力(TARGET):", TARGET_XLSX, "sheet=", TARGET_SHEET_NAME)
    print("[INFO] 出力:", OUT_XLSX)

    # 1) 結果からマップ作成
    map_rid_umaban_to_rank, map_rid_name_to_rank, payout_map = _load_results_maps(RESULTS_XLSX, RACE_DATE)

    # 2) Excel読み込み
    wb = load_workbook(TARGET_XLSX)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"TARGETシートが見つかりません: {TARGET_SHEET_NAME}  / 既存: {wb.sheetnames}")

    ws = wb[TARGET_SHEET_NAME]

    # 3) ヘッダ列を検出
    header_cols = _find_header_cols(ws)

    # rid列（必須）
    rid_col = None
    for k in ["rid_str", "レースID", "raceid", "rid"]:
        kn = _norm_colname(k)
        if kn in header_cols:
            rid_col = header_cols[kn]
            break
    if rid_col is None:
        raise ValueError("TARGETシートに rid_str（またはレースID）列が見つかりません")

    # 馬番列（あれば優先）
    umaban_col = None
    for k in ["馬番", "馬 番", "umaban"]:
        kn = _norm_colname(k)
        if kn in header_cols:
            umaban_col = header_cols[kn]
            break

    # 馬名列（馬番が無い時の保険）
    name_col = None
    for k in ["馬名", "馬 名", "horse_name"]:
        kn = _norm_colname(k)
        if kn in header_cols:
            name_col = header_cols[kn]
            break

    # 4) 結果_着順 列（Ag）
    rank_write_col = _ensure_header(ws, header_cols, "結果_着順", WRITE_RANK_COL_LETTER)

    # 5) 払戻列（Ah から順番）
    start_col = _excel_col_letter_to_index(WRITE_PAYOUT_START_COL_LETTER)
    payout_cols: Dict[str, int] = {}
    for i, kind in enumerate(PAYOUT_KINDS_ORDER):
        col_idx = start_col + i
        payout_cols[kind] = col_idx
        ws.cell(row=1, column=col_idx).value = kind

    # 6) 本体：1行ずつ、着順と払戻を書き込む
    hit = 0
    miss = 0
    payout_written_rows = 0

    for row in range(2, ws.max_row + 1):
        rid = _rid_to_str(ws.cell(row=row, column=rid_col).value)
        if not rid:
            continue

        # --- 着順 ---
        rank = None

        if umaban_col is not None:
            umaban = _to_int(ws.cell(row=row, column=umaban_col).value)
            if umaban is not None:
                rank = map_rid_umaban_to_rank.get((rid, umaban))

        if rank is None and name_col is not None:
            nm = _norm_name(ws.cell(row=row, column=name_col).value)
            if nm:
                rank = map_rid_name_to_rank.get((rid, nm))

        if rank is None:
            miss += 1
        else:
            ws.cell(row=row, column=rank_write_col).value = int(rank)
            hit += 1

        # --- 払戻（レース単位）---
        pm = payout_map.get(rid)
        wrote_any = False
        if pm:
            for kind in PAYOUT_KINDS_ORDER:
                val = pm.get(kind)
                if val is None:
                    continue
                ws.cell(row=row, column=payout_cols[kind]).value = val
                wrote_any = True

        if wrote_any:
            payout_written_rows += 1

    # 6.5) 買い目_レース別1行 シートへ反映
    rid_to_top3 = _build_rid_to_top3_umaban(map_rid_umaban_to_rank)
    hit_buy, miss_buy = _update_buysheet_hit_and_payout(
        wb, rid_to_top3, payout_map, sheet_name="買い目_レース別1行"
    )

    # 7) 保存
    wb.save(OUT_XLSX)

    # 8) ログ
    print("[OK] 書き込み完了")
    print(f"  RACE_DATE: {RACE_DATE}")
    print(f"  入力(結果): {RESULTS_XLSX}  sheet={RACE_DATE}")
    print(f"  入力(TARGET): {TARGET_XLSX}  sheet={TARGET_SHEET_NAME}")
    print(f"  出力: {OUT_XLSX}")
    print(f"  着順ヒット件数: {hit}")
    print(f"  着順見つからず(空欄): {miss}")
    print(f"  払戻を書いた行数（同一rid_str含む）: {payout_written_rows}")
    print(f"  買い目_レース別1行 的中(〇)件数: {hit_buy} / 不的中件数: {miss_buy}")
    print("  ※ miss が多い場合：rid_str / 馬番（または馬名）が結果側と一致しているか確認してください")
    print("  ※ 払戻が空の多い場合：結果側に「払戻種別/組番/払戻金」列があるか確認してください")
    print("  ※ AF列には『L列のランキング1位馬番が3着以内なら複勝払戻、それ以外は0』を入れます")


if __name__ == "__main__":
    main()