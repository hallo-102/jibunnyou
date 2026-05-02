# etc_py/fill_bet_odds_to_output.py
import os
import re
import unicodedata
import pandas as pd
from openpyxl import load_workbook

OUT_EXCEL = r"C:\Users\okino\OneDrive\ドキュメント\my_python_cursor\keiba_yosou_2026\data\output\馬の競走成績_with_feat_20260207.xlsx"
ODDS_DIR  = r"C:\Users\okino\OneDrive\ドキュメント\my_python_cursor\keiba_yosou_2026\data\ozzu_csv"
RACEDAY   = "20260207"  # 必要なら変更
SHEET     = "買い目_レース別1行"
ODDS_COL  = "単勝オッズ_1位"

def normalize_place(x):
    if not isinstance(x, str):
        return ""
    s = unicodedata.normalize("NFKC", x)
    return s.replace("競馬場", "").strip()

def race_no_from_rid(rid):
    m = re.search(r"(\d{2})$", str(rid))
    return m.group(1) if m else ""

def to_int(x):
    m = re.search(r"(\d+)", str(x))
    return int(m.group(1)) if m else None

def pick_ozzu_csv(odds_dir: str, raceday: str | None):
    csvs = [os.path.join(odds_dir, f) for f in os.listdir(odds_dir) if f.lower().endswith(".csv")]
    if not csvs:
        raise FileNotFoundError(f"CSVがありません: {odds_dir}")

    if raceday:
        hit = [p for p in csvs if raceday in os.path.basename(p)]
        if hit:
            # 同日に複数あれば更新日時が新しいもの
            return sorted(hit, key=lambda p: os.path.getmtime(p), reverse=True)[0]

    # 日付指定が無い/見つからないなら最新
    return sorted(csvs, key=lambda p: os.path.getmtime(p), reverse=True)[0]

ozzu_path = pick_ozzu_csv(ODDS_DIR, RACEDAY)
print("[INFO] use ozzu csv:", ozzu_path)

ozzu = pd.read_csv(ozzu_path, encoding="utf-8-sig")
need = {"racecourse", "race", "bet_type", "combination", "odds"}
if not need.issubset(set(ozzu.columns)):
    raise RuntimeError(f"OZZU形式の列が足りません。必要={need} 実際={set(ozzu.columns)}")

ozzu = ozzu[ozzu["bet_type"].astype(str).str.contains("単勝", na=False)].copy()
ozzu["place_norm"] = ozzu["racecourse"].map(normalize_place)
ozzu["race_no"] = ozzu["race"].astype(str).str.extract(r"(\d+)")[0].fillna("").apply(lambda v: str(v).zfill(2))
ozzu["umaban1"] = ozzu["combination"].map(to_int)
ozzu["odds_num"] = pd.to_numeric(ozzu["odds"].astype(str).str.replace(",", ""), errors="coerce")
ozzu = ozzu.dropna(subset=["place_norm", "race_no", "umaban1", "odds_num"])

odds_map = {(p, r, int(u)): float(o) for p, r, u, o in zip(ozzu["place_norm"], ozzu["race_no"], ozzu["umaban1"], ozzu["odds_num"])}

wb = load_workbook(OUT_EXCEL)
if SHEET not in wb.sheetnames:
    raise RuntimeError(f"{OUT_EXCEL} に {SHEET} シートがありません。シート一覧={wb.sheetnames}")

ws = wb[SHEET]

# ヘッダ位置
header = {}
for c in range(1, ws.max_column + 1):
    v = ws.cell(row=1, column=c).value
    if v is None:
        continue
    header[str(v).strip()] = c

for k in ["場所", "レースID", "1位馬番"]:
    if k not in header:
        raise RuntimeError(f"{SHEET} に必要列がありません: {k} / 実際の列={list(header.keys())}")

# 単勝オッズ列が無ければ作る（保険）
if ODDS_COL not in header:
    new_c = ws.max_column + 1
    ws.cell(row=1, column=new_c).value = ODDS_COL
    header[ODDS_COL] = new_c
    print(f"[INFO] {ODDS_COL} 列を新規作成しました（列={new_c}）")

col_place = header["場所"]
col_rid   = header["レースID"]
col_u1    = header["1位馬番"]
col_odds  = header[ODDS_COL]

filled = 0
total = 0

for r in range(2, ws.max_row + 1):
    rid = ws.cell(r, col_rid).value
    if rid is None:
        continue

    place = ws.cell(r, col_place).value
    u1 = ws.cell(r, col_u1).value
    u1i = to_int(u1)

    total += 1
    if u1i is None:
        continue

    key = (normalize_place(place), race_no_from_rid(rid), int(u1i))
    odds = odds_map.get(key)
    if odds is not None:
        ws.cell(r, col_odds).value = float(odds)
        filled += 1

wb.save(OUT_EXCEL)
print(f"[OK] {ODDS_COL} を反映しました: {filled}/{total} -> {OUT_EXCEL}")
