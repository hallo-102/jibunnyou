# -*- coding: utf-8 -*-
# 01_jizen_syuusyuu_race_info_classfix_20251212.py（完全版）
# -----------------------------------------------------------------------------
# 目的：
#   1) 指定日の race_list から race_id を取得
#   2) 各 race_id の出馬表(shutuba)から
#        「今走レース情報」シートを作る（1レース×登録馬数 行）
#        A〜I：レース情報（既存のまま）
#        J：馬番、K：馬名、L：年齢、M：斤量、N：騎手、O：厩舎
#   3) 各馬の競走成績（db.netkeiba）を race_id シートに出力（既存のまま）
#
# ★今回の修正点（ユーザー要望）：
#   - J列: 馬番、K列: 馬名 が取れている前提で
#   - 追加で L列: 年齢、M列: 斤量、N列: 騎手、O列: 厩舎 も追加
#   - 性齢は「牡2、牡4、牝6、セ5」など → 年齢は数字だけ取り出す
#   - 取得は scrape_shutuba_to_excel.py を参考に pandas.read_html() を使用（取りこぼし防止）
# -----------------------------------------------------------------------------

from __future__ import annotations

import os
import re
import io
import sys
import time
import configparser
from urllib.parse import urlparse, parse_qs
from datetime import datetime

import pandas as pd
import requests
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import TimeoutException, SessionNotCreatedException
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook


# ============================================================
# 設定（★ここを変えるだけで日付指定できます）
# ============================================================

# ★固定指定（ユーザー要望により復活）
raceday: str = "20260503"  # YYYYMMDD                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                                "  # YYYYMMDD

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/120.0.0.0 Safari/537.36"
)

HTTP_TIMEOUT = 20
SELENIUM_TIMEOUT = 40

# ★credentials.ini の場所（ユーザー指定）
CREDENTIALS_INI = r"C:\Users\okino\OneDrive\ドキュメント\my_python_cursor\keiba_yosou_2026\config\credentials.ini"

# ★出力先（ユーザー指定）
BASE_XLSX_DIR = r"C:\Users\okino\OneDrive\ドキュメント\my_python_cursor\keiba_yosou_2026\data\input"


# ============================================================
# ユーティリティ
# ============================================================

def _nfkc(s: str | None) -> str:
    if s is None:
        return ""
    try:
        import unicodedata
        return unicodedata.normalize("NFKC", str(s))
    except Exception:
        return str(s)


def load_credentials(ini_path: str = CREDENTIALS_INI) -> dict:
    """
    credentials.ini からログイン情報を読む
    想定:
      [netkeiba]
      username=....
      password=....
    """
    if not os.path.exists(ini_path):
        raise FileNotFoundError(f"credentials.ini が見つかりません: {ini_path}")

    config = configparser.ConfigParser()
    config.read(ini_path, encoding="utf-8")

    if "netkeiba" not in config:
        raise KeyError("credentials.ini に [netkeiba] セクションがありません")

    username = config["netkeiba"].get("username", "").strip()
    password = config["netkeiba"].get("password", "").strip()

    if not username or not password:
        raise ValueError("credentials.ini の username/password が空です")

    return {"username": username, "password": password}


def fix_headers_ag_to_as(xlsx_path: str) -> None:
    """
    既存互換の保険（READMEシートに最終更新時刻を書く程度）
    """
    try:
        wb = load_workbook(xlsx_path)
        if "README" not in wb.sheetnames:
            wb.create_sheet("README")
        ws = wb["README"]
        ws["A1"].value = "最終更新"
        ws["B1"].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        wb.save(xlsx_path)
    except Exception:
        pass


# ============================================================
# 1) Seleniumログイン → Cookieをrequests.Sessionに移す
# ============================================================

def selenium_login_to_session(sess: requests.Session, cred: dict) -> None:
    print("▶ Selenium による netkeiba ログイン開始")

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-software-rasterizer")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument(f"--user-agent={UA}")

    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

    try:
        driver.get("https://regist.netkeiba.com/account/?pid=login")
        time.sleep(1.5)

        driver.find_element("name", "login_id").send_keys(cred["username"])
        driver.find_element("name", "pswd").send_keys(cred["password"])
        driver.find_element("css selector", 'input[type="image"][alt="ログイン"]').click()

        time.sleep(2.5)
        driver.get("https://www.netkeiba.com/")  # Cookie発行待ち
        time.sleep(1.5)

        for c in driver.get_cookies():
            sess.cookies.set(c["name"], c["value"], domain=c.get("domain", None))

        print("✅ Selenium ログイン成功 (Cookie 移行完了)")
    finally:
        driver.quit()


# ============================================================
# 2) race_list ページから race_id 抽出
# ============================================================

def extract_race_ids(race_list_url: str) -> list[str]:
    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-software-rasterizer")
    opts.add_argument(f"--user-agent={UA}")

    try:
        driver = webdriver.Chrome(options=opts)
    except SessionNotCreatedException:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)

    driver.set_page_load_timeout(SELENIUM_TIMEOUT)

    try:
        driver.get(race_list_url)
    except TimeoutException:
        print(f"⚠️ {SELENIUM_TIMEOUT}s 超でタイムアウト。取得済 HTML 解析続行…", file=sys.stderr)

    soup = BeautifulSoup(driver.page_source, "html.parser")
    driver.quit()

    race_ids = {
        parse_qs(urlparse(a["href"]).query).get("race_id", [""])[0]
        for a in soup.find_all("a", href=True)
        if "race_id=" in a["href"]
    }
    race_ids = [rid for rid in race_ids if rid and rid.isdigit()]
    return sorted(race_ids)


# ============================================================
# 3) 出馬表HTML取得（同じHTMLを使い回す）
# ============================================================

def fetch_shutuba_soup(rid: str, sess: requests.Session) -> tuple[str, BeautifulSoup] | tuple[None, None]:
    url = f"https://race.netkeiba.com/race/shutuba.html?race_id={rid}"
    r = sess.get(url, timeout=HTTP_TIMEOUT, headers={"User-Agent": UA})
    if r.status_code != 200:
        print(f"❌ 出馬表取得失敗 [{r.status_code}] : {url}")
        return None, None
    return url, BeautifulSoup(r.content, "html.parser")


# ============================================================
# 4) 出馬表 → 馬名・詳細 URL（競走成績取得用）
# ============================================================

def get_horse_names_and_urls_from_soup(soup: BeautifulSoup) -> list[dict]:
    """
    競走成績取得用に db.netkeiba の horse URL を拾う（既存の考え方のまま）
    """
    horses = []
    for td in soup.find_all("td", class_="HorseInfo"):
        a = td.find("a")
        if a and a.has_attr("href"):
            horses.append({"馬名": a.get("title", a.text).strip(), "URL": a["href"]})
    return horses


# ============================================================
# 5) 出馬表 → 馬番/馬名/年齢/斤量/騎手/厩舎（★今回の追加）
# ============================================================

def _flatten_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    pd.read_html の列が MultiIndex の場合があるので、先頭要素だけでフラット化する
    """
    cols = []
    for c in df.columns:
        if isinstance(c, tuple):
            c0 = str(c[0])
        else:
            c0 = str(c)
        c0 = c0.replace(" ", "").replace("\u3000", "")
        cols.append(c0)
    df = df.copy()
    df.columns = cols
    return df


def _extract_age_from_seirei(seirei: str | None) -> int | None:
    """
    性齢例：牡2、牡4、牝6、セ5 → 数字だけ取り出して int にする
    """
    if seirei is None:
        return None
    s = str(seirei).strip()
    m = re.search(r"(\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def get_horse_rows_from_shutuba_soup(soup: BeautifulSoup) -> list[dict]:
    """
    出馬表ページから「登録馬の行」を抜く
    返すもの（例）:
      [{"馬番":"1","馬名":"○○","年齢":2,"斤量":"55.0","騎手":"○○","厩舎":"○○"}, ...]
    """
    out: list[dict] = []

    try:
        tables = pd.read_html(io.StringIO(str(soup)))
    except Exception:
        return out
    if not tables:
        return out

    # 出馬表のテーブルを探す（必要列が揃っているもの）
    needed = ["枠", "馬番", "馬名", "性齢", "斤量", "騎手", "厩舎"]

    for raw_df in tables:
        df = _flatten_columns(raw_df)

        if not all(col in df.columns for col in needed):
            continue

        df = df[needed].copy()

        # 行を辞書化
        for _, row in df.iterrows():
            umaban = None if pd.isna(row["馬番"]) else str(row["馬番"]).strip()
            name   = None if pd.isna(row["馬名"]) else str(row["馬名"]).strip()

            seirei = None if pd.isna(row["性齢"]) else str(row["性齢"]).strip()
            age = _extract_age_from_seirei(seirei)

            kinryo = None if pd.isna(row["斤量"]) else str(row["斤量"]).strip()
            jockey = None if pd.isna(row["騎手"]) else str(row["騎手"]).strip()
            stable = None if pd.isna(row["厩舎"]) else str(row["厩舎"]).strip()

            # 馬名に余計な空白が入るケースを軽く整形
            if name:
                name = re.sub(r"\s+", " ", name).strip()

            if umaban or name:
                out.append({
                    "馬番": umaban,
                    "馬名": name,
                    "年齢": age,
                    "斤量": kinryo,
                    "騎手": jockey,
                    "厩舎": stable,
                })

        if out:
            return out

    return out


# ============================================================
# 6) レース情報（RaceData01/02）＋ クラス推定（既存のまま）
# ============================================================

def _pick_race_class(race_name: str, d1: str, d2: str) -> str | None:
    s = _nfkc(" ".join([race_name or "", d1 or "", d2 or ""]))

    # G1/G2/G3
    if re.search(r"\bG1\b", s):
        return "G1"
    if re.search(r"\bG2\b", s):
        return "G2"
    if re.search(r"\bG3\b", s):
        return "G3"

    # 重賞系（例：OP/L/オープン）
    if re.search(r"(オープン|OP|リステッド|L)", s):
        return "オープン"

    # 条件戦（例：3勝クラス/2勝クラス/1勝クラス）
    m = re.search(r"([123])勝クラス", s)
    if m:
        return f"{m.group(1)}勝クラス"

    # 新馬・未勝利
    if "新馬" in s:
        return "新馬"
    if "未勝利" in s:
        return "未勝利"

    return None


def parse_today_race_info_from_soup(rid: str, soup: BeautifulSoup) -> list[dict[str, str | int | None]]:
    """
    今走レース情報シート用に、1レース×登録馬数の行を作る
    A〜I：レース情報（既存）
    J〜O：馬番/馬名/年齢/斤量/騎手/厩舎（今回追加）
    """
    base = soup.find("div", class_="RaceList_Item02")
    if not base:
        return []

    race_name = _nfkc(base.find("h1", class_="RaceName").get_text(strip=True))

    d1 = _nfkc(base.find("div", class_="RaceData01").get_text(" ", strip=True))
    m_time = re.search(r"(\d{1,2}:\d{2})発走", d1)
    m_course = re.search(r"(芝|ダ|ダート|障)[\s ]*(\d+)m", d1)
    m_baba = re.search(r"馬場:([^\s]+)", d1)

    d2 = _nfkc(base.find("div", class_="RaceData02").get_text(" ", strip=True))
    tokens = re.split(r"\s+", d2.strip())

    place = tokens[1] if len(tokens) > 1 else None
    head = next((t for t in tokens if t.endswith("頭")), None)
    heads = int(head.replace("頭", "")) if head else None
    race_type = next((t for t in tokens if re.match(r"(定量|別定|ハンデ|馬齢)", t)), None)

    race_class = _pick_race_class(race_name, d1, d2)
    course = f"{m_course.group(1)}{m_course.group(2)}" if m_course else None

    base_row = {
        "レースID": rid,
        "レース名": race_name,
        "発走時刻": m_time.group(1) if m_time else None,
        "場所": place,
        "コース": course,
        "馬場": m_baba.group(1) if m_baba else None,
        "頭数": heads,
        "レース種別": race_type,
        "クラス": race_class,
    }

    horse_rows = get_horse_rows_from_shutuba_soup(soup)

    rows: list[dict[str, str | int | None]] = []
    if horse_rows:
        for h in horse_rows:
            row = dict(base_row)
            row["馬番"] = h.get("馬番")
            row["馬名"] = h.get("馬名")
            row["年齢"] = h.get("年齢")
            row["斤量"] = h.get("斤量")
            row["騎手"] = h.get("騎手")
            row["厩舎"] = h.get("厩舎")
            rows.append(row)
    else:
        # 取れなかったときの保険（最低1行）
        row = dict(base_row)
        row["馬番"] = None
        row["馬名"] = None
        row["年齢"] = None
        row["斤量"] = None
        row["騎手"] = None
        row["厩舎"] = None
        rows.append(row)

    return rows


# ============================================================
# 7) 馬詳細ページ → 競走成績 DataFrame（既存のまま）
# ============================================================

RACE_META_CACHE: dict[str, dict[str, str | None]] = {}
RACE_LAP_CACHE: dict[str, str | None] = {}


def get_race_lap_times(rid: str | None, sess: requests.Session) -> str | None:
    if not rid:
        return None
    if rid in RACE_LAP_CACHE:
        return RACE_LAP_CACHE[rid]

    url = f"https://db.netkeiba.com/race/{rid}/"
    try:
        r = sess.get(url, timeout=HTTP_TIMEOUT, headers={"User-Agent": UA})
        r.raise_for_status()
    except requests.RequestException:
        RACE_LAP_CACHE[rid] = None
        return None

    soup = BeautifulSoup(r.content, "html.parser")
    td = soup.find("td", class_="race_lap_cell")
    lap = td.get_text(strip=True) if td else None
    RACE_LAP_CACHE[rid] = lap
    return lap


def get_race_basic_info(rid: str, sess: requests.Session) -> dict[str, str | None]:
    if rid in RACE_META_CACHE:
        return RACE_META_CACHE[rid]

    url = f"https://race.netkeiba.com/race/result.html?race_id={rid}"
    try:
        r = sess.get(url, timeout=HTTP_TIMEOUT, headers={"User-Agent": UA})
        r.raise_for_status()
    except requests.RequestException:
        RACE_META_CACHE[rid] = {}
        return {}

    soup = BeautifulSoup(r.content, "html.parser")
    base = soup.find("div", class_="RaceList_Item02")
    if not base:
        RACE_META_CACHE[rid] = {}
        return {}

    race_name = _nfkc(base.find("h1", class_="RaceName").get_text(strip=True))
    d1 = _nfkc(base.find("div", class_="RaceData01").get_text(" ", strip=True))
    d2 = _nfkc(base.find("div", class_="RaceData02").get_text(" ", strip=True))

    meta = {
        "レース名": race_name,
        "RaceData01": d1,
        "RaceData02": d2,
    }
    RACE_META_CACHE[rid] = meta
    return meta


def get_race_results(horse_url: str, sess: requests.Session) -> pd.DataFrame:
    """
    競走成績（db.netkeiba）を DataFrame にする
    """
    def to_abs(url: str) -> str:
        if url.startswith("http"):
            return url
        if url.startswith("/"):
            return "https://db.netkeiba.com" + url
        return url

    url_abs = to_abs(horse_url)

    m = re.search(r"https?://db\.netkeiba\.com/horse/(\d{10})/?", url_abs)
    if m:
        horse_id = m.group(1)
        url_abs = f"https://db.netkeiba.com/horse/result/{horse_id}/"

    try:
        r = sess.get(url_abs, timeout=HTTP_TIMEOUT, headers={"User-Agent": UA})
        r.raise_for_status()
    except requests.RequestException as e:
        print(f"❌ 競走成績取得失敗 : {url_abs} ({e})")
        return pd.DataFrame()

    soup = BeautifulSoup(r.content, "html.parser")

    table = soup.find("table", class_=lambda c: c and "db_h_race_results" in str(c))
    if table is None:
        box = soup.select_one("#horse_results_box")
        if box:
            table = box.find("table", class_=lambda c: c and "db_h_race_results" in str(c))

    if table is None:
        print(f"⚠️ 競走成績テーブル無し : {url_abs}")
        return pd.DataFrame()

    df = pd.read_html(io.StringIO(str(table)))[0]

    race_ids, laps = [], []
    for tr in table.find_all("tr")[1:]:
        a = tr.find("a", href=re.compile(r"/race/\d{12}"))
        rid = None
        if a and a.has_attr("href"):
            m2 = re.search(r"/race/(\d{12})", a["href"])
            if m2:
                rid = m2.group(1)
        race_ids.append(rid)
        laps.append(get_race_lap_times(rid, sess) if rid else None)
        time.sleep(0.15)

    df["race_id"] = race_ids
    df["ラップタイム"] = laps
    return df


# ============================================================
# 8) メイン処理
# ============================================================

def main() -> None:
    # 0) 準備
    cred = load_credentials(CREDENTIALS_INI)

    race_list_url = f"https://race.netkeiba.com/top/race_list.html?kaisai_date={raceday}"

    sess = requests.Session()
    sess.headers.update({"User-Agent": UA})

    # 1) ログイン（Cookie移行）
    selenium_login_to_session(sess, cred)

    # 2) race_id 抽出
    race_ids = extract_race_ids(race_list_url)
    if not race_ids:
        print("❌ race_id 取得失敗")
        return
    print("▶ 取得 race_id:", race_ids)

    # 3) データ収集
    race_id_dfs: dict[str, pd.DataFrame] = {}
    today_info_rows: list[dict] = []

    for rid in race_ids:
        print(f"\n===== race_id {rid} =====")

        shutuba_url, soup = fetch_shutuba_soup(rid, sess)
        if not shutuba_url or soup is None:
            continue

        # 今走レース情報（レース×登録馬数）
        rows = parse_today_race_info_from_soup(rid, soup)
        if rows:
            today_info_rows.extend(rows)

        # 競走成績取得用（馬URL）
        horses = get_horse_names_and_urls_from_soup(soup)
        if not horses:
            continue

        meta_info = get_race_basic_info(rid, sess)

        horse_dfs = []
        for h in horses:
            print(f"  ▶ {h['馬名']} 成績取得中…")
            df = get_race_results(h["URL"], sess)
            if df.empty:
                continue
            df.insert(0, "馬名", h["馬名"])
            horse_dfs.append(df)
            time.sleep(0.5)

        if not horse_dfs:
            continue

        combined = pd.concat(horse_dfs, ignore_index=True)

        # 追加情報（レース名・RaceData）をラップ列の右に挿入（既存互換）
        if "ラップタイム" in combined.columns:
            lap_pos = combined.columns.get_loc("ラップタイム")
        else:
            lap_pos = 0

        # meta_info は最低限：レース名 / RaceData01/02
        add_cols = ["レース名", "RaceData01", "RaceData02"]
        for col in reversed(add_cols):
            if col in combined.columns:
                continue
            combined.insert(lap_pos + 1, col, meta_info.get(col))

        race_id_dfs[rid] = combined
        time.sleep(2.0)

    # 4) 出力
    out_xlsx = os.path.join(BASE_XLSX_DIR, f"馬の競走成績_{raceday}.xlsx")

    # ★今走レース情報：列順（J〜Oを追加）
    cols_today = [
        "レースID", "レース名", "発走時刻", "場所", "コース", "馬場",
        "頭数", "レース種別", "クラス",
        "馬番", "馬名", "年齢", "斤量", "騎手", "厩舎",
    ]

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # 各 race_id シート（競走成績）
        for rid in sorted(race_id_dfs.keys()):
            race_id_dfs[rid].to_excel(writer, sheet_name=rid, index=False)

        # 今走レース情報（1レース×登録馬数）
        df_today = pd.DataFrame(today_info_rows)
        for c in cols_today:
            if c not in df_today.columns:
                df_today[c] = None
        df_today = df_today[cols_today]
        df_today.to_excel(writer, sheet_name="今走レース情報", index=False)

    fix_headers_ag_to_as(out_xlsx)

    print(f"\n✅ 完了: {out_xlsx} を保存しました")


if __name__ == "__main__":
    main()
