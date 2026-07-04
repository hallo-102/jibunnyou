# =========================
# 1_keibayosou_pipeline.py
# =========================
# penalties（新規ファイル）を呼び出す形に整理した完全版。
# 旧：pipeline内に _calc_extra_penalty / _calc_rest_dist_risk を直書き
# 新：1_keibayosou_penalties.py に分離し、ここは「流れ」に集中
#
# 今回の追加修正:
# - レース登録馬の過去走情報が不足しているレースを予想対象から除外
# - 判定基準:
#   レースIDごとに
#     1) 今走レース情報シートの頭数
#     2) feat_df 側で過去走特徴量を作れた馬名ユニーク数
#   を比較し、
#     過去走特徴量を作れた馬名ユニーク数 < 頭数
#   のレースは予想除外する
#
# 例:
#   頭数=16、過去走あり馬数=15 → 初出走馬などがいるとみなし、そのレースは除外
#
# ※ 既存ロジックはなるべくそのまま維持しています。

# -*- coding: utf-8 -*-
"""パイプライン全体の実行フローをまとめたモジュール。"""

from __future__ import annotations

import importlib
import os
import re
import shutil
import sys
import unicodedata
from datetime import datetime
from typing import Dict, Optional, Tuple


def _register_renamed_keibayosou_modules() -> None:
    """1_ 始まりへリネームした自作モジュールを、旧import名でも参照できるようにする。"""
    module_aliases = [
        ("keibayosou_config", "1_keibayosou_config"),
        ("keibayosou_utils", "1_keibayosou_utils"),
        ("keibayosou_course_style", "1_keibayosou_course_style"),
        ("keibayosou_loaders", "1_keibayosou_loaders"),
        ("keibayosou_features", "1_keibayosou_features"),
        ("keibayosou_penalties", "1_keibayosou_penalties"),
    ]
    for old_name, new_name in module_aliases:
        if old_name not in sys.modules:
            sys.modules[old_name] = importlib.import_module(new_name)


_register_renamed_keibayosou_modules()

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from keibayosou_config import (
    TARGET_SHEET,
    NOW_SHEET,
    ALPHA,
    EXTRA_ALPHA,
    DL_PROB_BLEND,
    DL_RANK_BLEND,
    DL_SCORE_BONUS,
    RACE_LEVEL_XLSX,
    BASE_TIME_XLSX,
    ODDS_CSV,
    SUCCESS_REPORT,
    FEAT_COLS,
    JAPANESE_FEATURE_NAMES,
)
from keibayosou_features import (
    _normalize_rid_series,
    _normalize_umaban_series,
    apply_weights,
    build_calc_favorite_risk,
    build_features_from_excel,
    normalize_score,
    score_sum,
)
from keibayosou_loaders import load_base_time, load_odds_csv, load_race_levels
from keibayosou_penalties import calc_extra_penalty, calc_rest_dist_risk
from keibayosou_utils import (
    _build_feature_sheet_for_export,
    _normalize_place,
    _normalize_surface,
    _to_int,
)


def compute_scores_with_pipeline_logic(
    feat_df: pd.DataFrame,
    place_map: Dict[str, str],
    surface_map: Dict[str, str],
    calc_fav_risk,
    alpha: float = ALPHA,
    extra_alpha: float = EXTRA_ALPHA,
) -> pd.DataFrame:
    """pipeline 本番と同じ式で total / score / rank を計算する。"""
    out = feat_df.copy()

    out["total_raw"] = out.apply(
        lambda r: score_sum(
            apply_weights(
                {k: r.get(k) for k in FEAT_COLS},
                place=_normalize_place(place_map.get(str(r.get("rid_str", "")))),
                surface=_normalize_surface(surface_map.get(str(r.get("rid_str", "")))),
            )
        ),
        axis=1,
    )

    out["favorite_risk"] = out.apply(calc_fav_risk, axis=1)
    out["rest_dist_risk"] = out.apply(calc_rest_dist_risk, axis=1)
    out["extra_penalty"] = out.apply(
        lambda r: calc_extra_penalty(r, rest_dist_risk=r.get("rest_dist_risk")),
        axis=1,
    )

    # dl_score は 0.5 を中立点として total に反映する。
    # 1回目は dl 系列が無いので 0.5 扱いとなり、2回目だけ順位へ効く。
    if "dl_score" not in out.columns:
        out["dl_score"] = 0.5

    out["dl_bonus"] = (pd.to_numeric(out["dl_score"], errors="coerce").fillna(0.5) - 0.5) * DL_SCORE_BONUS
    out["total"] = (
        out["total_raw"]
        + out["dl_bonus"]
        - alpha * out["favorite_risk"]
        - extra_alpha * out["extra_penalty"]
    )
    out["score"] = out.groupby("rid_str")["total"].transform(normalize_score).round(2)
    out["rank"] = out.groupby("rid_str")["score"].rank("dense", ascending=False).astype(int)
    return out


# ================================================================
# 今回追加：過去走不足レースを除外するための補助関数
# ================================================================
def _pick_first_existing_col(df: pd.DataFrame, candidates: list[str]) -> Optional[str]:
    """候補の中から、最初に存在する列名を返す。"""
    for col in candidates:
        if col in df.columns:
            return col
    return None


def _exclude_races_with_missing_history(
    merged: pd.DataFrame,
    feat_df: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    レース登録頭数に対して、実際に過去走特徴量を作れた馬数が不足しているレースを除外する。

    判定:
      レースIDごとに
        実際に過去走特徴量あり馬数 < 登録頭数
      なら、そのレースは予想対象外

    ここでいう「実際に過去走特徴量あり馬」とは、
    feat_df に行があるだけではなく、
    ta_n（または同等列）が 1 以上ある馬を指す。
    つまり、行だけ存在して主要特徴量が実質作れていない馬は数えない。

    merged:
      build_features_from_excel の戻り値（今走ベース）
    feat_df:
      build_features_from_excel の戻り値（過去走から特徴量を作れた馬だけ入る想定）

    戻り値:
      (filtered_merged, filtered_feat_df)
    """
    if merged is None or merged.empty:
        return merged, feat_df

    if feat_df is None or feat_df.empty:
        print("[WARN] feat_df が空のため、全レースを予想除外します")
        return merged.iloc[0:0].copy(), feat_df.iloc[0:0].copy()

    work_merged = merged.copy()
    work_feat = feat_df.copy()

    if "rid_str" not in work_merged.columns:
        if "レースID" in work_merged.columns:
            work_merged["rid_str"] = work_merged["レースID"]
        else:
            print("[WARN] merged に rid_str/レースID が無いため、過去走不足レース除外をスキップします")
            return merged, feat_df

    if "rid_str" not in work_feat.columns:
        print("[WARN] feat_df に rid_str が無いため、過去走不足レース除外をスキップします")
        return merged, feat_df

    work_merged["rid_str"] = _normalize_rid_series(work_merged["rid_str"])
    work_feat["rid_str"] = _normalize_rid_series(work_feat["rid_str"])

    # merged 側の馬名列候補
    merged_name_col = _pick_first_existing_col(work_merged, ["馬名", "horse_name", "name"])
    feat_name_col = _pick_first_existing_col(work_feat, ["馬名", "horse_name", "name"])

    if feat_name_col is None:
        print("[WARN] feat_df に馬名列が無いため、過去走不足レース除外をスキップします")
        return merged, feat_df

    # 登録頭数
    field_col = _pick_first_existing_col(work_merged, ["頭数", "頭 数", "field_size"])

    if field_col is not None:
        race_field_df = (
            work_merged[["rid_str", field_col]]
            .copy()
            .assign(**{field_col: pd.to_numeric(work_merged[field_col], errors="coerce")})
            .groupby("rid_str", as_index=False)[field_col]
            .first()
            .rename(columns={field_col: "registered_field_size"})
        )
    else:
        # 念のためのフォールバック
        if merged_name_col is None:
            print("[WARN] merged に頭数列も馬名列も無いため、過去走不足レース除外をスキップします")
            return merged, feat_df

        race_field_df = (
            work_merged[["rid_str", merged_name_col]]
            .dropna(subset=[merged_name_col])
            .copy()
        )
        race_field_df[merged_name_col] = race_field_df[merged_name_col].astype(str).str.strip()
        race_field_df = (
            race_field_df.groupby("rid_str", as_index=False)[merged_name_col]
            .nunique()
            .rename(columns={merged_name_col: "registered_field_size"})
        )

    # 実際に過去走特徴量を作れた馬数
    # 重要:
    # feat_df に行があるだけでは数えず、ta_n（同等列）が 1 以上ある馬だけ数える。
    feat_horse_df = work_feat[["rid_str", feat_name_col]].dropna(subset=[feat_name_col]).copy()
    feat_horse_df[feat_name_col] = feat_horse_df[feat_name_col].astype(str).str.strip()
    feat_horse_df = feat_horse_df[feat_horse_df[feat_name_col] != ""]

    ta_n_col = _pick_first_existing_col(work_feat, ["ta_n", "f_race_count", "レース数"])
    if ta_n_col is not None:
        feat_horse_df[ta_n_col] = pd.to_numeric(work_feat.loc[feat_horse_df.index, ta_n_col], errors="coerce")
        feat_horse_df = feat_horse_df[feat_horse_df[ta_n_col].fillna(0) > 0]
    else:
        # ta_n 相当列が無い場合は、既存挙動に近い保険として全件を数える
        print("[WARN] feat_df に ta_n/f_race_count/レース数 列が無いため、行ベースで過去走あり馬数を数えます")

    history_count_df = (
        feat_horse_df.groupby("rid_str", as_index=False)[feat_name_col]
        .nunique()
        .rename(columns={feat_name_col: "history_horse_count"})
    )

    audit_df = pd.merge(
        race_field_df,
        history_count_df,
        on="rid_str",
        how="left",
    )

    audit_df["history_horse_count"] = pd.to_numeric(audit_df["history_horse_count"], errors="coerce").fillna(0).astype(int)
    audit_df["registered_field_size"] = pd.to_numeric(audit_df["registered_field_size"], errors="coerce")

    exclude_rids = audit_df.loc[
        audit_df["registered_field_size"].notna()
        & (audit_df["history_horse_count"] < audit_df["registered_field_size"]),
        "rid_str",
    ].astype(str).tolist()

    if exclude_rids:
        preview_df = audit_df[audit_df["rid_str"].isin(exclude_rids)].copy()
        preview_df = preview_df.sort_values(["rid_str"], kind="mergesort")

        print(
            f"[INFO] 過去走不足レースを除外します: {len(exclude_rids)}レース "
            f"(全{len(audit_df)}レース中)"
        )
        for _, r in preview_df.iterrows():
            print(
                f"[INFO] 除外 rid={r['rid_str']} "
                f"登録頭数={int(r['registered_field_size']) if pd.notna(r['registered_field_size']) else 'NaN'} "
                f"過去走あり馬数={int(r['history_horse_count'])}"
            )
    else:
        print("[INFO] 過去走不足による除外レースはありません")

    filtered_merged = work_merged.loc[~work_merged["rid_str"].astype(str).isin(exclude_rids)].copy()
    filtered_feat = work_feat.loc[~work_feat["rid_str"].astype(str).isin(exclude_rids)].copy()

    return filtered_merged, filtered_feat



# ================================================================
# 追加で作るシート名（過去の出力と互換）
# ================================================================
BET_SHEET = "買い目_レース別1行"
ROI_FOCUS_BET_SHEET = "回収率重視_買い目候補"
README_SHEET = "README"
BET_RANK_README_START = "【買い目_レース別1行 ランク条件】"
BET_RANK_README_END = "【買い目_レース別1行 ランク条件ここまで】"
S_RANK_SCORE_MIN = 64.0
S_RANK_RATING4_AVG_MIN = 400.0
S_RANK_RECENT3_TIME_IDX_MIN = 80.0
S_RANK_EXTRA_PENALTY_MAX = 1.25
RATING4_AVG_COL_CANDIDATES = ["rating4平均", "rating4_avg", "rating4_mean", "rating4"]
RECENT3_TIME_IDX_COL_CANDIDATES = ["近3走タイム指数", "近3走平均タイム指数", "recent3_time_idx", "recent3_time_index"]
EXTRA_PENALTY_COL_CANDIDATES = ["extra_penalty", "追加ペナルティ", "補正ペナルティ"]

ROI_FOCUS_BET_COLUMNS = [
    "レースID",
    "レース名",
    "発走時刻",
    "場所",
    "コース",
    "馬場",
    "頭数",
    "クラス",
    "score1",
    "score2",
    "gap12",
    "dango_2_5",
    "1位馬番",
    "2位馬番",
    "3位馬番",
    "4位馬番",
    "5位馬番",
    "購入判定",
    "購入理由",
    "3連複1点目_馬番1",
    "3連複1点目_馬番2",
    "3連複1点目_馬番3",
    "3連複2点目_馬番1",
    "3連複2点目_馬番2",
    "3連複2点目_馬番3",
    "3連複3点目_馬番1",
    "3連複3点目_馬番2",
    "3連複3点目_馬番3",
    "3連複_点数",
    "3連複_金額",
    "馬連1点目_馬番1",
    "馬連1点目_馬番2",
    "馬連2点目_馬番1",
    "馬連2点目_馬番2",
    "馬連_点数",
    "馬連_金額",
    "ワイド1点目_馬番1",
    "ワイド1点目_馬番2",
    "ワイド2点目_馬番1",
    "ワイド2点目_馬番2",
    "ワイド_点数",
    "ワイド_金額",
    "合計購入金額",
]


# ================================================================
# 買い目シート作成（過去版の出力互換）
# ================================================================
def _to_int_safe(x: object) -> Optional[int]:
    try:
        if pd.isna(x):
            return None
        if isinstance(x, str) and x.strip() == "":
            return None
        return int(float(x))
    except Exception:
        return None


def _to_float_safe(x: object) -> Optional[float]:
    try:
        if pd.isna(x):
            return None
        if isinstance(x, str) and x.strip() == "":
            return None
        return float(x)
    except Exception:
        return None


def _first_float(*values: object) -> Optional[float]:
    """複数候補から最初に数値化できる値を返す。"""
    for value in values:
        float_value = _to_float_safe(value)
        if float_value is not None:
            return float_value
    return None


def _metric_from_row(row: pd.Series, candidates: list[str]) -> Optional[float]:
    """行データから候補列名の順に数値指標を取り出す。"""
    for col in candidates:
        if col in row.index:
            float_value = _to_float_safe(row.get(col))
            if float_value is not None:
                return float_value
    return None


def _normalize_race_key(x: object) -> Optional[str]:
    """Excel読み戻し時の 123.0 のような表記ゆれをレースID比較用にそろえる。"""
    if pd.isna(x):
        return None
    if isinstance(x, (int, np.integer)):
        return str(int(x))
    if isinstance(x, (float, np.floating)):
        if np.isfinite(x) and float(x).is_integer():
            return str(int(x))
        return str(x).strip()

    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return None
    m = re.fullmatch(r"(\d+)\.0+", s)
    if m:
        return m.group(1)
    return s


def _race_no_to_2digits(x: object) -> str:
    """'11R' や 11 を '11' にそろえる。"""
    m = re.search(r"(\d+)", str(x))
    return m.group(1).zfill(2) if m else ""


def _race_no_from_race_key(x: object) -> str:
    """netkeiba由来のレースID末尾2桁からR番号を取り出す。"""
    race_key = _normalize_race_key(x)
    if race_key is None:
        return ""
    m = re.search(r"(\d{2})$", str(race_key))
    return m.group(1) if m else ""


def _normalize_horse_name_for_odds_key(x: object) -> str:
    """オッズ照合キー用に馬名の全角半角・空白ゆらぎを吸収する。"""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    s = unicodedata.normalize("NFKC", str(x))
    s = re.sub(r"[\s\u3000]+", "", s)
    return s.strip()


def _extract_yyyymmdd_for_odds_key(x: object) -> str:
    """日付列の値から YYYYMMDD を取り出す。"""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    digits = re.sub(r"\D", "", str(x))
    return digits[:8] if len(digits) >= 8 else ""


def _put_unique_odds_value(target: Dict[tuple, float], key: tuple, value: float, context: str) -> None:
    """同一キーが複数ある場合は、上書きせず例外で停止する。"""
    if key in target:
        raise ValueError(f"{context} に同一キーのオッズが複数あります。上書きせず中止します: key={key}")
    target[key] = float(value)


def _lookup_rid_tansho_odds(
    rid_odds_map: Dict[Tuple[str, int, str], float],
    race_key: str,
    umaban: int,
    horse_name_norm: str,
) -> Optional[float]:
    """rid_str・馬番・馬名で単勝オッズを引く。旧形式は空馬名キーを許容する。"""
    keys = [(str(race_key), int(umaban), horse_name_norm)]
    if horse_name_norm:
        keys.append((str(race_key), int(umaban), ""))
    for key in keys:
        if key in rid_odds_map:
            return rid_odds_map[key]
    return None


def _lookup_place_tansho_odds(
    place_odds_map: Dict[Tuple[str, str, str, int, str], float],
    date_key: str,
    place_norm: str,
    race_no: str,
    umaban: int,
    horse_name_norm: str,
) -> Optional[float]:
    """日付・場所・R番号・馬番・馬名で単勝オッズを引く。旧形式は空日付/空馬名キーを許容する。"""
    keys = []
    if date_key:
        keys.append((str(date_key), str(place_norm), str(race_no), int(umaban), horse_name_norm))
        if horse_name_norm:
            keys.append((str(date_key), str(place_norm), str(race_no), int(umaban), ""))
    keys.append(("", str(place_norm), str(race_no), int(umaban), horse_name_norm))
    if horse_name_norm:
        keys.append(("", str(place_norm), str(race_no), int(umaban), ""))
    for key in keys:
        if key in place_odds_map:
            return place_odds_map[key]
    return None


def _build_tansho_odds_maps_for_bet_sheet(
    odds_df: Optional[pd.DataFrame],
) -> Tuple[Dict[Tuple[str, int, str], float], Dict[Tuple[str, str, str, int, str], float]]:
    """
    買い目シート用に単勝オッズ辞書を作る。

    標準CSVは (rid_str, 馬番, 馬名)、JRA OZZU CSVは
    (日付, 場所, R番号, 馬番, 馬名) で参照する。
    """
    rid_odds_map: Dict[Tuple[str, int, str], float] = {}
    place_odds_map: Dict[Tuple[str, str, str, int, str], float] = {}

    if odds_df is None or not isinstance(odds_df, pd.DataFrame) or odds_df.empty:
        return rid_odds_map, place_odds_map

    od = odds_df.copy()
    name_col = None
    for cand in ["name_norm", "馬名_norm", "name", "馬名", "horse_name"]:
        if cand in od.columns:
            name_col = cand
            break

    if {"rid_str", "umaban", "tansho"}.issubset(od.columns):
        for _, row in od.iterrows():
            r = row.get("rid_str")
            u = row.get("umaban")
            t = row.get("tansho")
            race_key = _normalize_race_key(r)
            umaban = _to_int_safe(u)
            tansho = _to_float_safe(t)
            if race_key is None or umaban is None or tansho is None:
                continue
            horse_name_norm = _normalize_horse_name_for_odds_key(row.get(name_col)) if name_col else ""
            key = (str(race_key), int(umaban), horse_name_norm)
            _put_unique_odds_value(rid_odds_map, key, float(tansho), "標準単勝オッズ")

    place_col = "place" if "place" in od.columns else "racecourse" if "racecourse" in od.columns else None
    race_col = "race_no" if "race_no" in od.columns else "race" if "race" in od.columns else None
    date_col = "date" if "date" in od.columns else "日付" if "日付" in od.columns else None
    if place_col is not None and race_col is not None and {"umaban", "tansho"}.issubset(od.columns):
        for _, row in od.iterrows():
            p = row.get(place_col)
            r = row.get(race_col)
            u = row.get("umaban")
            t = row.get("tansho")
            place_norm = _normalize_place(p)
            race_no = _race_no_to_2digits(r)
            umaban = _to_int_safe(u)
            tansho = _to_float_safe(t)
            if not place_norm or not race_no or umaban is None or tansho is None:
                continue
            date_key = _extract_yyyymmdd_for_odds_key(row.get(date_col)) if date_col else ""
            horse_name_norm = _normalize_horse_name_for_odds_key(row.get(name_col)) if name_col else ""
            key = (str(date_key), str(place_norm), str(race_no), int(umaban), horse_name_norm)
            _put_unique_odds_value(place_odds_map, key, float(tansho), "OZZU単勝オッズ")

    return rid_odds_map, place_odds_map


def _to_umaban_int_for_roi(x: object) -> Optional[int]:
    """買い目用の馬番を整数化する。欠損・非整数・0以下は異常値として扱う。"""
    if pd.isna(x):
        return None
    if isinstance(x, str) and x.strip() == "":
        return None

    try:
        v = float(str(x).strip())
    except Exception:
        return None

    if not np.isfinite(v) or not v.is_integer():
        return None

    n = int(v)
    if n <= 0:
        return None
    return n


def _class_map_from_now_df(now_df: Optional[pd.DataFrame]) -> Dict[str, object]:
    """今走シートから レースID -> クラス の対応を作る。"""
    if now_df is None or not isinstance(now_df, pd.DataFrame) or now_df.empty:
        return {}

    now = now_df.copy()
    race_col = _pick_first_existing_col(now, ["レースID", "rid_str", "race_id"])
    class_col = _pick_first_existing_col(now, ["クラス", "レースクラス", "条件クラス"])
    if race_col is None or class_col is None:
        return {}

    now["_race_key_for_roi"] = now[race_col].map(_normalize_race_key)
    now = now.dropna(subset=["_race_key_for_roi"])
    if now.empty:
        return {}

    return now.groupby("_race_key_for_roi")[class_col].first().to_dict()


def _normalize_surface_for_roi(value: object) -> str:
    """回収率重視シート用に、芝/ダだけを有効なsurfaceとして返す。"""
    if pd.isna(value):
        return ""
    surface = _normalize_surface(str(value))
    return surface if surface in {"芝", "ダ"} else ""


def _surface_map_from_now_df(now_df: Optional[pd.DataFrame]) -> Dict[str, str]:
    """今走シートから レースID -> surface の対応を作る。判定不能は空文字にする。"""
    if now_df is None or not isinstance(now_df, pd.DataFrame) or now_df.empty:
        return {}

    now = now_df.copy()
    race_col = _pick_first_existing_col(now, ["レースID", "rid_str", "race_id"])
    surface_col = _pick_first_existing_col(now, ["芝・ダ", "芝ダ", "surface", "Surface"])
    course_col = _pick_first_existing_col(now, ["コース", "course"])
    if race_col is None or (surface_col is None and course_col is None):
        return {}

    now["_race_key_for_roi"] = now[race_col].map(_normalize_race_key)
    if surface_col is not None:
        now["_surface_for_roi"] = now[surface_col].map(_normalize_surface_for_roi)
    else:
        now["_surface_for_roi"] = ""

    if course_col is not None:
        course_surface = now[course_col].map(_normalize_surface_for_roi)
        now["_surface_for_roi"] = now["_surface_for_roi"].where(now["_surface_for_roi"] != "", course_surface)

    now = now.dropna(subset=["_race_key_for_roi"])
    now = now[now["_surface_for_roi"] != ""]
    if now.empty:
        return {}

    return now.groupby("_race_key_for_roi")["_surface_for_roi"].first().to_dict()


def _pick_popularity_col_for_roi(df: pd.DataFrame) -> Optional[str]:
    """人気列の表記ゆれを吸収して、回収率重視シートの条件判定に使う列を返す。"""
    col = _pick_first_existing_col(df, ["人気", "単勝人気", "予想人気", "ninki", "popularity", "pop"])
    if col is not None:
        return col

    for c in df.columns:
        if "人気" in str(c):
            return c
    return None


def _horse_metric_map_from_prediction_df(pred_df: Optional[pd.DataFrame]) -> Dict[Tuple[str, int], Dict[str, float]]:
    """予想系シートから (レースID, 馬番) -> ランク判定用指標 の対応を作る。"""
    if pred_df is None or not isinstance(pred_df, pd.DataFrame) or pred_df.empty:
        return {}

    pred = pred_df.copy()
    race_cols = [c for c in ["レースID", "rid_str", "race_id"] if c in pred.columns]
    umaban_col = _pick_first_existing_col(pred, ["馬番", "umaban", "馬 番", "馬番 "])
    metric_cols = {
        "popularity": _pick_popularity_col_for_roi(pred),
        "score": _pick_first_existing_col(pred, ["score"]),
        "extra_penalty": _pick_first_existing_col(pred, EXTRA_PENALTY_COL_CANDIDATES),
        "rating4_avg": _pick_first_existing_col(pred, RATING4_AVG_COL_CANDIDATES),
        "recent3_time_idx": _pick_first_existing_col(pred, RECENT3_TIME_IDX_COL_CANDIDATES),
    }
    if not race_cols or umaban_col is None or not any(metric_cols.values()):
        return {}

    # レースID表記は出力元によって rid_str / レースID のどちらでも来るため、両方をキー化する。
    pred["_race_keys_for_roi"] = pred.apply(
        lambda r: [
            str(key)
            for key in (_normalize_race_key(r.get(c)) for c in race_cols)
            if key is not None
        ],
        axis=1,
    )
    pred["_umaban_for_roi"] = pred[umaban_col].map(_to_umaban_int_for_roi)
    for metric_name, col in metric_cols.items():
        pred[f"_{metric_name}_for_roi"] = pd.to_numeric(pred[col], errors="coerce") if col is not None else np.nan

    pred = pred.dropna(subset=["_umaban_for_roi"])
    pred = pred[pred["_race_keys_for_roi"].map(bool)]
    if pred.empty:
        return {}

    metric_map: Dict[Tuple[str, int], Dict[str, float]] = {}
    for _, r in pred.iterrows():
        metrics = {
            "popularity": _to_float_safe(r.get("_popularity_for_roi")),
            "score": _to_float_safe(r.get("_score_for_roi")),
            "extra_penalty": _to_float_safe(r.get("_extra_penalty_for_roi")),
            "rating4_avg": _to_float_safe(r.get("_rating4_avg_for_roi")),
            "recent3_time_idx": _to_float_safe(r.get("_recent3_time_idx_for_roi")),
        }
        for race_key in r["_race_keys_for_roi"]:
            metric_map[(str(race_key), int(r["_umaban_for_roi"]))] = metrics
    return metric_map


def _rank3_metric_map_from_now_df(now_df: Optional[pd.DataFrame]) -> Dict[Tuple[str, int], Dict[str, float]]:
    """今走シートから (レースID, 馬番) -> 人気・score・extra_penalty の対応を作る。"""
    return _horse_metric_map_from_prediction_df(now_df)


def _horse_metrics_for_race_keys(
    metric_map: Dict[Tuple[str, int], Dict[str, float]],
    race_key_sources: Tuple[object, ...],
    umaban: object,
) -> Dict[str, float]:
    """複数のレースID表記候補から、指定馬番の指標を探す。"""
    umaban_int = _to_umaban_int_for_roi(umaban)
    if umaban_int is None:
        return {}

    for race_key_src in race_key_sources:
        race_key = _normalize_race_key(race_key_src)
        if race_key is None:
            continue
        metrics = metric_map.get((str(race_key), int(umaban_int)), {})
        if metrics:
            return metrics
    return {}


def _format_metric_value(value: Optional[float]) -> str:
    """理由欄に出す数値を、欠損時も読みやすい文字列へそろえる。"""
    return "不明" if value is None else f"{value:.2f}"


def _judge_bet_rank(
    score1: Optional[float],
    gap12: Optional[float],
    dango_2_5: Optional[float],
    rank1_rating4_avg: Optional[float],
    rank1_recent3_time_idx: Optional[float],
    rank1_extra_penalty: Optional[float],
    rank3_popularity: Optional[float],
    rank3_score: Optional[float],
    rank3_extra_penalty: Optional[float],
) -> Tuple[str, str, str]:
    """
    買い目_レース別1行のランク・判定・理由を決める。

    ランキング1位を軸にした3連複向けに、Sは1位馬の指定条件で判定する。
    A/Bは従来どおりscore1と2位との差を使い、補助指標は理由欄に残す。
    """
    rank1_detail = (
        f"1位rating4平均={_format_metric_value(rank1_rating4_avg)}"
        f" / 1位近3走タイム指数={_format_metric_value(rank1_recent3_time_idx)}"
        f" / 1位extra_penalty={_format_metric_value(rank1_extra_penalty)}"
    )
    rank3_detail = (
        f"3位人気={_format_metric_value(rank3_popularity)}"
        f" / 3位score={_format_metric_value(rank3_score)}"
        f" / 3位extra_penalty={_format_metric_value(rank3_extra_penalty)}"
    )
    base_detail = (
        f"score1={_format_metric_value(score1)}"
        f" / dango_2_5={_format_metric_value(dango_2_5)}"
        f" / gap12={_format_metric_value(gap12)}"
    )

    s_rank_ok = (
        score1 is not None
        and score1 >= S_RANK_SCORE_MIN
        and rank1_rating4_avg is not None
        and rank1_rating4_avg >= S_RANK_RATING4_AVG_MIN
        and rank1_recent3_time_idx is not None
        and rank1_recent3_time_idx >= S_RANK_RECENT3_TIME_IDX_MIN
        and rank1_extra_penalty is not None
        and rank1_extra_penalty <= S_RANK_EXTRA_PENALTY_MAX
    )
    if s_rank_ok:
        return (
            "S",
            "1位軸3連複の推奨レース",
            "S条件一致（ランキング1位: score>=64 かつ rating4平均>=400 "
            f"かつ 近3走タイム指数>=80 かつ extra_penalty<=1.25）。{base_detail} / {rank1_detail}、補助情報: {rank3_detail}",
        )

    a_score_ok = score1 is not None and score1 >= 68.0
    a_gap_ok = gap12 is not None and gap12 >= 5.0
    a_balanced_ok = score1 is not None and score1 >= 65.0 and gap12 is not None and gap12 >= 3.0
    if a_score_ok or a_gap_ok or a_balanced_ok:
        matched = []
        if a_score_ok:
            matched.append("score1>=68")
        if a_gap_ok:
            matched.append("gap12>=5")
        if a_balanced_ok:
            matched.append("score1>=65 かつ gap12>=3")
        matched_text = "、".join(matched)
        return (
            "A",
            "1位軸候補だがSより一段慎重",
            f"A条件一致（{matched_text}）。S条件は未達。{base_detail} / {rank1_detail}、補助情報: {rank3_detail}",
        )

    b_score_ok = score1 is not None and score1 >= 63.0
    b_gap_ok = gap12 is not None and gap12 >= 2.0
    if b_score_ok or b_gap_ok:
        return (
            "B",
            "少額・参考",
            f"B条件一致（score1>=63 または gap12>=2）。{base_detail} / {rank1_detail}、補助情報: {rank3_detail}",
        )

    return (
        "-",
        "見送り",
        f"見送り条件（S/A/Bに届かず）。{base_detail} / {rank1_detail}、補助情報: {rank3_detail}",
    )


def build_roi_focus_bet_sheet(
    bet_df: pd.DataFrame,
    now_df: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    回収率重視の条件に合うレースだけを、購入候補シート用に1レース1行へ整形する。

    注意:
    - 着順、払戻、的中判定などの結果情報は使わない。
    - score1/score2/gap12/dango_2_5/1位馬番〜5位馬番は既存の買い目シート値を使う。
    - ランキング3位の人気・score・extra_penaltyは今走シートの馬ごとの値を使う。
    """
    if bet_df is None or not isinstance(bet_df, pd.DataFrame) or bet_df.empty:
        return pd.DataFrame(columns=ROI_FOCUS_BET_COLUMNS)

    work = bet_df.copy()
    class_map = _class_map_from_now_df(now_df)
    surface_map = _surface_map_from_now_df(now_df)
    rank3_metric_map = _rank3_metric_map_from_now_df(now_df)

    if "クラス" not in work.columns:
        work["クラス"] = pd.NA

    if class_map:
        race_col = _pick_first_existing_col(work, ["レースID", "rid_str", "race_id"])
        if race_col is not None:
            race_key = work[race_col].map(_normalize_race_key)
            mapped_class = race_key.map(class_map)
            work["クラス"] = work["クラス"].combine_first(mapped_class)

    rows: list[dict] = []

    for _, row in work.iterrows():
        score1 = _to_float_safe(row.get("score1"))
        score2 = _to_float_safe(row.get("score2"))
        gap12 = _to_float_safe(row.get("gap12"))
        dango_2_5 = _to_float_safe(row.get("dango_2_5"))

        if score1 is None or dango_2_5 is None:
            continue
        if not (score1 >= 65.0 and dango_2_5 >= 6.0):
            continue
        if gap12 is not None and gap12 < 0.5:
            continue

        race_id = row.get("レースID", row.get("rid_str", pd.NA))
        race_key = _normalize_race_key(race_id)
        surface = surface_map.get(str(race_key), "") if race_key is not None else ""
        if surface == "":
            surface = _normalize_surface_for_roi(row.get("コース", pd.NA))
        if surface == "":
            continue

        ranks = [_to_umaban_int_for_roi(row.get(f"{i}位馬番")) for i in range(1, 6)]
        if any(v is None for v in ranks) or len(set(ranks)) != 5:
            print(
                f"[WARN] '{ROI_FOCUS_BET_SHEET}' 出力除外: "
                f"レースID={race_id} 1位〜5位馬番に欠損/重複/変換不可があります"
            )
            continue

        rank1, rank2, rank3, rank4, rank5 = [int(v) for v in ranks]
        rank3_metrics = rank3_metric_map.get((str(race_key), rank3), {}) if race_key is not None else {}
        rank3_popularity = _to_float_safe(rank3_metrics.get("popularity"))
        rank3_score = _to_float_safe(rank3_metrics.get("score"))
        rank3_extra_penalty = _to_float_safe(rank3_metrics.get("extra_penalty"))

        if rank3_popularity is None or rank3_score is None or rank3_extra_penalty is None:
            continue
        if not (rank3_popularity <= 5.0 and rank3_extra_penalty < 2.0 and rank3_score >= 57.0):
            continue

        rows.append(
            {
                "レースID": race_id,
                "レース名": row.get("レース名", pd.NA),
                "発走時刻": row.get("発走時刻", pd.NA),
                "場所": row.get("場所", pd.NA),
                "コース": row.get("コース", pd.NA),
                "馬場": row.get("馬場", pd.NA),
                "頭数": row.get("頭数", pd.NA),
                "クラス": row.get("クラス", pd.NA),
                "score1": score1,
                "score2": score2,
                "gap12": gap12,
                "dango_2_5": dango_2_5,
                "1位馬番": rank1,
                "2位馬番": rank2,
                "3位馬番": rank3,
                "4位馬番": rank4,
                "5位馬番": rank5,
                "購入判定": "購入",
                "購入理由": "score1>=65 かつ dango_2_5>=6 かつ gap12>=0.5 かつ surfaceが芝/ダ かつ 3位人気<=5 かつ 3位extra_penalty<2 かつ 3位score>=57",
                "3連複1点目_馬番1": rank1,
                "3連複1点目_馬番2": rank3,
                "3連複1点目_馬番3": rank2,
                "3連複2点目_馬番1": rank1,
                "3連複2点目_馬番2": rank3,
                "3連複2点目_馬番3": rank4,
                "3連複3点目_馬番1": rank1,
                "3連複3点目_馬番2": rank3,
                "3連複3点目_馬番3": rank5,
                "3連複_点数": 3,
                "3連複_金額": 300,
                "馬連1点目_馬番1": rank1,
                "馬連1点目_馬番2": rank3,
                "馬連2点目_馬番1": rank2,
                "馬連2点目_馬番2": rank3,
                "馬連_点数": 2,
                "馬連_金額": 200,
                "ワイド1点目_馬番1": rank1,
                "ワイド1点目_馬番2": rank3,
                "ワイド2点目_馬番1": rank2,
                "ワイド2点目_馬番2": rank3,
                "ワイド_点数": 2,
                "ワイド_金額": 200,
                "合計購入金額": 700,
            }
        )

    out = pd.DataFrame(rows)
    for c in ROI_FOCUS_BET_COLUMNS:
        if c not in out.columns:
            out[c] = pd.NA
    out = out[ROI_FOCUS_BET_COLUMNS]

    int_cols = [
        c
        for c in ROI_FOCUS_BET_COLUMNS
        if "馬番" in c or c.endswith("_点数") or c.endswith("_金額") or c == "合計購入金額"
    ]
    if not out.empty:
        for c in int_cols:
            out[c] = out[c].astype("int64")

    return out


def _delete_excel_sheet_if_exists(excel_path: str, sheet_name: str) -> bool:
    """同名シートをいったん削除して、次の書き込みで末尾に追加できる状態にする。"""
    if not os.path.exists(excel_path):
        return False

    wb = None
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            return False
        if len(wb.sheetnames) <= 1:
            return False

        del wb[sheet_name]
        wb.save(excel_path)
        return True
    finally:
        if wb is not None:
            wb.close()


def _bet_rank_readme_rows() -> list[list[object]]:
    """READMEシートに出す、買い目ランク条件の説明行を作る。"""
    return [
        [BET_RANK_README_START, ""],
        ["対象シート", BET_SHEET],
        ["対象列", "R列: ランク(S/A/B)"],
        ["目的", "ランキング1位を軸にした3連複を買うレース選択の目安"],
        ["前提", "Sはランキング1位馬のscore・rating4平均・近3走タイム指数・extra_penaltyで判定"],
        ["", ""],
        ["ランク", "条件", "使い方", "考え方"],
        [
            "S",
            "ランキング1位の score >= 64 かつ rating4平均 >= 400 かつ 近3走タイム指数 >= 80 かつ extra_penalty <= 1.25",
            "ランキング1位を軸にした3連複の主候補",
            "1位馬の絶対評価・近走時計・リスクの条件がそろっている状態",
        ],
        [
            "A",
            "score1 >= 68 または gap12 >= 5、または score1 >= 65 かつ gap12 >= 3",
            "Sより一段慎重な1位軸候補",
            "1位の絶対評価または2位との差はあるが、S条件ほど強くそろっていない状態",
        ],
        [
            "B",
            "score1 >= 63 または gap12 >= 2",
            "少額・参考候補",
            "1位軸としての根拠は弱めなので、他条件やオッズを確認して扱う",
        ],
        [
            "-",
            "上記条件を満たさない",
            "見送り候補",
            "ランキング1位を軸にする根拠が不足している状態",
        ],
        ["", ""],
        ["補足", "1位馬のrating4平均・近3走タイム指数・extra_penaltyと、3位馬の人気・score・extra_penaltyは理由欄に残す"],
        ["注意", "結果後にしか分からない着順・払戻はランク判定に使わない"],
        [BET_RANK_README_END, ""],
    ]


def _write_bet_rank_readme_to_excel(out_excel_path: str) -> None:
    """出力ExcelのREADMEシートへ、買い目ランク条件の説明ブロックを書き込む。"""
    if not os.path.exists(out_excel_path):
        return

    wb = None
    try:
        wb = load_workbook(out_excel_path)
        ws = wb[README_SHEET] if README_SHEET in wb.sheetnames else wb.create_sheet(README_SHEET)

        start_row = None
        end_row = None
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=1).value
            if value == BET_RANK_README_START:
                start_row = row_idx
            if start_row is not None and value == BET_RANK_README_END:
                end_row = row_idx
                break

        if start_row is not None and end_row is not None:
            ws.delete_rows(start_row, end_row - start_row + 1)

        has_existing_content = any(
            ws.cell(row=row_idx, column=col_idx).value not in (None, "")
            for row_idx in range(1, ws.max_row + 1)
            for col_idx in range(1, ws.max_column + 1)
        )
        write_row = ws.max_row + 2 if has_existing_content else 1

        for row_values in _bet_rank_readme_rows():
            for col_idx, value in enumerate(row_values, start=1):
                ws.cell(row=write_row, column=col_idx).value = value
            write_row += 1

        for row in ws.iter_rows():
            first_value = row[0].value if row else None
            if first_value in {BET_RANK_README_START, BET_RANK_README_END}:
                row[0].font = Font(bold=True)
            if first_value == "ランク":
                for cell in row:
                    cell.font = Font(bold=True)

        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 58
        ws.column_dimensions["C"].width = 34
        ws.column_dimensions["D"].width = 58

        wb.save(out_excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。README更新をスキップします: {out_excel_path}")
    except Exception as e:
        print(f"[WARN] READMEシートへのランク条件説明の書き込みに失敗しました: {e}")
    finally:
        if wb is not None:
            wb.close()


def _format_roi_focus_bet_worksheet(excel_path: str) -> None:
    """回収率重視_買い目候補 シートだけ最低限見やすく整える。"""
    wb = None
    try:
        wb = load_workbook(excel_path)
        if ROI_FOCUS_BET_SHEET not in wb.sheetnames:
            return

        ws = wb[ROI_FOCUS_BET_SHEET]
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.font = Font(bold=True)
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

        sample_max_row = min(ws.max_row, 200)
        for col_idx in range(1, ws.max_column + 1):
            values = [
                ws.cell(row=row_idx, column=col_idx).value
                for row_idx in range(1, sample_max_row + 1)
            ]
            max_len = max((len(str(v)) for v in values if v is not None), default=0)
            width = min(max(max_len + 2, 10), 32)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。整形をスキップします: {excel_path}")
    except Exception as e:
        print(f"[WARN] '{ROI_FOCUS_BET_SHEET}' の整形に失敗しました: {e}")
    finally:
        if wb is not None:
            wb.close()


def append_roi_focus_bet_sheet_to_excel(out_excel_path: str) -> int:
    """
    既存の 買い目_レース別1行 から、回収率重視_買い目候補 をブック末尾に作り直す。
    """
    if not os.path.exists(out_excel_path):
        print(f"[WARN] 出力Excelが見つからないため、'{ROI_FOCUS_BET_SHEET}' 作成をスキップ: {out_excel_path}")
        return 0

    xls = None
    try:
        xls = pd.ExcelFile(out_excel_path, engine="openpyxl")
        if BET_SHEET not in xls.sheet_names:
            print(f"[WARN] '{BET_SHEET}' シートが無いため、'{ROI_FOCUS_BET_SHEET}' 作成をスキップします")
            return 0

        bet_df = pd.read_excel(out_excel_path, sheet_name=BET_SHEET, engine="openpyxl")
        now_df = (
            pd.read_excel(out_excel_path, sheet_name=NOW_SHEET, engine="openpyxl")
            if NOW_SHEET in xls.sheet_names
            else pd.DataFrame()
        )
    except Exception as e:
        print(f"[WARN] '{ROI_FOCUS_BET_SHEET}' 用のExcel読み込みに失敗しました: {e}")
        return 0
    finally:
        if xls is not None:
            xls.close()

    roi_focus_bet_df = build_roi_focus_bet_sheet(bet_df=bet_df, now_df=now_df)

    try:
        _delete_excel_sheet_if_exists(out_excel_path, ROI_FOCUS_BET_SHEET)
        with pd.ExcelWriter(out_excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            roi_focus_bet_df.to_excel(writer, sheet_name=ROI_FOCUS_BET_SHEET, index=False)
        _format_roi_focus_bet_worksheet(out_excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。Excelを閉じてから再実行してください: {out_excel_path}")
        return 0
    except Exception as e:
        print(f"[WARN] '{ROI_FOCUS_BET_SHEET}' のExcel書き込みに失敗しました: {e}")
        return 0

    print(f"[INFO] '{ROI_FOCUS_BET_SHEET}' 作成完了: {len(roi_focus_bet_df)}レース -> {out_excel_path}")
    return int(len(roi_focus_bet_df))


def _build_bet_sheet(
    feat_export: pd.DataFrame,
    now_export: pd.DataFrame,
    odds_df: Optional[pd.DataFrame] = None,
    gap_min: float = 3.5,
    extra_th: float = 0.8,
    rest_th: float = 0.4,
) -> pd.DataFrame:
    """TARGET（feat_export）と今走（now_export）から
    買い目_レース別1行 を作る。

    ここでのポイント（過去出力の互換）：
    - score1/score2/gap12 は「上位2頭（同点含む）」で計算
    - dango_2_5 は「rank2 と rank5 の score差」で計算（rank不足なら 999）
    """

    # ----------------------------
    # オッズマップ
    # ----------------------------
    rid_odds_map, place_odds_map = _build_tansho_odds_maps_for_bet_sheet(odds_df)
    place_odds_dates = {k[0] for k in place_odds_map.keys() if k[0]}
    default_odds_date = next(iter(place_odds_dates)) if len(place_odds_dates) == 1 else ""

    # ----------------------------
    # 今走：レース情報（1レース1行）
    # ----------------------------
    now = now_export.copy()
    if "rid_str" not in now.columns:
        if "レースID" in now.columns:
            now["rid_str"] = now["レースID"].astype(str)
        else:
            now["rid_str"] = pd.NA
    now["rid_str"] = now["rid_str"].astype(str)

    race_info_cols = [
        "レースID",
        "レース名",
        "発走時刻",
        "場所",
        "コース",
        "馬場",
        "頭数",
        "レース種別",
        "クラス",
        "日付",
        "開催日",
        "年月日",
        "raceday",
        "date",
    ]
    for c in race_info_cols:
        if c not in now.columns:
            now[c] = pd.NA

    race_info = now.groupby("rid_str", as_index=False)[race_info_cols].first()
    now_metric_map = _horse_metric_map_from_prediction_df(now_export)

    # ----------------------------
    # TARGET：rid_str+馬番+score+rank を前提
    # ----------------------------
    ft = feat_export.copy()
    if "rid_str" not in ft.columns:
        raise ValueError("TARGET（feat_export）に rid_str 列がありません")

    ft["rid_str"] = ft["rid_str"].astype(str)

    if "馬番" not in ft.columns:
        for cand in ["umaban", "馬 番", "馬番 "]:
            if cand in ft.columns:
                ft["馬番"] = ft[cand]
                break

    ft["馬番"] = pd.to_numeric(ft["馬番"], errors="coerce").astype("Int64")
    ft["score"] = pd.to_numeric(ft.get("score", pd.Series([pd.NA] * len(ft))), errors="coerce")
    ft["rank"] = pd.to_numeric(ft.get("rank", pd.Series([pd.NA] * len(ft))), errors="coerce").astype("Int64")
    target_metric_map = _horse_metric_map_from_prediction_df(ft)

    bet_rows: list[dict] = []

    for rid, sub in ft.groupby("rid_str", sort=True):
        # 上位馬番（同点は馬番昇順）を取る
        sub2 = sub.sort_values(["score", "馬番"], ascending=[False, True], kind="mergesort")
        top7 = sub2.head(7)
        top1_row = top7.iloc[0] if not top7.empty else pd.Series(dtype=object)

        horses7 = [_to_int_safe(v) for v in top7["馬番"].tolist()]
        horses6 = horses7[:6] + [None] * (6 - len(horses7[:6]))
        horses7 = horses7[:7] + [None] * (7 - len(horses7[:7]))

        # score1/score2/gap12（上位2頭）
        top_scores = [_to_float_safe(v) for v in top7["score"].tolist()]
        score1 = top_scores[0] if len(top_scores) > 0 else None
        score2 = top_scores[1] if len(top_scores) > 1 else score1
        gap12 = round(float(score1 - score2), 2) if (score1 is not None and score2 is not None) else 0.0

        # dango_2_5：rank2とrank5のscore差（rank不足なら999）
        rank_score = sub2.dropna(subset=["rank", "score"]).groupby("rank")["score"].max().sort_index()
        if 2 in rank_score.index and 5 in rank_score.index:
            dango_2_5 = round(float(rank_score.loc[2] - rank_score.loc[5]), 2)
        else:
            dango_2_5 = 999.0

        # レース情報（無ければ最低限）
        info = race_info[race_info["rid_str"] == rid]
        info_row = info.iloc[0].to_dict() if not info.empty else {c: pd.NA for c in race_info_cols}
        if pd.isna(info_row.get("レースID")):
            info_row["レースID"] = rid
        race_key_sources = (info_row.get("レースID"), rid)
        row_date_key = ""
        for date_col in ["日付", "開催日", "年月日", "raceday", "date"]:
            row_date_key = _extract_yyyymmdd_for_odds_key(info_row.get(date_col))
            if row_date_key:
                break
        if not row_date_key:
            row_date_key = default_odds_date

        # 単勝オッズ（上位1頭）
        odds_top1 = None
        if horses6[0] is not None:
            top_umaban = int(horses6[0])
            top_horse_name_norm = _normalize_horse_name_for_odds_key(top1_row.get("馬名"))
            for race_key_src in race_key_sources:
                race_key = _normalize_race_key(race_key_src)
                if race_key is None:
                    continue
                odds_top1 = _lookup_rid_tansho_odds(rid_odds_map, str(race_key), top_umaban, top_horse_name_norm)
                if odds_top1 is not None:
                    break

            if odds_top1 is None:
                place_norm = _normalize_place(info_row.get("場所"))
                race_no = ""
                for race_key_src in race_key_sources:
                    race_no = _race_no_from_race_key(race_key_src)
                    if race_no:
                        break
                if place_norm and race_no:
                    odds_top1 = _lookup_place_tansho_odds(
                        place_odds_map,
                        row_date_key,
                        str(place_norm),
                        str(race_no),
                        top_umaban,
                        top_horse_name_norm,
                    )

        # 1位馬のrating4平均・近3走タイム指数・extra_penaltyをS判定に使う。
        rank1_target_metrics = _horse_metrics_for_race_keys(target_metric_map, race_key_sources, horses6[0])
        rank1_now_metrics = _horse_metrics_for_race_keys(now_metric_map, race_key_sources, horses6[0])
        rank1_rating4_avg = _first_float(
            _metric_from_row(top1_row, RATING4_AVG_COL_CANDIDATES),
            rank1_target_metrics.get("rating4_avg"),
            rank1_now_metrics.get("rating4_avg"),
        )
        rank1_recent3_time_idx = _first_float(
            _metric_from_row(top1_row, RECENT3_TIME_IDX_COL_CANDIDATES),
            rank1_target_metrics.get("recent3_time_idx"),
            rank1_now_metrics.get("recent3_time_idx"),
        )
        rank1_extra_penalty = _first_float(
            _metric_from_row(top1_row, EXTRA_PENALTY_COL_CANDIDATES),
            rank1_target_metrics.get("extra_penalty"),
            rank1_now_metrics.get("extra_penalty"),
        )

        # 3位馬の人気・score・extra_penaltyは理由欄と回収率重視シート用に残す。
        rank3_umaban = horses6[2]
        rank3_target_metrics = _horse_metrics_for_race_keys(target_metric_map, race_key_sources, rank3_umaban)
        rank3_now_metrics = _horse_metrics_for_race_keys(now_metric_map, race_key_sources, rank3_umaban)

        rank3_popularity = _first_float(rank3_now_metrics.get("popularity"), rank3_target_metrics.get("popularity"))
        rank3_score = _first_float(rank3_target_metrics.get("score"), rank3_now_metrics.get("score"))
        rank3_extra_penalty = _first_float(
            rank3_target_metrics.get("extra_penalty"),
            rank3_now_metrics.get("extra_penalty"),
        )

        rank_label, judge, reason = _judge_bet_rank(
            score1=score1,
            gap12=gap12,
            dango_2_5=dango_2_5,
            rank1_rating4_avg=rank1_rating4_avg,
            rank1_recent3_time_idx=rank1_recent3_time_idx,
            rank1_extra_penalty=rank1_extra_penalty,
            rank3_popularity=rank3_popularity,
            rank3_score=rank3_score,
            rank3_extra_penalty=rank3_extra_penalty,
        )

        bet_rows.append(
            {
                "レースID": info_row.get("レースID", rid),
                "レース名": info_row.get("レース名", pd.NA),
                "発走時刻": info_row.get("発走時刻", pd.NA),
                "場所": info_row.get("場所", pd.NA),
                "コース": info_row.get("コース", pd.NA),
                "馬場": info_row.get("馬場", pd.NA),
                "頭数": info_row.get("頭数", pd.NA),
                "score1": score1,
                "score2": score2,
                "gap12": gap12,
                "dango_2_5": dango_2_5,
                "1位馬番": horses6[0],
                "2位馬番": horses6[1],
                "3位馬番": horses6[2],
                "4位馬番": horses6[3],
                "5位馬番": horses6[4],
                "6位馬番": horses6[5],
                "ランク(S/A/B)": rank_label,
                "判定": judge,
                "理由": reason,
                "軸馬番": "-",
                "単勝オッズ_1位": odds_top1,
            }
        )

    bet_df = pd.DataFrame(bet_rows)
    bet_cols = [
        "レースID",
        "レース名",
        "発走時刻",
        "場所",
        "コース",
        "馬場",
        "頭数",
        "score1",
        "score2",
        "gap12",
        "dango_2_5",
        "1位馬番",
        "2位馬番",
        "3位馬番",
        "4位馬番",
        "5位馬番",
        "6位馬番",
        "ランク(S/A/B)",
        "判定",
        "理由",
        "軸馬番",
        "単勝オッズ_1位",
    ]
    for c in bet_cols:
        if c not in bet_df.columns:
            bet_df[c] = pd.NA
    bet_df = bet_df[bet_cols]

    return bet_df


# ================================================================
# Excel 出力処理
# ================================================================
def write_features_to_excel(
    src_excel: str,
    out_excel: str,
    feat_df: pd.DataFrame,
    now_df: pd.DataFrame,
    odds_df: Optional[pd.DataFrame] = None,
) -> None:
    """もとの EXCEL をコピーし、TARGET シートと今走シートを上書き。
    さらに過去版互換で 買い目_レース別1行 も作成する。
    """
    print(f"[INFO] 特徴量を {out_excel} に出力します")

    src_abs = os.path.normcase(os.path.abspath(src_excel))
    out_abs = os.path.normcase(os.path.abspath(out_excel))
    if src_abs == out_abs:
        print("[INFO] 入力Excelと出力Excelが同じため、コピーせず既存ブックを更新します")
    else:
        try:
            shutil.copy2(src_excel, out_excel)
        except PermissionError:
            stem, ext = os.path.splitext(out_excel)
            alt = f"{stem}_{datetime.now().strftime('%H%M%S')}{ext}"
            print(f"[WARN] 出力先ファイルに書き込めません（Excelで開いている可能性）: {out_excel} -> {alt}")
            shutil.copy2(src_excel, alt)
            out_excel = alt

    feat_export = _build_feature_sheet_for_export(feat_df, FEAT_COLS, JAPANESE_FEATURE_NAMES)

    # TARGETは rid_str ごとに rank 昇順（上位=1が先）に並べる
    if {"rid_str", "rank"}.issubset(feat_export.columns):
        feat_export = feat_export.copy()
        feat_export["_rid_sort"] = feat_export["rid_str"].astype(str).str.replace(r"\D", "", regex=True)
        feat_export["_rank_sort"] = pd.to_numeric(feat_export["rank"], errors="coerce")
        feat_export = feat_export.sort_values(
            ["_rid_sort", "_rank_sort"],
            ascending=[True, True],
            kind="mergesort",
            na_position="last",
        ).drop(columns=["_rid_sort", "_rank_sort"])

    now_export = now_df.copy()

    # ★追加：買い目シートを生成
    try:
        bet_df = _build_bet_sheet(feat_export=feat_export, now_export=now_export, odds_df=odds_df)
    except ValueError:
        raise
    except Exception as e:
        print(f"[WARN] '{BET_SHEET}' の作成に失敗したためスキップします: {e}")
        bet_df = pd.DataFrame()

    with pd.ExcelWriter(out_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        feat_export.to_excel(writer, sheet_name=TARGET_SHEET, index=False)
        now_export.to_excel(writer, sheet_name=NOW_SHEET, index=False)

        if not bet_df.empty:
            bet_df.to_excel(writer, sheet_name=BET_SHEET, index=False)

    if not bet_df.empty:
        _write_bet_rank_readme_to_excel(out_excel)
        append_roi_focus_bet_sheet_to_excel(out_excel)


def append_success_report(df: pd.DataFrame, report_path: str) -> None:
    """success_report.xlsx に簡易集計を追記。"""
    n_races = df["rid_str"].nunique() if "rid_str" in df.columns else 0
    n_horses = len(df)

    row = {
        "日付": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "レース数": n_races,
        "頭数": n_horses,
    }

    if os.path.exists(report_path):
        rep = pd.read_excel(report_path, engine="openpyxl")
        rep = pd.concat([rep, pd.DataFrame([row])], ignore_index=True)
    else:
        rep = pd.DataFrame([row])

    rep.to_excel(report_path, index=False)
    print(f"[INFO] success_report.xlsx を更新しました: {report_path}")


def _merge_dl_rank_override(merged: pd.DataFrame, dl_rank_df: Optional[pd.DataFrame]) -> pd.DataFrame:
    """メモリ上で作成した dl_rank / dl_prob を今走データへ結合する。"""
    if dl_rank_df is None or dl_rank_df.empty:
        return merged
    if merged.empty:
        return merged
    if not {"rid_str", "馬番"}.issubset(merged.columns):
        raise RuntimeError("今走データに rid_str / 馬番 列が無く、DL順位データを結合できません")

    required_cols = {"rid_str", "馬番"}
    if not required_cols.issubset(dl_rank_df.columns):
        missing = ", ".join(sorted(required_cols - set(dl_rank_df.columns)))
        raise RuntimeError(f"DL順位データに必要列がありません: {missing}")

    value_cols = [c for c in ["dl_rank", "dl_prob"] if c in dl_rank_df.columns]
    if not value_cols:
        return merged

    work = dl_rank_df[["rid_str", "馬番", *value_cols]].copy()
    work["rid_str"] = _normalize_rid_series(work["rid_str"])
    work["馬番"] = _normalize_umaban_series(work["馬番"])
    if "dl_rank" in work.columns:
        work["dl_rank"] = pd.to_numeric(work["dl_rank"], errors="coerce")
    if "dl_prob" in work.columns:
        work["dl_prob"] = pd.to_numeric(work["dl_prob"], errors="coerce")
    work = work.drop_duplicates(subset=["rid_str", "馬番"], keep="last")

    out = merged.copy()
    out["rid_str"] = _normalize_rid_series(out["rid_str"])
    out["馬番"] = _normalize_umaban_series(out["馬番"])
    out = out.drop(columns=[c for c in value_cols if c in out.columns], errors="ignore")
    out = pd.merge(out, work, on=["rid_str", "馬番"], how="left")
    print(f"[INFO] DL順位データをメモリ上で結合しました: {len(work)}行")
    return out


# ================================================================
# メイン処理
# ================================================================
def run_pipeline(
    SRC_EXCEL: str,
    OUT_EXCEL: str,
    LEVELS_XL: str = str(RACE_LEVEL_XLSX),
    BASE_TIME: str = str(BASE_TIME_XLSX),
    ODDS_CSV_PATH: str = str(ODDS_CSV),
    RACEDAY: str | None = None,
    DL_RANK_DF: Optional[pd.DataFrame] = None,
) -> None:
    # 各種マスタ読み込み
    levels_df = load_race_levels(LEVELS_XL)
    base_time_df = load_base_time(BASE_TIME)
    odds_df = load_odds_csv(ODDS_CSV_PATH, raceday=RACEDAY)

    # 特徴量構築
    merged, feat_df = build_features_from_excel(
        SRC_EXCEL,
        levels_df,
        base_time_df,
        odds_df,
        raceday=RACEDAY,
    )

    # ============================================================
    # 今回追加:
    # レース登録馬の過去走情報が足りないレースを予想対象から除外
    # ============================================================
    merged, feat_df = _exclude_races_with_missing_history(merged, feat_df)
    merged = _merge_dl_rank_override(merged, DL_RANK_DF)

    # 除外後に空になった場合
    if merged.empty or feat_df.empty:
        print("[WARN] 過去走不足レース除外後、予想対象がありませんでした")
        out_df = merged.copy()

        # score系の列が無いと後続や出力で困るので、念のため空列を作る
        for c in ["score", "rank", "favorite_risk", "extra_penalty", "rest_dist_risk", "dl_rank_score"]:
            if c not in out_df.columns:
                out_df[c] = pd.NA

        write_features_to_excel(
            src_excel=SRC_EXCEL,
            out_excel=OUT_EXCEL,
            feat_df=feat_df,
            now_df=out_df,
            odds_df=odds_df,
        )
        append_success_report(out_df, str(SUCCESS_REPORT))
        return

    # 場所・馬場を rid_str ごとに取得
    place_map: Dict[str, str] = {}
    surface_map: Dict[str, str] = {}

    if "場所" in merged.columns:
        place_map = merged.groupby("rid_str")["場所"].first().to_dict()
    if "芝・ダ" in merged.columns:
        surface_map = merged.groupby("rid_str")["芝・ダ"].first().to_dict()
    elif "芝ダ" in merged.columns:
        surface_map = merged.groupby("rid_str")["芝ダ"].first().to_dict()

    # 距離マップ
    dist_map = {}
    for col in ["距離", "距離(m)", "距離 ", "Distance"]:
        if col in merged.columns:
            dist_map = (
                merged.groupby("rid_str")[col]
                .first()
                .apply(
                    lambda v: _to_int(re.search(r"(\d+)", str(v)).group(1))
                    if pd.notna(v) and re.search(r"(\d+)", str(v))
                    else None
                )
                .to_dict()
            )
            break

    # 頭数マップ
    field_size_map = {}
    for col in ["頭数", "頭 数", "field_size"]:
        if col in merged.columns:
            field_size_map = merged.groupby("rid_str")[col].first().apply(lambda v: _to_int(v)).to_dict()
            break

    # 馬場マップ
    baba_map = {}
    for col in ["馬場状態", "馬場", "馬 場"]:
        if col in merged.columns:
            baba_map = merged.groupby("rid_str")[col].first().to_dict()
            break

    # 人気マップ（rid_str, 馬番 -> 人気）
    pop_map = {}
    pop_col = None
    for col in merged.columns:
        if "人気" in str(col):
            pop_col = col
            break
    if pop_col:
        pop_series = pd.to_numeric(merged[pop_col], errors="coerce")
        pop_map = {
            (str(rid), _to_int(uma)): _to_int(pop)
            for rid, uma, pop in zip(merged.get("rid_str"), merged.get("馬番"), pop_series)
            if _to_int(uma) is not None and _to_int(pop) is not None
        }

    calc_fav_risk = build_calc_favorite_risk(place_map, surface_map, dist_map, field_size_map, pop_map, baba_map)

    # dl_rank を feat_df に付与（rid_str+馬番で結合）
    dl_join = merged[["rid_str", "馬番"]].copy()
    dl_join["dl_rank"] = merged["dl_rank"] if "dl_rank" in merged.columns else pd.NA
    dl_join["dl_prob"] = merged["dl_prob"] if "dl_prob" in merged.columns else pd.NA
    dl_join["頭数"] = merged["頭数"] if "頭数" in merged.columns else pd.NA
    dl_join["rid_str"] = _normalize_rid_series(dl_join["rid_str"])
    dl_join["馬番"] = _normalize_umaban_series(dl_join["馬番"])
    dl_join["dl_rank"] = pd.to_numeric(dl_join["dl_rank"], errors="coerce")
    dl_join["dl_prob"] = pd.to_numeric(dl_join["dl_prob"], errors="coerce")
    dl_join["頭数"] = pd.to_numeric(dl_join["頭数"], errors="coerce")

    feat_df["rid_str"] = _normalize_rid_series(feat_df["rid_str"])
    feat_df["馬番"] = _normalize_umaban_series(feat_df["馬番"])
    feat_df = pd.merge(feat_df, dl_join, on=["rid_str", "馬番"], how="left")

    # dl_rank_score の計算（事故防止の例外ルール付き）
    def _calc_dl_rank_score(row: pd.Series) -> float:
        r = row.get("dl_rank")
        n = row.get("頭数")
        if pd.isna(r) or pd.isna(n):
            return 0.5
        try:
            r_f = float(r)
            n_f = float(n)
        except Exception:
            return 0.5
        if n_f <= 1 or r_f < 1 or r_f > n_f:
            return 0.5
        return (n_f - r_f) / (n_f - 1.0)

    feat_df["dl_rank_score"] = feat_df.apply(_calc_dl_rank_score, axis=1)

    # dl_prob をレース内 0-1 に正規化して、dl_rank_score と混合する。
    # 確率差の情報を残しつつ、順位情報も少し残す。
    feat_df["dl_prob_score"] = pd.to_numeric(feat_df.get("dl_prob"), errors="coerce")

    def _normalize_prob_within_race(s: pd.Series) -> pd.Series:
        x = pd.to_numeric(s, errors="coerce")
        if x.notna().sum() == 0:
            return pd.Series([0.5] * len(s), index=s.index, dtype=float)
        mn = x.min(skipna=True)
        mx = x.max(skipna=True)
        if pd.isna(mn) or pd.isna(mx) or mx == mn:
            return pd.Series([0.5] * len(s), index=s.index, dtype=float)
        return ((x - mn) / (mx - mn)).fillna(0.5)

    feat_df["dl_prob_score"] = feat_df.groupby("rid_str")["dl_prob_score"].transform(_normalize_prob_within_race)
    feat_df["dl_score"] = (
        pd.to_numeric(feat_df["dl_prob_score"], errors="coerce").fillna(0.5) * DL_PROB_BLEND
        + pd.to_numeric(feat_df["dl_rank_score"], errors="coerce").fillna(0.5) * DL_RANK_BLEND
    )
    feat_df = feat_df.drop(columns=["頭数"], errors="ignore")

    feat_df = compute_scores_with_pipeline_logic(
        feat_df,
        place_map=place_map,
        surface_map=surface_map,
        calc_fav_risk=calc_fav_risk,
    )

    # 既に同名列が merged 側に入っている場合（過去の with_feat を入力にした等）
    # pandas merge が suffix（_x/_y）を付ける際に列名が衝突して MergeError になることがあります。
    # ここでは「今回あらためて計算した列」で上書きする前提で、古い同名列を削除してから結合します。
    _cols_to_add = ["score", "rank", "favorite_risk", "extra_penalty", "rest_dist_risk", "dl_rank_score"]
    _cols_to_drop = []
    for _c in _cols_to_add:
        _cols_to_drop.extend([_c, f"{_c}_x", f"{_c}_y"])
    merged = merged.drop(columns=[c for c in _cols_to_drop if c in merged.columns], errors="ignore")

    # 今走情報へ結合（rest_dist_risk も出力する）
    out_df = pd.merge(
        merged,
        feat_df[["rid_str", "馬番", "score", "rank", "favorite_risk", "extra_penalty", "rest_dist_risk", "dl_rank_score"]],
        on=["rid_str", "馬番"],
        how="left",
    )

    # Excel 出力
    write_features_to_excel(
        src_excel=SRC_EXCEL,
        out_excel=OUT_EXCEL,
        feat_df=feat_df,
        now_df=out_df,
        odds_df=odds_df,  # ★追加：買い目シートの単勝オッズ反映に使う
    )

    # 集計
    append_success_report(out_df, str(SUCCESS_REPORT))
