# -*- coding: utf-8 -*-
"""
keibayosou_best_import_roi_runner.py

やりたいこと
- 1回目: 元の入力Excelから特徴量計算して with_feat を作る
- 2回目の前: with_feat を元に dl_rank / dl_prob をメモリ上で自動作成する
- 2回目: dl_rank / dl_prob をメモリ渡しして最終予想Excelを作る

このコードでは、
1) keibayosou_best_import_roi_runner.py を1回目実行
2) make_dl_rank_from_racedata_results.py 相当のDL順位作成
3) keibayosou_best_import_roi_runner.py の2回目予想

という3回実行を、

keibayosou_best_import_roi_runner.py の1回実行

だけで完了できるようにしています。

今回の修正ポイント
- 馬の競走成績_with_feat_YYYYMMDD_with_dl.xlsx は作成しない
- dl_rank / dl_prob はDataFrameで2回目の予想処理へ渡す
- 既存の動きはなるべく崩さない
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
import torch
from openpyxl import load_workbook
from torch import nn
from torch.utils.data import DataLoader, TensorDataset

from keibayosou_config import BASE_DIR, HORSE_RESULTS_DIR, RACE_LEVEL_XLSX, BASE_TIME_XLSX, ODDS_CSV
from keibayosou_loaders import load_odds_csv
from keibayosou_pipeline import append_roi_focus_bet_sheet_to_excel, run_pipeline
from keibayosou_utils import _normalize_place


# ============================================================
# 基本パス
# ============================================================
OUTPUT_DIR = BASE_DIR / "data" / "output"
TRAIN_XLSX = BASE_DIR / "data" / "master" / "racedata_results.xlsx"
OZZU_SCRAPER_SCRIPT = BASE_DIR / "etc_py" / "02_scrape_jra_odds_2.py"
NOW_SHEET = "今走レース情報"
TARGET_SHEET = "TARGET"
EST_IN3_SHEET = "推定馬券内率"
VALUE_HORSE_SHEET = "妙味あり馬"
DANGER_HORSE_SHEET = "危険馬"
RANK_RATE_TABLE_SHEET = "rank_rate_table"
SCORE_RATE_TABLE_SHEET = "score_rate_table"
BET_SHEET = "買い目_レース別1行"
ROI_FOCUS_BET_SHEET = "回収率重視_買い目候補"
RACE_ID_FORMAT_SHEETS = [BET_SHEET, ROI_FOCUS_BET_SHEET, DANGER_HORSE_SHEET]
AXIS_UMABAN_COLUMN = "軸馬番"
OLD_AXIS_BET_COLUMNS = ["1頭軸_馬番", "1頭軸_相手", "1頭軸_金額", "保険BOX_馬番", "保険BOX_金額"]


# ============================================================
# 推定馬券内率用のデフォルト値
# ============================================================
DEFAULT_RANK_IN3_RATE: Dict[int, float] = {
    1: 45.0,
    2: 36.0,
    3: 30.0,
    4: 24.0,
    5: 19.0,
    6: 15.0,
    7: 12.0,
    8: 9.0,
    9: 7.0,
    10: 5.0,
}

DEFAULT_SCORE_IN3_RATE: Dict[str, float] = {
    "90以上": 55.0,
    "80〜90未満": 43.0,
    "70〜80未満": 31.0,
    "60〜70未満": 20.0,
    "50〜60未満": 12.0,
    "50未満": 6.0,
}

SCORE_BAND_ORDER = ["90以上", "80〜90未満", "70〜80未満", "60〜70未満", "50〜60未満", "50未満"]
EST_IN3_RESULT_COLS = [
    "レースID",
    "馬番",
    "馬名",
    "予想順位",
    "score",
    "頭数",
    "単勝オッズ",
    "複勝オッズ",
    "人気",
    "順位別馬券内率",
    "score帯馬券内率",
    "基本馬券内率",
    "score差補正",
    "頭数補正係数",
    "条件適性補正",
    "オッズ補正",
    "穴馬救済補正",
    "危険馬補正",
    "補正後馬券内率",
    "レース内調整係数",
    "推定馬券内率",
    "推定馬券内率_補正前",
    "オッズ帯基準馬券内率",
    "オッズ帯補正係数",
    "推定馬券内率_オッズ補正後",
    "市場評価オッズ種別",
    "市場馬券内率",
    "期待値_補正前",
    "期待値",
    "妙味判定",
]

DANGER_HORSE_RESULT_COLS = [
    "レースID",
    "場所",
    "危険人気馬_馬番",
    "危険人気馬_馬名",
    "危険人気馬_人気",
    "危険人気馬_rank",
    "危険人気馬_score",
    "危険人気馬スコア",
    "危険人気馬_判定",
    "危険人気馬_理由",
    "危険人気馬_単勝オッズ",
    "危険人気馬_複勝オッズ",
    "危険人気馬_extra_penalty",
    "危険人気馬_favorite_risk",
    "危険人気馬_rest_dist_risk",
    "危険人気馬_score1差",
    "4_8番人気_score上位頭数",
    "レースscore1",
    "頭数",
]


# ============================================================
# DL順位作成用の列候補
# ============================================================
COLS_CANDIDATES: Dict[str, List[str]] = {
    "horse_name": ["馬名"],
    "finish": ["着順", "着 順"],
    "race_id": ["レースID", "race_id"],
    "umaban": ["馬番", "馬 番"],
    "popularity": ["人気", "人 気"],
    "odds": ["単勝オッズ", "単勝 オッズ", "単勝"],
    "frame": ["枠"],
    "weight": ["斤量"],
    "sex_age": ["性齢"],
    "body_weight": ["馬体重(増減)", "馬体重 (増減)", "馬体重"],
}

DL_FEATURE_COLS = [
    "popularity",
    "odds",
    "frame",
    "weight",
    "age",
    "sex",
    "body_weight",
    "body_weight_diff",
]


# ============================================================
# パス解決
# ============================================================
def _resolve_raw_src_out_paths(raceday_str: Optional[str]) -> Tuple[str, str]:
    """
    1回目実行用
    - 入力: 元の 馬の競走成績_YYYYMMDD.xlsx
    - 出力: data/output/馬の競走成績_with_feat_YYYYMMDD.xlsx
    """
    if raceday_str:
        candidate = HORSE_RESULTS_DIR / f"馬の競走成績_{raceday_str}.xlsx"
        if os.path.exists(candidate):
            src_excel = str(candidate)
        else:
            src_excel = str(HORSE_RESULTS_DIR / "馬の競走成績.xlsx")
        out_excel = str(OUTPUT_DIR / f"馬の競走成績_with_feat_{raceday_str}.xlsx")
    else:
        src_excel = str(HORSE_RESULTS_DIR / "馬の競走成績.xlsx")
        out_excel = str(OUTPUT_DIR / "馬の競走成績_with_feat.xlsx")
    return src_excel, out_excel


def _pick_actual_out_excel(expected_out_excel: str) -> str:
    """
    pipeline 側が PermissionError 回避のため、
    末尾に _HHMMSS を付けて保存することがある。
    その場合でも「実際にできた出力Excel」を拾う。
    """
    if os.path.exists(expected_out_excel):
        return expected_out_excel

    out_dir = os.path.dirname(expected_out_excel)
    base = os.path.splitext(os.path.basename(expected_out_excel))[0]
    ext = os.path.splitext(expected_out_excel)[1]

    if not os.path.isdir(out_dir):
        return expected_out_excel

    cands = [
        os.path.join(out_dir, f)
        for f in os.listdir(out_dir)
        if f.startswith(base) and f.endswith(ext)
    ]
    if not cands:
        return expected_out_excel

    return max(cands, key=lambda p: os.path.getmtime(p))


# ============================================================
# 推定馬券内率
# ============================================================
def _coerce_float_series(s: pd.Series) -> pd.Series:
    """
    数値列または「1.2-1.5」のような文字列から、先頭の数値を取り出す。
    複勝オッズの範囲表記にも最低限対応するための補助関数。
    """
    direct = pd.to_numeric(s, errors="coerce")
    text = s.astype(str).str.replace(",", "", regex=False)
    picked = pd.to_numeric(text.str.extract(r"([-+]?\d+(?:\.\d+)?)", expand=False), errors="coerce")
    return direct.combine_first(picked)


def _rank_bucket_value(rank_val: object) -> int:
    """予想順位を 1〜9 / 10位以下 の集計キーへ寄せる。"""
    rank_num = pd.to_numeric(pd.Series([rank_val]), errors="coerce").iloc[0]
    if pd.isna(rank_num):
        return 10
    rank_int = int(rank_num)
    if rank_int <= 0:
        return 10
    return rank_int if rank_int <= 9 else 10


def _rank_bucket_label(bucket: int) -> object:
    """順位別集計表に出す表示値を返す。"""
    return "10位以下" if int(bucket) >= 10 else int(bucket)


def _score_band_label(score_val: object) -> str:
    """scoreを、過去集計と現在予想で共通利用するscore帯へ変換する。"""
    score_num = pd.to_numeric(pd.Series([score_val]), errors="coerce").iloc[0]
    if pd.isna(score_num):
        return "50未満"
    score_float = float(score_num)
    if score_float >= 90.0:
        return "90以上"
    if score_float >= 80.0:
        return "80〜90未満"
    if score_float >= 70.0:
        return "70〜80未満"
    if score_float >= 60.0:
        return "60〜70未満"
    if score_float >= 50.0:
        return "50〜60未満"
    return "50未満"


def _canonical_prediction_frame(df: pd.DataFrame) -> pd.DataFrame:
    """
    予想DataFrameの列名ゆらぎを吸収し、推定馬券内率計算に必要な基本列を作る。
    必須に近い列が無い場合も、後続で可能な範囲で補完する。
    """
    out = df.copy()

    c_rid = _pick_col(out, ["rid_str", "レースID", "race_id"])
    if c_rid is None:
        raise RuntimeError("推定馬券内率計算に必要な rid_str/レースID 列が見つかりません")
    out["rid_str"] = _normalize_rid_series(out[c_rid])

    c_umaban = _pick_col(out, ["馬番", "馬 番", "umaban"])
    if c_umaban is None:
        raise RuntimeError("推定馬券内率計算に必要な 馬番 列が見つかりません")
    out["馬番"] = _normalize_umaban_series(out[c_umaban])

    c_name = _pick_col(out, ["馬名", "horse_name", "name"])
    if c_name is not None and c_name != "馬名":
        out["馬名"] = out[c_name]
    elif "馬名" not in out.columns:
        out["馬名"] = pd.NA

    c_score = _pick_col(out, ["score", "スコア"])
    if c_score is not None and c_score != "score":
        out["score"] = _coerce_float_series(out[c_score])
    elif "score" in out.columns:
        out["score"] = _coerce_float_series(out["score"])
    else:
        out["score"] = np.nan

    c_rank = _pick_col(out, ["予想順位", "rank", "順位"])
    if c_rank is not None:
        out["予想順位"] = pd.to_numeric(out[c_rank], errors="coerce")
    elif out["score"].notna().any():
        out["予想順位"] = out.groupby("rid_str")["score"].rank(ascending=False, method="dense")
    else:
        out["予想順位"] = np.nan

    c_field = _pick_col(out, ["頭数", "頭 数", "field_size"])
    if c_field is not None:
        out["頭数"] = pd.to_numeric(out[c_field], errors="coerce")
    else:
        out["頭数"] = out.groupby("rid_str")["馬番"].transform("count")

    c_pop = _pick_col(out, ["人気", "人 気", "popularity"])
    if c_pop is not None:
        out["人気"] = pd.to_numeric(out[c_pop], errors="coerce")
    elif "人気" not in out.columns:
        out["人気"] = np.nan

    c_tansho = _pick_col(out, ["単勝オッズ", "単勝 オッズ", "単勝", "tansho", "オッズ"])
    if c_tansho is not None:
        out["単勝オッズ"] = _coerce_float_series(out[c_tansho])
    elif "単勝オッズ" not in out.columns:
        out["単勝オッズ"] = np.nan

    c_fukusho = _pick_col(out, ["複勝オッズ", "複勝", "fukusho"])
    if c_fukusho is not None:
        out["複勝オッズ"] = _coerce_float_series(out[c_fukusho])
    elif "複勝オッズ" not in out.columns:
        out["複勝オッズ"] = np.nan

    if "レースID" not in out.columns:
        out["レースID"] = out["rid_str"]

    return out


def _load_actual_in3_df(results_path: Path = TRAIN_XLSX) -> pd.DataFrame:
    """
    実結果Excelから、(rid_str, 馬番) ごとの3着内フラグを作る。
    過去予想Excel側に実着順列が無い場合でも、このマスタと照合して集計できる。
    """
    if not results_path.exists():
        print(f"[WARN] 実結果Excelが見つかりません: {results_path}")
        return pd.DataFrame(columns=["rid_str", "馬番", "馬名", "着順", "馬券内"])

    book = pd.read_excel(results_path, sheet_name=None, engine="openpyxl")
    rows: List[pd.DataFrame] = []

    for _, df in book.items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue

        c_race_id = _pick_col(df, COLS_CANDIDATES["race_id"])
        c_umaban = _pick_col(df, COLS_CANDIDATES["umaban"])
        c_name = _pick_col(df, COLS_CANDIDATES["horse_name"])
        c_finish = _pick_col(df, COLS_CANDIDATES["finish"])

        if c_race_id is None or c_umaban is None or c_finish is None:
            continue

        rid_str = _normalize_rid_series(df[c_race_id])
        umaban = _normalize_umaban_series(df[c_umaban])
        finish = pd.to_numeric(df[c_finish], errors="coerce")

        use = pd.DataFrame(
            {
                "rid_str": rid_str,
                "馬番": umaban,
                "馬名": df[c_name].astype(str) if c_name is not None else pd.NA,
                "着順": finish,
            }
        )
        use = use[(use["rid_str"].str.len() == 12) & use["馬番"].notna() & use["着順"].notna()].copy()
        if use.empty:
            continue

        use["馬番"] = use["馬番"].astype("Int64")
        use["馬券内"] = use["着順"].between(1, 3).astype(int)
        rows.append(use)

    if not rows:
        return pd.DataFrame(columns=["rid_str", "馬番", "馬名", "着順", "馬券内"])

    actual = pd.concat(rows, ignore_index=True)
    actual = actual.drop_duplicates(subset=["rid_str", "馬番"], keep="first")
    return actual


def _load_historical_prediction_df(exclude_paths: Optional[List[Path]] = None) -> pd.DataFrame:
    """
    data/output の過去予想Excelから、順位・scoreを持つ行を集める。
    _with_dl は中間ファイルなので除外し、最終 with_feat だけを使う。
    """
    exclude_resolved = {p.resolve() for p in (exclude_paths or []) if p is not None and p.exists()}
    files = sorted(OUTPUT_DIR.glob("馬の競走成績_with_feat_*.xlsx"))
    rows: List[pd.DataFrame] = []

    for path in files:
        if "_with_dl" in path.stem:
            continue
        if path.exists() and path.resolve() in exclude_resolved:
            continue

        try:
            xls = pd.ExcelFile(path, engine="openpyxl")
            sheet_name = TARGET_SHEET if TARGET_SHEET in xls.sheet_names else None
            if sheet_name is None and NOW_SHEET in xls.sheet_names:
                sheet_name = NOW_SHEET
            if sheet_name is None:
                continue

            pred_raw = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
            pred = _canonical_prediction_frame(pred_raw)
            pred = pred[["rid_str", "馬番", "馬名", "score", "予想順位"]].copy()
            pred = pred.dropna(subset=["rid_str", "馬番", "score", "予想順位"])
            pred["馬番"] = pred["馬番"].astype("Int64")
            pred["source_file"] = path.name
            if not pred.empty:
                rows.append(pred)
        except Exception as e:
            print(f"[WARN] 過去予想Excelの読み込みをスキップします: {path.name} / {e}")

    if not rows:
        return pd.DataFrame(columns=["rid_str", "馬番", "馬名", "score", "予想順位", "source_file"])

    return pd.concat(rows, ignore_index=True)


def _load_historical_prediction_result_df(exclude_paths: Optional[List[Path]] = None) -> pd.DataFrame:
    """
    過去予想と実結果を照合して、順位別・score帯別の馬券内率集計に使う明細を作る。
    """
    pred = _load_historical_prediction_df(exclude_paths=exclude_paths)
    if pred.empty:
        return pd.DataFrame(columns=["rid_str", "馬番", "馬名", "score", "予想順位", "馬券内"])

    actual = _load_actual_in3_df(TRAIN_XLSX)
    if actual.empty:
        return pd.DataFrame(columns=["rid_str", "馬番", "馬名", "score", "予想順位", "馬券内"])

    merged = pd.merge(
        pred,
        actual[["rid_str", "馬番", "馬券内"]],
        on=["rid_str", "馬番"],
        how="inner",
    )

    if merged.empty and "馬名" in pred.columns and "馬名" in actual.columns:
        pred_name = pred.dropna(subset=["馬名"]).copy()
        actual_name = actual.dropna(subset=["馬名"]).copy()
        pred_name["馬名"] = pred_name["馬名"].astype(str).str.strip()
        actual_name["馬名"] = actual_name["馬名"].astype(str).str.strip()
        merged = pd.merge(
            pred_name,
            actual_name[["rid_str", "馬名", "馬券内"]],
            on=["rid_str", "馬名"],
            how="inner",
        )

    return merged


def build_rank_rate_table(history_df: pd.DataFrame) -> pd.DataFrame:
    """
    ステップ1:
    過去の予想結果から「予想順位ごとの3着内率」を集計する。
    """
    rows: List[Dict[str, object]] = []

    work = history_df.copy()
    if not work.empty:
        work["rank_bucket"] = work["予想順位"].map(_rank_bucket_value)
        grouped = work.groupby("rank_bucket")["馬券内"].agg(["count", "sum"])
    else:
        grouped = pd.DataFrame(columns=["count", "sum"])

    for bucket in range(1, 11):
        if bucket in grouped.index:
            count = int(grouped.loc[bucket, "count"])
            in3_count = int(grouped.loc[bucket, "sum"])
            rate = round((in3_count / count) * 100.0, 2) if count > 0 else DEFAULT_RANK_IN3_RATE[bucket]
        else:
            count = 0
            in3_count = 0
            rate = DEFAULT_RANK_IN3_RATE[bucket]

        rows.append(
            {
                "予想順位": _rank_bucket_label(bucket),
                "件数": count,
                "馬券内数": in3_count,
                "馬券内率": rate,
            }
        )

    return pd.DataFrame(rows)


def build_score_rate_table(history_df: pd.DataFrame) -> pd.DataFrame:
    """
    ステップ2:
    過去の予想結果から「score帯ごとの3着内率」を集計する。
    """
    rows: List[Dict[str, object]] = []

    work = history_df.copy()
    if not work.empty:
        work["score帯"] = work["score"].map(_score_band_label)
        grouped = work.groupby("score帯")["馬券内"].agg(["count", "sum"])
    else:
        grouped = pd.DataFrame(columns=["count", "sum"])

    for band in SCORE_BAND_ORDER:
        if band in grouped.index:
            count = int(grouped.loc[band, "count"])
            in3_count = int(grouped.loc[band, "sum"])
            rate = round((in3_count / count) * 100.0, 2) if count > 0 else DEFAULT_SCORE_IN3_RATE[band]
        else:
            count = 0
            in3_count = 0
            rate = DEFAULT_SCORE_IN3_RATE[band]

        rows.append(
            {
                "score帯": band,
                "件数": count,
                "馬券内数": in3_count,
                "馬券内率": rate,
            }
        )

    return pd.DataFrame(rows)


def _rank_rate_map(rank_rate_table: pd.DataFrame) -> Dict[int, float]:
    out = DEFAULT_RANK_IN3_RATE.copy()
    if rank_rate_table is None or rank_rate_table.empty:
        return out

    for _, row in rank_rate_table.iterrows():
        raw_rank = row.get("予想順位")
        if str(raw_rank) == "10位以下":
            bucket = 10
        else:
            bucket = _rank_bucket_value(raw_rank)
        rate = pd.to_numeric(pd.Series([row.get("馬券内率")]), errors="coerce").iloc[0]
        if pd.notna(rate):
            out[bucket] = float(rate)
    return out


def _score_rate_map(score_rate_table: pd.DataFrame) -> Dict[str, float]:
    out = DEFAULT_SCORE_IN3_RATE.copy()
    if score_rate_table is None or score_rate_table.empty:
        return out

    for _, row in score_rate_table.iterrows():
        band = str(row.get("score帯"))
        rate = pd.to_numeric(pd.Series([row.get("馬券内率")]), errors="coerce").iloc[0]
        if band in out and pd.notna(rate):
            out[band] = float(rate)
    return out


def _condition_correction(work: pd.DataFrame) -> pd.Series:
    """今回条件適性スコアがある場合だけ、軽い上下補正を入れる。"""
    c_cond = _pick_col(work, ["今回条件適性スコア", "条件適性スコア", "condition_fit_score"])
    if c_cond is None:
        return pd.Series(0.0, index=work.index)

    cond = _coerce_float_series(work[c_cond])
    valid = cond.dropna()
    if not valid.empty and (valid.between(0.0, 1.0).mean() >= 0.8):
        cond = cond * 100.0

    corr = pd.Series(0.0, index=work.index)
    corr = corr.mask(cond >= 70.0, 3.0)
    corr = corr.mask(cond <= 40.0, -3.0)
    return corr.astype(float)


def _odds_correction(work: pd.DataFrame) -> pd.Series:
    """人気を優先し、無ければ単勝オッズから軽い市場補正を入れる。"""
    pop = pd.to_numeric(work.get("人気", pd.Series(np.nan, index=work.index)), errors="coerce")
    odds = _coerce_float_series(work.get("単勝オッズ", pd.Series(np.nan, index=work.index)))

    corr = pd.Series(0.0, index=work.index)
    corr = corr.mask(pop.between(1, 3, inclusive="both"), 2.0)
    corr = corr.mask(pop.between(9, 12, inclusive="both"), -2.0)
    corr = corr.mask(pop >= 13, -4.0)

    pop_missing = pop.isna()
    corr = corr.mask(pop_missing & odds.notna() & (odds <= 5.0), 2.0)
    corr = corr.mask(pop_missing & odds.notna() & (odds > 20.0) & (odds <= 50.0), -2.0)
    corr = corr.mask(pop_missing & odds.notna() & (odds > 50.0), -4.0)
    return corr.astype(float)


def _hole_rescue_correction(work: pd.DataFrame) -> pd.Series:
    """
    穴馬救済スコアがあれば最大+5%で使う。
    無い場合は、人気薄なのに予想順位・scoreが悪くない馬だけ軽く救済する。
    """
    c_rescue = _pick_col(work, ["穴馬救済スコア", "穴馬スコア", "妙味スコア", "value_score"])
    if c_rescue is not None:
        rescue = _coerce_float_series(work[c_rescue]).fillna(0.0)
        valid = rescue.dropna()
        if not valid.empty and (valid.between(0.0, 1.0).mean() >= 0.8):
            rescue = rescue * 100.0
        return (rescue.clip(lower=0.0, upper=100.0) / 100.0 * 5.0).clip(0.0, 5.0)

    rank_num = pd.to_numeric(work.get("予想順位", pd.Series(np.nan, index=work.index)), errors="coerce")
    score = pd.to_numeric(work.get("score", pd.Series(np.nan, index=work.index)), errors="coerce")
    pop = pd.to_numeric(work.get("人気", pd.Series(np.nan, index=work.index)), errors="coerce")
    odds = _coerce_float_series(work.get("単勝オッズ", pd.Series(np.nan, index=work.index)))

    corr = pd.Series(0.0, index=work.index)
    corr = corr.mask((rank_num <= 5) & (pop >= 9), 2.0)
    corr = corr.mask((rank_num <= 5) & (odds >= 20.0), 2.0)
    corr = corr.mask((rank_num <= 8) & (score >= 70.0) & ((pop >= 9) | (odds >= 20.0)), 3.0)
    return corr.clip(0.0, 5.0).astype(float)


def _danger_correction(work: pd.DataFrame) -> pd.Series:
    """危険馬スコアがあれば使い、無ければ既存リスク列から最大-5%で減点する。"""
    c_danger = _pick_col(work, ["危険馬スコア", "danger_score"])
    if c_danger is not None:
        danger = _coerce_float_series(work[c_danger]).fillna(0.0)
        valid = danger.dropna()
        if not valid.empty and (valid.between(0.0, 1.0).mean() >= 0.8):
            danger = danger * 100.0
        return -(danger.clip(lower=0.0, upper=100.0) / 100.0 * 5.0).clip(0.0, 5.0)

    risk_cols = [
        _pick_col(work, ["extra_penalty"]),
        _pick_col(work, ["rest_dist_risk", "休養×距離差リスク"]),
        _pick_col(work, ["favorite_risk"]),
    ]
    risk_parts = []
    for col in risk_cols:
        if col is not None and col in work.columns:
            risk_parts.append(_coerce_float_series(work[col]).fillna(0.0))

    if not risk_parts:
        return pd.Series(0.0, index=work.index)

    risk_df = pd.concat(risk_parts, axis=1)
    risk_max = risk_df.max(axis=1).fillna(0.0)
    return -(risk_max * 2.0).clip(0.0, 5.0)


def _field_size_factor(field_size: pd.Series) -> pd.Series:
    """頭数による入りやすさを、後段の300%調整前に軽く反映する。"""
    field = pd.to_numeric(field_size, errors="coerce")
    factor = pd.Series(1.0, index=field.index)
    factor = factor.mask(field <= 8, 1.15)
    factor = factor.mask(field.between(9, 12, inclusive="both"), 1.05)
    factor = factor.mask(field >= 16, 0.92)
    return factor.fillna(1.0).astype(float)


def _normalize_in3_rates_to_race_total(raw_rates: pd.Series) -> pd.Series:
    """
    レース内の合計が通常300%に近づくように調整する。
    1%〜75%の上下限を守るため、数回だけ再スケールする。
    """
    rates = pd.to_numeric(raw_rates, errors="coerce").fillna(0.0).astype(float)
    if len(rates) == 0:
        return rates

    lower = 1.0
    upper = 75.0
    target_total = min(300.0, float(len(rates)) * upper)
    target_total = max(target_total, float(len(rates)) * lower)

    if rates.sum() <= 0:
        even = target_total / max(len(rates), 1)
        return pd.Series(even, index=rates.index).clip(lower, upper).round(2)

    def _scaled_total(scale: float) -> float:
        return float((rates * scale).clip(lower, upper).sum())

    low = 0.0
    high = 1.0
    while _scaled_total(high) < target_total and high < 1_000_000.0:
        high *= 2.0

    for _ in range(80):
        mid = (low + high) / 2.0
        if _scaled_total(mid) < target_total:
            low = mid
        else:
            high = mid

    adjusted = (rates * high).clip(lower, upper)
    rounded = adjusted.round(2)

    diff = round(float(target_total - rounded.sum()), 2)
    step = 0.01 if diff > 0 else -0.01
    steps = int(round(abs(diff) / 0.01))
    if steps > 0:
        if step > 0:
            priority = (adjusted - rounded).sort_values(ascending=False)
            candidates = [idx for idx in priority.index if rounded.loc[idx] < upper]
        else:
            priority = (rounded - adjusted).sort_values(ascending=False)
            candidates = [idx for idx in priority.index if rounded.loc[idx] > lower]

        if candidates:
            pos = 0
            for _ in range(steps):
                idx = candidates[pos % len(candidates)]
                if step > 0 and rounded.loc[idx] + step <= upper:
                    rounded.loc[idx] = round(float(rounded.loc[idx] + step), 2)
                elif step < 0 and rounded.loc[idx] + step >= lower:
                    rounded.loc[idx] = round(float(rounded.loc[idx] + step), 2)
                pos += 1

    return rounded.round(2)


def _market_odds_and_kind(work: pd.DataFrame) -> Tuple[pd.Series, pd.Series]:
    """市場馬券内率と期待値計算に使う複勝オッズを返す。"""
    fukusho = _coerce_float_series(work["複勝オッズ"]) if "複勝オッズ" in work.columns else pd.Series(np.nan, index=work.index)

    kind = pd.Series("未取得", index=work.index, dtype=object)
    kind = kind.mask(fukusho.notna(), "複勝オッズ")
    return fukusho, kind


def get_fukusho_odds_base_in3_rate(fukusho_odds: object) -> Optional[float]:
    """
    複勝オッズ帯から、一般的な基準馬券内率を返す。
    単位は %。最初の仮値なので、将来は過去データ由来の実績率へ置き換える。
    """
    if pd.isna(fukusho_odds):
        return None

    try:
        odds = float(fukusho_odds)
    except (TypeError, ValueError):
        return None

    if not np.isfinite(odds) or odds <= 0:
        return None
    if odds < 1.5:
        return 70.0
    if odds < 2.0:
        return 55.0
    if odds < 3.0:
        return 42.0
    if odds < 4.0:
        return 33.0
    if odds < 6.0:
        return 25.0
    if odds < 8.0:
        return 20.0
    if odds < 12.0:
        return 15.0
    if odds < 20.0:
        return 10.0
    return 6.0


def get_odds_blend_weight(fukusho_odds: object) -> float:
    """
    複勝オッズが高いほど、市場基準側を強める。
    戻り値は、モデル推定を何割使うかを表す係数。
    """
    if pd.isna(fukusho_odds):
        return 1.0

    try:
        odds = float(fukusho_odds)
    except (TypeError, ValueError):
        return 1.0

    if not np.isfinite(odds) or odds <= 0:
        return 1.0
    if odds < 2.0:
        return 1.00
    if odds < 4.0:
        return 0.85
    if odds < 8.0:
        return 0.70
    if odds < 12.0:
        return 0.55
    return 0.40


def adjust_in3_rate_by_fukusho_odds(row: pd.Series) -> object:
    """
    推定馬券内率を、複勝オッズ帯の基準馬券内率で下方補正する。
    row["推定馬券内率"] は % 表記、row["複勝オッズ"] は倍率を前提にする。
    """
    model_rate_raw = row.get("推定馬券内率")
    fukusho_odds = row.get("複勝オッズ")

    if pd.isna(model_rate_raw):
        return model_rate_raw

    try:
        model_rate = float(model_rate_raw)
    except (TypeError, ValueError):
        return model_rate_raw

    if not np.isfinite(model_rate):
        return model_rate_raw

    base_rate = get_fukusho_odds_base_in3_rate(fukusho_odds)
    if base_rate is None:
        return round(model_rate, 2)

    model_weight = get_odds_blend_weight(fukusho_odds)
    market_weight = 1.0 - model_weight

    # モデル推定とオッズ帯基準をブレンドし、高オッズの過大評価を抑える。
    adjusted = model_rate * model_weight + base_rate * market_weight

    # 基準馬券内率の1.5倍を上限にし、低オッズ馬だけは最低35%まで許容する。
    cap = max(base_rate * 1.5, 35.0)
    adjusted = min(adjusted, cap)

    # 最終的な安全範囲を 1%〜70% に収める。
    adjusted = max(1.0, min(adjusted, 70.0))
    return round(adjusted, 2)


def add_estimated_in3_rate(
    pred_df: pd.DataFrame,
    rank_rate_table: Optional[pd.DataFrame] = None,
    score_rate_table: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    今走予想DataFrameに、各馬の推定馬券内率・市場評価・期待値を追加する。
    """
    work = _canonical_prediction_frame(pred_df)

    rank_map = _rank_rate_map(rank_rate_table if rank_rate_table is not None else pd.DataFrame())
    score_map = _score_rate_map(score_rate_table if score_rate_table is not None else pd.DataFrame())

    work["_rank_bucket"] = work["予想順位"].map(_rank_bucket_value)
    work["score帯"] = work["score"].map(_score_band_label)
    work["順位別馬券内率"] = work["_rank_bucket"].map(lambda x: rank_map.get(int(x), DEFAULT_RANK_IN3_RATE[10]))
    work["score帯馬券内率"] = work["score帯"].map(lambda x: score_map.get(str(x), DEFAULT_SCORE_IN3_RATE["50未満"]))
    work["基本馬券内率"] = (work["順位別馬券内率"] * 0.6 + work["score帯馬券内率"] * 0.4).round(2)

    score_sorted = work.sort_values(["rid_str", "score", "馬番"], ascending=[True, False, True], kind="mergesort")
    top_score_map: Dict[str, List[float]] = {}
    for rid, sub in score_sorted.groupby("rid_str", sort=False):
        vals = pd.to_numeric(sub["score"], errors="coerce").dropna().head(2).astype(float).tolist()
        top_score_map[str(rid)] = vals

    def _gap12(rid: object) -> float:
        vals = top_score_map.get(str(rid), [])
        if len(vals) < 2:
            return 0.0
        return round(float(vals[0] - vals[1]), 2)

    work["gap12"] = work["rid_str"].map(_gap12)
    rank_num = pd.to_numeric(work["予想順位"], errors="coerce")
    gap = pd.to_numeric(work["gap12"], errors="coerce").fillna(0.0)
    score_gap_corr = pd.Series(0.0, index=work.index)
    score_gap_corr = score_gap_corr.mask((rank_num == 1) & (gap >= 8.0), 5.0)
    score_gap_corr = score_gap_corr.mask((rank_num == 1) & (gap >= 5.0) & (gap < 8.0), 3.0)
    score_gap_corr = score_gap_corr.mask((rank_num == 1) & (gap < 2.0), -3.0)
    work["score差補正"] = score_gap_corr.astype(float)

    work["頭数補正係数"] = _field_size_factor(work["頭数"]).round(3)
    work["条件適性補正"] = _condition_correction(work).round(2)
    work["オッズ補正"] = _odds_correction(work).round(2)
    work["穴馬救済補正"] = _hole_rescue_correction(work).round(2)
    work["危険馬補正"] = _danger_correction(work).round(2)

    additive_rate = (
        work["基本馬券内率"]
        + work["score差補正"]
        + work["条件適性補正"]
        + work["オッズ補正"]
        + work["穴馬救済補正"]
        + work["危険馬補正"]
    )
    work["補正後馬券内率"] = (additive_rate.clip(lower=0.1) * work["頭数補正係数"]).round(2)

    race_sum = work.groupby("rid_str")["補正後馬券内率"].transform("sum")
    work["レース内調整係数"] = np.where(race_sum > 0, 300.0 / race_sum, 1.0)
    work["レース内調整係数"] = pd.to_numeric(work["レース内調整係数"], errors="coerce").fillna(1.0).round(4)

    work["_raw_estimated_in3"] = work["補正後馬券内率"] * work["レース内調整係数"]
    work["推定馬券内率"] = work.groupby("rid_str")["_raw_estimated_in3"].transform(_normalize_in3_rates_to_race_total)

    market_odds, market_kind = _market_odds_and_kind(work)
    work["市場評価オッズ種別"] = market_kind
    work["市場馬券内率"] = np.where(market_odds > 0, (1.0 / market_odds) * 100.0, np.nan)
    work["市場馬券内率"] = pd.to_numeric(work["市場馬券内率"], errors="coerce").round(2)
    work["推定馬券内率_補正前"] = work["推定馬券内率"]
    work["オッズ帯基準馬券内率"] = market_odds.apply(get_fukusho_odds_base_in3_rate)
    work["オッズ帯補正係数"] = market_odds.apply(get_odds_blend_weight).round(2)
    work["推定馬券内率_オッズ補正後"] = work.apply(adjust_in3_rate_by_fukusho_odds, axis=1)
    work["期待値_補正前"] = ((work["推定馬券内率_補正前"] / 100.0) * market_odds).round(2)
    work["期待値"] = ((work["推定馬券内率_オッズ補正後"] / 100.0) * market_odds).round(2)

    ev = pd.to_numeric(work["期待値"], errors="coerce")
    est = pd.to_numeric(work["推定馬券内率_オッズ補正後"], errors="coerce")
    work["妙味判定"] = "オッズ未取得"
    work["妙味判定"] = work["妙味判定"].mask(ev.notna() & (est >= 45.0) & (ev < 1.0), "来そうだが妙味なし")
    work["妙味判定"] = work["妙味判定"].mask(ev.notna() & (ev < 0.95), "妙味なし")
    work["妙味判定"] = work["妙味判定"].mask(ev.notna() & (ev >= 0.95) & (ev < 1.15), "ほぼ妥当")
    work["妙味判定"] = work["妙味判定"].mask(ev.notna() & (ev >= 1.15) & (ev < 1.30), "穴候補")
    work["妙味判定"] = work["妙味判定"].mask(ev.notna() & (ev >= 1.30), "妙味あり")

    return work.drop(columns=["_rank_bucket", "_raw_estimated_in3"], errors="ignore")


def _merge_now_and_odds_for_estimation(target_df: pd.DataFrame, now_df: pd.DataFrame, raceday_str: Optional[str]) -> pd.DataFrame:
    """
    TARGETに、今走情報と単勝オッズを結合して推定馬券内率計算用DataFrameを作る。
    """
    target = _canonical_prediction_frame(target_df)
    now = _canonical_prediction_frame(now_df) if now_df is not None and not now_df.empty else pd.DataFrame()

    if not now.empty:
        info_cols = [
            "rid_str",
            "馬番",
            "レースID",
            "場所",
            "頭数",
            "人気",
            "単勝オッズ",
            "複勝オッズ",
            "今回条件適性スコア",
            "穴馬救済スコア",
            "危険馬スコア",
            "extra_penalty",
            "favorite_risk",
            "rest_dist_risk",
            "休養×距離差リスク",
        ]
        info_cols = [c for c in dict.fromkeys(info_cols) if c in now.columns]
        now_info = now[info_cols].drop_duplicates(subset=["rid_str", "馬番"], keep="first")
        target = pd.merge(target, now_info, on=["rid_str", "馬番"], how="left", suffixes=("", "_now"))

        for col in [
            "レースID",
            "頭数",
            "人気",
            "単勝オッズ",
            "複勝オッズ",
            "今回条件適性スコア",
            "穴馬救済スコア",
            "危険馬スコア",
            "extra_penalty",
            "favorite_risk",
            "rest_dist_risk",
            "休養×距離差リスク",
        ]:
            now_col = f"{col}_now"
            if now_col in target.columns:
                if col in target.columns:
                    if target[now_col].notna().any():
                        target[col] = target[col].combine_first(target[now_col])
                else:
                    target[col] = target[now_col]
                target = target.drop(columns=[now_col])

    try:
        odds_df = load_odds_csv(str(ODDS_CSV), raceday=raceday_str)
    except Exception as e:
        print(f"[WARN] オッズCSV読み込みに失敗したため、期待値計算の一部をスキップします: {e}")
        odds_df = pd.DataFrame(columns=["rid_str", "umaban", "tansho"])

    if odds_df is not None and not odds_df.empty and {"rid_str", "umaban", "tansho"}.issubset(odds_df.columns):
        odds_use = odds_df[["rid_str", "umaban", "tansho"]].copy()
        odds_use["rid_str"] = _normalize_rid_series(odds_use["rid_str"])
        odds_use["馬番"] = _normalize_umaban_series(odds_use["umaban"])
        odds_use["単勝オッズ_odds_csv"] = pd.to_numeric(odds_use["tansho"], errors="coerce")
        odds_use = odds_use.dropna(subset=["rid_str", "馬番", "単勝オッズ_odds_csv"])
        odds_use = odds_use[["rid_str", "馬番", "単勝オッズ_odds_csv"]].drop_duplicates(
            subset=["rid_str", "馬番"],
            keep="first",
        )

        target = pd.merge(target, odds_use, on=["rid_str", "馬番"], how="left")
        target["単勝オッズ"] = target["単勝オッズ_odds_csv"].combine_first(target["単勝オッズ"])
        target = target.drop(columns=["単勝オッズ_odds_csv"], errors="ignore")

    ozzu_path = _pick_ozzu_csv(str(ODDS_CSV), raceday_str)
    if ozzu_path and os.path.exists(ozzu_path) and "場所" in target.columns:
        try:
            ozzu_raw = _read_csv_any_encoding(ozzu_path)
            tansho_map = _build_tansho_map_from_ozzu(ozzu_raw)
            fukusho_map = _build_fukusho_map_from_ozzu(ozzu_raw)

            def _race_no_from_rid_for_estimation(rid_val: object) -> str:
                m = re.search(r"(\d{2})$", str(rid_val))
                return m.group(1) if m else ""

            def _umaban_for_estimation(umaban_val: object) -> Optional[int]:
                umaban_num = pd.to_numeric(pd.Series([umaban_val]), errors="coerce").iloc[0]
                if pd.isna(umaban_num):
                    return None
                return int(umaban_num)

            def _lookup_tansho_by_place(row: pd.Series) -> float:
                place_norm = _normalize_place(row.get("場所"))
                race_no = _race_no_from_rid_for_estimation(row.get("レースID", row.get("rid_str")))
                umaban = _umaban_for_estimation(row.get("馬番"))
                if not place_norm or not race_no or umaban is None:
                    return np.nan
                return float(tansho_map.get((place_norm, race_no, int(umaban)), np.nan))

            def _lookup_fukusho_by_place(row: pd.Series) -> float:
                place_norm = _normalize_place(row.get("場所"))
                race_no = _race_no_from_rid_for_estimation(row.get("レースID", row.get("rid_str")))
                umaban = _umaban_for_estimation(row.get("馬番"))
                if not place_norm or not race_no or umaban is None:
                    return np.nan
                return float(fukusho_map.get((place_norm, race_no, int(umaban)), np.nan))

            target["単勝オッズ_ozzu_place"] = target.apply(_lookup_tansho_by_place, axis=1)
            target["単勝オッズ"] = target["単勝オッズ_ozzu_place"].combine_first(target["単勝オッズ"])
            target["複勝オッズ_ozzu_place"] = target.apply(_lookup_fukusho_by_place, axis=1)
            target["複勝オッズ"] = target["複勝オッズ_ozzu_place"].combine_first(target["複勝オッズ"])
            target = target.drop(columns=["単勝オッズ_ozzu_place", "複勝オッズ_ozzu_place"], errors="ignore")
        except Exception as e:
            print(f"[WARN] OZZU場所+R番号照合によるオッズ補完に失敗しました: {e}")

    if "人気" in target.columns and "単勝オッズ" in target.columns:
        target["人気"] = pd.to_numeric(target["人気"], errors="coerce")
        target["単勝オッズ"] = _coerce_float_series(target["単勝オッズ"])
        odds_pop = target.groupby("rid_str")["単勝オッズ"].rank(ascending=True, method="min")
        target["人気"] = odds_pop.combine_first(target["人気"])
        target["人気"] = pd.to_numeric(target["人気"], errors="coerce").round().astype("Int64")

    return target


def _append_estimated_cols(base_df: pd.DataFrame, estimated_df: pd.DataFrame) -> pd.DataFrame:
    """既存シートへ推定馬券内率の主要列を戻す。"""
    base = _canonical_prediction_frame(base_df)
    add_cols = ["rid_str", "馬番"] + [c for c in EST_IN3_RESULT_COLS if c in estimated_df.columns and c != "馬番"]
    add_cols = list(dict.fromkeys(add_cols))

    base = base.drop(columns=[c for c in add_cols if c not in {"rid_str", "馬番"} and c in base.columns], errors="ignore")
    return pd.merge(
        base,
        estimated_df[add_cols].drop_duplicates(subset=["rid_str", "馬番"], keep="first"),
        on=["rid_str", "馬番"],
        how="left",
    )


def _risk_value_series(work: pd.DataFrame, candidates: List[str]) -> pd.Series:
    """候補列のどれかから、危険馬判定用の数値Seriesを取り出す。"""
    col = _pick_col(work, candidates)
    if col is None or col not in work.columns:
        return pd.Series(np.nan, index=work.index)
    return _coerce_float_series(work[col])


def _danger_horse_label(score: int) -> str:
    """危険馬スコアを、運用しやすい判定ラベルへ変換する。"""
    if score >= 7:
        return "強い切り候補"
    if score >= 5:
        return "切り候補"
    if score >= 3:
        return "3列目残し候補"
    return "切らない"


def _fmt_reason_num(value: object, digits: int = 1) -> str:
    """理由文に入れる数値を、欠損に強い短い表記へ寄せる。"""
    num = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.isna(num):
        return "不明"
    return f"{float(num):.{digits}f}"


def _build_danger_horse_sheet(estimated_df: pd.DataFrame) -> pd.DataFrame:
    """
    1〜3番人気の中で危険馬スコアが最も高い馬を、レースごとに1頭選ぶ。
    人気は単勝オッズをレース内で昇順rankし、OZZU由来のオッズ評価に寄せる。
    """
    if estimated_df is None or estimated_df.empty:
        return pd.DataFrame(columns=DANGER_HORSE_RESULT_COLS)

    work = _canonical_prediction_frame(estimated_df)
    work["score"] = pd.to_numeric(work["score"], errors="coerce")
    work["予想順位"] = pd.to_numeric(work["予想順位"], errors="coerce")
    work["単勝オッズ"] = _coerce_float_series(work["単勝オッズ"])
    work["複勝オッズ"] = _coerce_float_series(work["複勝オッズ"])
    work["_danger_popularity"] = work.groupby("rid_str")["単勝オッズ"].rank(ascending=True, method="min")
    work["_extra_penalty"] = _risk_value_series(work, ["extra_penalty"])
    work["_favorite_risk"] = _risk_value_series(work, ["favorite_risk"])
    work["_rest_dist_risk"] = _risk_value_series(work, ["rest_dist_risk", "休養×距離差リスク"])

    rows: List[Dict[str, object]] = []
    for rid, race in work.groupby("rid_str", sort=True):
        if not str(rid).strip():
            continue

        candidates = race[race["_danger_popularity"].between(1, 3, inclusive="both")].copy()
        if candidates.empty:
            continue

        top_score = pd.to_numeric(race["score"], errors="coerce").max()
        fav_score_min = pd.to_numeric(candidates["score"], errors="coerce").min()
        scored_rows: List[Tuple[Tuple[float, float, float, float, float], Dict[str, object]]] = []

        for _, row in candidates.iterrows():
            danger_score = 0
            reasons: List[str] = []

            pred_rank = row.get("予想順位")
            horse_score = row.get("score")
            pop = row.get("_danger_popularity")
            tansho = row.get("単勝オッズ")
            extra_penalty = row.get("_extra_penalty")
            favorite_risk = row.get("_favorite_risk")
            rest_dist_risk = row.get("_rest_dist_risk")

            pred_rank_num = pd.to_numeric(pd.Series([pred_rank]), errors="coerce").iloc[0]
            horse_score_num = pd.to_numeric(pd.Series([horse_score]), errors="coerce").iloc[0]
            pop_num = pd.to_numeric(pd.Series([pop]), errors="coerce").iloc[0]
            tansho_num = pd.to_numeric(pd.Series([tansho]), errors="coerce").iloc[0]
            score_gap = (
                float(top_score - horse_score_num)
                if pd.notna(top_score) and pd.notna(horse_score_num)
                else np.nan
            )

            if pd.notna(pred_rank_num) and pred_rank_num >= 4:
                danger_score += 2
                reasons.append(f"{int(pop_num)}番人気だが予想rank{int(pred_rank_num)}位")
            if pd.notna(pred_rank_num) and pred_rank_num >= 6:
                danger_score += 2
                reasons.append("予想rank6位以下")

            if pd.notna(horse_score_num) and pd.notna(fav_score_min) and horse_score_num == fav_score_min:
                danger_score += 2
                reasons.append("1〜3番人気内でscore最下位")

            if pd.notna(score_gap) and score_gap >= 5.0:
                danger_score += 1
                reasons.append(f"score1との差{_fmt_reason_num(score_gap)}")
            if pd.notna(score_gap) and score_gap >= 8.0:
                danger_score += 1
                reasons.append("score1との差8以上")

            if pd.notna(extra_penalty) and float(extra_penalty) >= 2.0:
                danger_score += 1
                reasons.append(f"extra_penalty{_fmt_reason_num(extra_penalty)}")
            if pd.notna(extra_penalty) and float(extra_penalty) >= 3.0:
                danger_score += 1
                reasons.append("extra_penalty3以上")

            if pd.notna(favorite_risk) and float(favorite_risk) >= 2.0:
                danger_score += 1
                reasons.append(f"favorite_risk{_fmt_reason_num(favorite_risk)}")
            if pd.notna(favorite_risk) and float(favorite_risk) >= 3.0:
                danger_score += 1
                reasons.append("favorite_risk3以上")

            if pd.notna(rest_dist_risk) and float(rest_dist_risk) >= 2.0:
                danger_score += 1
                reasons.append(f"rest_dist_risk{_fmt_reason_num(rest_dist_risk)}")
            if pd.notna(rest_dist_risk) and float(rest_dist_risk) >= 3.0:
                danger_score += 1
                reasons.append("rest_dist_risk3以上")

            mid_higher_count = 0
            if pd.notna(horse_score_num):
                mid = race[race["_danger_popularity"].between(4, 8, inclusive="both")]
                mid_higher_count = int((pd.to_numeric(mid["score"], errors="coerce") > horse_score_num).sum())
            if mid_higher_count >= 2:
                danger_score += 2
                reasons.append(f"4〜8番人気にscore上位馬{mid_higher_count}頭")

            if (
                pd.notna(tansho_num)
                and tansho_num <= 5.0
                and pd.notna(pred_rank_num)
                and pred_rank_num >= 5
            ):
                danger_score += 1
                reasons.append(f"単勝{_fmt_reason_num(tansho_num)}倍でrank5位以下")

            row_dict = {
                "レースID": row.get("レースID", rid),
                "場所": row.get("場所", pd.NA),
                "危険人気馬_馬番": row.get("馬番", pd.NA),
                "危険人気馬_馬名": row.get("馬名", pd.NA),
                "危険人気馬_人気": int(pop_num) if pd.notna(pop_num) else pd.NA,
                "危険人気馬_rank": int(pred_rank_num) if pd.notna(pred_rank_num) else pd.NA,
                "危険人気馬_score": round(float(horse_score_num), 2) if pd.notna(horse_score_num) else pd.NA,
                "危険人気馬スコア": int(danger_score),
                "危険人気馬_判定": _danger_horse_label(int(danger_score)),
                "危険人気馬_理由": "、".join(reasons) if reasons else "大きな危険条件なし",
                "危険人気馬_単勝オッズ": round(float(tansho_num), 2) if pd.notna(tansho_num) else pd.NA,
                "危険人気馬_複勝オッズ": round(float(row.get("複勝オッズ")), 2) if pd.notna(row.get("複勝オッズ")) else pd.NA,
                "危険人気馬_extra_penalty": round(float(extra_penalty), 2) if pd.notna(extra_penalty) else pd.NA,
                "危険人気馬_favorite_risk": round(float(favorite_risk), 2) if pd.notna(favorite_risk) else pd.NA,
                "危険人気馬_rest_dist_risk": round(float(rest_dist_risk), 2) if pd.notna(rest_dist_risk) else pd.NA,
                "危険人気馬_score1差": round(float(score_gap), 2) if pd.notna(score_gap) else pd.NA,
                "4_8番人気_score上位頭数": int(mid_higher_count),
                "レースscore1": round(float(top_score), 2) if pd.notna(top_score) else pd.NA,
                "頭数": row.get("頭数", len(race)),
            }
            scored_rows.append(
                (
                    (
                        -float(danger_score),
                        -float(score_gap) if pd.notna(score_gap) else 0.0,
                        -float(pred_rank_num) if pd.notna(pred_rank_num) else 0.0,
                        float(horse_score_num) if pd.notna(horse_score_num) else float("inf"),
                        float(pop_num) if pd.notna(pop_num) else float("inf"),
                    ),
                    row_dict,
                )
            )

        if scored_rows:
            scored_rows.sort(key=lambda x: x[0])
            rows.append(scored_rows[0][1])

    if not rows:
        return pd.DataFrame(columns=DANGER_HORSE_RESULT_COLS)
    return pd.DataFrame(rows, columns=DANGER_HORSE_RESULT_COLS)


def _add_estimated_in3_rate_to_excel(out_excel_path: str, raceday_str: Optional[str]) -> int:
    """
    最終出力Excelの既存シートへ、推定馬券内率の主要列を書き戻す。
    """
    if not os.path.exists(out_excel_path):
        print(f"[WARN] 出力Excelが見つからないため、推定馬券内率付与をスキップ: {out_excel_path}")
        return 0

    try:
        xls = pd.ExcelFile(out_excel_path, engine="openpyxl")
        if TARGET_SHEET not in xls.sheet_names:
            print(f"[WARN] '{TARGET_SHEET}' シートが無いため、推定馬券内率付与をスキップします")
            return 0

        target_df = pd.read_excel(out_excel_path, sheet_name=TARGET_SHEET, engine="openpyxl")
        now_df = (
            pd.read_excel(out_excel_path, sheet_name=NOW_SHEET, engine="openpyxl")
            if NOW_SHEET in xls.sheet_names
            else pd.DataFrame()
        )
    except Exception as e:
        print(f"[WARN] 推定馬券内率用のExcel読み込みに失敗しました: {e}")
        return 0

    history_df = _load_historical_prediction_result_df(exclude_paths=[Path(out_excel_path)])
    rank_rate_table = build_rank_rate_table(history_df)
    score_rate_table = build_score_rate_table(history_df)

    if history_df.empty:
        print("[WARN] 過去予想と実結果の照合が0件だったため、デフォルト馬券内率で推定します")
    else:
        print(f"[INFO] 推定馬券内率の過去集計件数: {len(history_df)}頭")

    current_df = _merge_now_and_odds_for_estimation(target_df, now_df, raceday_str)
    estimated_df = add_estimated_in3_rate(
        current_df,
        rank_rate_table=rank_rate_table,
        score_rate_table=score_rate_table,
    )

    target_aug = _append_estimated_cols(target_df, estimated_df)
    now_aug = _append_estimated_cols(now_df, estimated_df) if now_df is not None and not now_df.empty else now_df
    danger_horse_df = _build_danger_horse_sheet(estimated_df)
    estimated_count = int(len(estimated_df))
    danger_race_count = int(len(danger_horse_df))

    try:
        with pd.ExcelWriter(out_excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            # 旧バージョンで作成済みの不要シートが残らないよう削除する。
            for stale_sheet in [
                EST_IN3_SHEET,
                RANK_RATE_TABLE_SHEET,
                SCORE_RATE_TABLE_SHEET,
                VALUE_HORSE_SHEET,
                DANGER_HORSE_SHEET,
            ]:
                if stale_sheet in writer.book.sheetnames:
                    del writer.book[stale_sheet]

            target_aug.to_excel(writer, sheet_name=TARGET_SHEET, index=False)
            if now_aug is not None and not now_aug.empty:
                now_aug.to_excel(writer, sheet_name=NOW_SHEET, index=False)
            danger_horse_df.to_excel(writer, sheet_name=DANGER_HORSE_SHEET, index=False)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。Excelを閉じてから再実行してください: {out_excel_path}")
        return 0
    except Exception as e:
        print(f"[WARN] 推定馬券内率のExcel書き戻しに失敗しました: {e}")
        return 0

    print(f"[INFO] 推定馬券内率付与完了: {estimated_count}頭 -> {out_excel_path}")
    print(f"[INFO] '{DANGER_HORSE_SHEET}' シート作成完了: {danger_race_count}レース -> {out_excel_path}")
    return estimated_count


def _danger_horse_key_set(danger_df: pd.DataFrame) -> set[Tuple[str, int]]:
    """危険馬シートから、軸候補から除外する (レースID, 馬番) を作る。"""
    if danger_df is None or not isinstance(danger_df, pd.DataFrame) or danger_df.empty:
        return set()

    c_rid = _pick_col(danger_df, ["レースID", "rid_str", "race_id"])
    c_umaban = _pick_col(danger_df, ["危険人気馬_馬番", "危険馬_馬番", "馬番", "umaban"])
    if c_rid is None or c_umaban is None:
        return set()

    c_label = _pick_col(danger_df, ["危険人気馬_判定", "危険馬_判定", "判定"])
    c_score = _pick_col(danger_df, ["危険人気馬スコア", "危険馬スコア", "danger_score"])

    work = danger_df.copy()
    work["_axis_rid"] = _normalize_rid_series(work[c_rid])
    work["_axis_umaban"] = _normalize_umaban_series(work[c_umaban])
    work["_axis_label"] = work[c_label].astype(str).str.strip() if c_label is not None else ""
    work["_axis_score"] = pd.to_numeric(work[c_score], errors="coerce") if c_score is not None else np.nan

    no_danger_labels = {"", "-", "なし", "無し", "切らない", "大きな危険条件なし", "nan", "<NA>"}
    dangerous_keys: set[Tuple[str, int]] = set()
    for _, row in work.iterrows():
        rid = str(row.get("_axis_rid", "")).strip()
        umaban = row.get("_axis_umaban")
        if not rid or pd.isna(umaban):
            continue

        label = str(row.get("_axis_label", "")).strip()
        score = pd.to_numeric(pd.Series([row.get("_axis_score")]), errors="coerce").iloc[0]
        if pd.notna(score):
            if float(score) < 3.0:
                continue
            dangerous_keys.add((rid, int(umaban)))
            continue
        if label in no_danger_labels:
            continue

        dangerous_keys.add((rid, int(umaban)))

    return dangerous_keys


def _build_axis_umaban_map(target_df: pd.DataFrame, danger_df: pd.DataFrame) -> Dict[str, object]:
    """
    ranking 1〜3位から、指定条件を満たす軸馬番をレースごとに1頭選ぶ。
    条件未達のレースは "-" を返す。
    """
    if target_df is None or not isinstance(target_df, pd.DataFrame) or target_df.empty:
        return {}

    work = _canonical_prediction_frame(target_df)
    work["予想順位"] = pd.to_numeric(work["予想順位"], errors="coerce")
    work["単勝オッズ"] = _coerce_float_series(work["単勝オッズ"])
    work["score"] = pd.to_numeric(work.get("score", pd.Series(np.nan, index=work.index)), errors="coerce")

    c_extra = _pick_col(work, ["extra_penalty"])
    c_ev = _pick_col(work, ["期待値"])
    work["_axis_extra_penalty"] = (
        _coerce_float_series(work[c_extra]) if c_extra is not None else pd.Series(np.nan, index=work.index)
    )
    work["_axis_ev"] = _coerce_float_series(work[c_ev]) if c_ev is not None else pd.Series(np.nan, index=work.index)

    dangerous_keys = _danger_horse_key_set(danger_df)
    work["_axis_is_danger"] = [
        (str(rid), int(umaban)) in dangerous_keys if pd.notna(umaban) else True
        for rid, umaban in zip(work["rid_str"], work["馬番"])
    ]

    axis_map: Dict[str, object] = {}
    for rid, race in work.groupby("rid_str", sort=True):
        if not str(rid).strip():
            continue

        candidates = race[
            race["予想順位"].between(1, 3, inclusive="both")
            & race["単勝オッズ"].le(12.0)
            & race["_axis_extra_penalty"].le(1.5)
            & race["_axis_ev"].ge(0.80)
            & (~race["_axis_is_danger"])
        ].copy()

        if candidates.empty:
            axis_map[str(rid)] = "-"
            continue

        candidates["_axis_odds_tier"] = np.where(candidates["単勝オッズ"].le(8.0), 0, 1)
        candidates["_axis_extra_tier"] = np.where(candidates["_axis_extra_penalty"].le(0.5), 0, 1)
        candidates = candidates.sort_values(
            [
                "_axis_odds_tier",
                "_axis_extra_tier",
                "予想順位",
                "単勝オッズ",
                "_axis_extra_penalty",
                "score",
                "_axis_ev",
                "馬番",
            ],
            ascending=[True, True, True, True, True, False, False, True],
            kind="mergesort",
            na_position="last",
        )

        axis_umaban = candidates.iloc[0].get("馬番")
        axis_map[str(rid)] = int(axis_umaban) if pd.notna(axis_umaban) else "-"

    return axis_map


def _fill_axis_umaban_to_bet_sheet(out_excel_path: str) -> int:
    """
    「買い目_レース別1行」のU列に、新条件で選んだ軸馬番を書き戻す。
    旧1頭軸・保険BOX列は、この最終出力から除外する。
    """
    if not os.path.exists(out_excel_path):
        print(f"[WARN] 出力Excelが見つからないため、軸馬番反映をスキップ: {out_excel_path}")
        return 0

    try:
        with pd.ExcelFile(out_excel_path, engine="openpyxl") as xls:
            if BET_SHEET not in xls.sheet_names:
                print(f"[WARN] '{BET_SHEET}' シートが無いため、軸馬番反映をスキップします")
                return 0
            if TARGET_SHEET not in xls.sheet_names:
                print(f"[WARN] '{TARGET_SHEET}' シートが無いため、軸馬番反映をスキップします")
                return 0

            bet_df = pd.read_excel(out_excel_path, sheet_name=BET_SHEET, engine="openpyxl")
            target_df = pd.read_excel(out_excel_path, sheet_name=TARGET_SHEET, engine="openpyxl")
            danger_df = (
                pd.read_excel(out_excel_path, sheet_name=DANGER_HORSE_SHEET, engine="openpyxl")
                if DANGER_HORSE_SHEET in xls.sheet_names
                else pd.DataFrame()
            )
    except Exception as e:
        print(f"[WARN] 軸馬番反映用のExcel読み込みに失敗しました: {e}")
        return 0

    axis_map = _build_axis_umaban_map(target_df, danger_df)
    race_col = _pick_col(bet_df, ["レースID", "rid_str", "race_id"])
    if race_col is None:
        bet_df[AXIS_UMABAN_COLUMN] = "-"
    else:
        race_keys = _normalize_rid_series(bet_df[race_col])
        bet_df[AXIS_UMABAN_COLUMN] = race_keys.map(lambda x: axis_map.get(str(x), "-"))

    base_cols = [c for c in bet_df.columns if c not in OLD_AXIS_BET_COLUMNS and c != AXIS_UMABAN_COLUMN]
    insert_at = base_cols.index("理由") + 1 if "理由" in base_cols else min(20, len(base_cols))
    bet_cols = base_cols[:insert_at] + [AXIS_UMABAN_COLUMN] + base_cols[insert_at:]
    bet_df = bet_df[bet_cols]

    filled = int((bet_df[AXIS_UMABAN_COLUMN].astype(str).str.strip() != "-").sum())
    total = int(len(bet_df))

    try:
        with pd.ExcelWriter(out_excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            bet_df.to_excel(writer, sheet_name=BET_SHEET, index=False)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。Excelを閉じてから再実行してください: {out_excel_path}")
        return 0
    except Exception as e:
        print(f"[WARN] 軸馬番のExcel書き戻しに失敗しました: {e}")
        return 0

    print(f"[INFO] 軸馬番反映完了: {filled}/{total}レース -> {out_excel_path}")
    return filled


# ============================================================
# OZZU単勝オッズ反映
# ============================================================
def _pick_ozzu_csv(base: str, raceday: Optional[str]) -> Optional[str]:
    """
    ODDS_CSV（フォルダ or ファイル）から、使う OZZU CSV を1つ選ぶ。
    - base がファイルならそれを使う
    - base がフォルダでracedayがあるなら、racedayを含むファイルだけを使う
    - base がフォルダでracedayが無いなら、最新ファイルを使う
    """
    if os.path.isfile(base):
        return base
    if not os.path.isdir(base):
        return None

    csv_files = [os.path.join(base, f) for f in os.listdir(base) if f.lower().endswith(".csv")]
    if not csv_files:
        return None

    if raceday:
        preferred = [p for p in csv_files if str(raceday) in os.path.basename(p)]
        if preferred:
            return max(preferred, key=lambda p: os.path.getmtime(p))
        return None

    return max(csv_files, key=lambda p: os.path.getmtime(p))


def _require_target_ozzu_csv_for_raceday(raceday: Optional[str]) -> Optional[str]:
    """
    対象日指定時に、対象日のOZZU CSVが無ければ明示エラーで中止する。
    古い別日CSVを誤って使わないため、スクレイピングの自動実行はしない。
    """
    if not raceday:
        return None

    ozzu_path = _pick_ozzu_csv(str(ODDS_CSV), raceday)
    if ozzu_path and os.path.exists(ozzu_path):
        return ozzu_path

    raise FileNotFoundError(
        "対象日のオッズCSVが見つからないため処理を中止します。"
        f" 対象日={raceday}"
        f" / 探索先={ODDS_CSV}"
        f" / 先に {OZZU_SCRAPER_SCRIPT} を実行して、対象日のOZZU CSVを作成してください。"
    )


def _read_csv_any_encoding(path: str) -> pd.DataFrame:
    last_err: Optional[Exception] = None
    for enc in ("utf-8-sig", "utf-8", "cp932", "shift_jis", "utf-16"):
        try:
            return pd.read_csv(path, encoding=enc)
        except Exception as e:
            last_err = e
            continue
    if last_err:
        raise last_err
    raise RuntimeError("CSV 読み込みに失敗しました")


def _build_tansho_map_from_ozzu(ozzu_raw: pd.DataFrame) -> Dict[Tuple[str, str, int], float]:
    """
    OZZU CSV（racecourse/race/bet_type/combination/odds）から
    (place_norm, race_no, umaban) -> odds の辞書を作る
    """
    need_cols = {"racecourse", "race", "bet_type", "combination", "odds"}
    if not need_cols.issubset(set(ozzu_raw.columns)):
        missing = need_cols - set(ozzu_raw.columns)
        raise ValueError(f"OZZU CSV に必要列が不足しています: {missing}")

    ozzu = ozzu_raw.copy()
    ozzu = ozzu[ozzu["bet_type"].astype(str).str.contains("単勝", na=False)].copy()

    def _to_race_no(x: object) -> str:
        m = re.search(r"(\d+)", str(x))
        return m.group(1).zfill(2) if m else ""

    def _to_umaban(x: object) -> Optional[int]:
        m = re.search(r"(\d+)", str(x))
        return int(m.group(1)) if m else None

    def _to_odds(x: object) -> Optional[float]:
        s = str(x).replace(",", "")
        if not re.search(r"\d", s):
            return None
        try:
            return float(s)
        except Exception:
            return None

    ozzu["place_norm"] = ozzu["racecourse"].map(_normalize_place)
    ozzu["race_no"] = ozzu["race"].map(_to_race_no)
    ozzu["umaban"] = ozzu["combination"].map(_to_umaban)
    ozzu["tansho"] = ozzu["odds"].map(_to_odds)

    ozzu = ozzu.dropna(subset=["place_norm", "race_no", "umaban", "tansho"])
    out: Dict[Tuple[str, str, int], float] = {}
    for p, r, u, o in zip(ozzu["place_norm"], ozzu["race_no"], ozzu["umaban"], ozzu["tansho"]):
        out[(str(p), str(r), int(u))] = float(o)
    return out


def _build_fukusho_map_from_ozzu(ozzu_raw: pd.DataFrame) -> Dict[Tuple[str, str, int], float]:
    """
    OZZU CSV の複勝オッズから (place_norm, race_no, umaban) -> 複勝オッズ下限 の辞書を作る。
    複勝オッズが「1.3-2.3」のような範囲の場合、期待値は保守的に下限を使う。
    """
    need_cols = {"racecourse", "race", "bet_type", "combination", "odds"}
    if not need_cols.issubset(set(ozzu_raw.columns)):
        missing = need_cols - set(ozzu_raw.columns)
        raise ValueError(f"OZZU CSV に必要列が不足しています: {missing}")

    ozzu = ozzu_raw.copy()
    ozzu = ozzu[ozzu["bet_type"].astype(str).str.contains("複勝", na=False)].copy()

    def _to_race_no(x: object) -> str:
        m = re.search(r"(\d+)", str(x))
        return m.group(1).zfill(2) if m else ""

    def _to_umaban(x: object) -> Optional[int]:
        m = re.search(r"(\d+)", str(x))
        return int(m.group(1)) if m else None

    def _to_fukusho_lower(x: object) -> Optional[float]:
        s = str(x).replace(",", "")
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        if not m:
            return None
        try:
            return float(m.group(1))
        except Exception:
            return None

    ozzu["place_norm"] = ozzu["racecourse"].map(_normalize_place)
    ozzu["race_no"] = ozzu["race"].map(_to_race_no)
    ozzu["umaban"] = ozzu["combination"].map(_to_umaban)
    ozzu["fukusho"] = ozzu["odds"].map(_to_fukusho_lower)

    ozzu = ozzu.dropna(subset=["place_norm", "race_no", "umaban", "fukusho"])
    out: Dict[Tuple[str, str, int], float] = {}
    for p, r, u, o in zip(ozzu["place_norm"], ozzu["race_no"], ozzu["umaban"], ozzu["fukusho"]):
        out[(str(p), str(r), int(u))] = float(o)
    return out


def _fill_tansho_odds_to_bet_sheet(out_excel_path: str, raceday_str: Optional[str]) -> int:
    """
    出力Excelの「買い目_レース別1行」シートに、単勝オッズ_1位 を直接書き込む。
    """
    if not os.path.exists(out_excel_path):
        print(f"[WARN] 出力Excelが見つからないため、単勝オッズ反映をスキップ: {out_excel_path}")
        return 0

    ozzu_path = _pick_ozzu_csv(str(ODDS_CSV), raceday_str)
    if not ozzu_path or not os.path.exists(ozzu_path):
        print("[WARN] オッズCSVが見つからないため単勝オッズ付与をスキップします")
        return 0

    try:
        ozzu_raw = _read_csv_any_encoding(ozzu_path)
        tansho_map = _build_tansho_map_from_ozzu(ozzu_raw)
    except Exception as e:
        print(f"[WARN] OZZU CSV から単勝オッズマップ作成に失敗しました: {e}")
        return 0

    try:
        wb = load_workbook(out_excel_path)
    except Exception as e:
        print(f"[WARN] 出力Excelを開けませんでした: {e}")
        return 0

    sheet_name = "買い目_レース別1行"
    if sheet_name not in wb.sheetnames:
        print(f"[INFO] '{sheet_name}' シートが無いため単勝オッズ反映をスキップします")
        return 0

    ws = wb[sheet_name]

    header_to_col = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        header_to_col[str(v).strip()] = c

    need = ["場所", "レースID", "1位馬番"]
    for k in need:
        if k not in header_to_col:
            print(f"[WARN] '{sheet_name}' に必要列がありません: {k} -> 単勝オッズ反映をスキップ")
            return 0

    odds_col_name = "単勝オッズ_1位"
    if odds_col_name not in header_to_col:
        new_c = ws.max_column + 1
        ws.cell(row=1, column=new_c).value = odds_col_name
        header_to_col[odds_col_name] = new_c
        print(f"[INFO] '{sheet_name}' に '{odds_col_name}' 列が無かったため新規作成しました（列={new_c}）")

    col_place = header_to_col["場所"]
    col_rid = header_to_col["レースID"]
    col_umaban1 = header_to_col["1位馬番"]
    col_odds = header_to_col[odds_col_name]

    def _race_no_from_rid(rid_val: object) -> str:
        m = re.search(r"(\d{2})$", str(rid_val))
        return m.group(1) if m else ""

    def _to_int_safe(x: object) -> Optional[int]:
        m = re.search(r"(\d+)", str(x))
        return int(m.group(1)) if m else None

    filled = 0
    total = 0

    for r in range(2, ws.max_row + 1):
        rid = ws.cell(row=r, column=col_rid).value
        if rid is None:
            continue

        place = ws.cell(row=r, column=col_place).value
        umaban1 = ws.cell(row=r, column=col_umaban1).value

        total += 1

        place_norm = _normalize_place(place)
        race_no = _race_no_from_rid(rid)
        u1 = _to_int_safe(umaban1)
        if not place_norm or not race_no or u1 is None:
            continue

        odds_val = tansho_map.get((place_norm, race_no, int(u1)))
        if odds_val is None:
            continue

        ws.cell(row=r, column=col_odds).value = float(odds_val)
        filled += 1

    try:
        wb.save(out_excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。Excelを閉じてから再実行してください: {out_excel_path}")
        return 0

    print(f"[INFO] 単勝オッズ反映完了: {filled}/{total} -> {out_excel_path}")
    return filled


def _format_race_id_column_to_integer(out_excel_path: str) -> int:
    """
    指定シートのA列（レースID）を、Excelのユーザー定義書式「0」にする。
    """
    if not os.path.exists(out_excel_path):
        print(f"[WARN] 出力Excelが見つからないため、レースID列の書式設定をスキップ: {out_excel_path}")
        return 0

    wb = None
    formatted_sheets = 0
    try:
        wb = load_workbook(out_excel_path)
        for sheet_name in RACE_ID_FORMAT_SHEETS:
            if sheet_name not in wb.sheetnames:
                print(f"[INFO] '{sheet_name}' シートが無いため、レースID列の書式設定をスキップします")
                continue

            ws = wb[sheet_name]
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=1).number_format = "0"
            formatted_sheets += 1

        wb.save(out_excel_path)
    except PermissionError:
        print(f"[WARN] 出力Excelが開かれている可能性があります。Excelを閉じてから再実行してください: {out_excel_path}")
        return 0
    except Exception as e:
        print(f"[WARN] レースID列の書式設定に失敗しました: {e}")
        return 0
    finally:
        if wb is not None:
            wb.close()

    print(f"[INFO] レースID列の書式設定完了: {formatted_sheets}シート -> {out_excel_path}")
    return formatted_sheets


# ============================================================
# DL順位作成ユーティリティ
# ============================================================
def _normalize_colname(name: object) -> str:
    s = "" if name is None else str(name)
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", "", s)
    return s.lower()


def _pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    norm_map = {_normalize_colname(c): str(c) for c in df.columns}
    for cand in candidates:
        key = _normalize_colname(cand)
        if key in norm_map:
            return norm_map[key]
    return None


def _normalize_rid_series(s: pd.Series) -> pd.Series:
    def _normalize_one(v: object) -> str:
        if pd.isna(v):
            return ""
        text = str(v).strip()
        if text == "" or text.lower() == "nan" or text == "<NA>":
            return ""

        m = re.fullmatch(r"(\d+)(?:\.0+)?", text)
        if m:
            digits = m.group(1)
        else:
            num = pd.to_numeric(text, errors="coerce")
            if pd.notna(num) and float(num).is_integer():
                digits = str(int(num))
            else:
                digits = re.sub(r"\D", "", text)

        return digits[-12:] if len(digits) > 12 else digits

    return s.map(_normalize_one).astype(str)


def _normalize_umaban_series(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").astype("Int64")


def _parse_sex_age(val: object) -> Tuple[float, float]:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return np.nan, np.nan
    s = str(val).strip()
    if s == "":
        return np.nan, np.nan

    sex_map = {"牡": 0.0, "牝": 1.0, "セ": 2.0}
    sex_val = np.nan
    for k, v in sex_map.items():
        if s.startswith(k):
            sex_val = v
            break

    m = re.search(r"(\d+)", s)
    age_val = float(m.group(1)) if m else np.nan
    return sex_val, age_val


def _parse_body_weight(val: object) -> Tuple[float, float]:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return np.nan, np.nan
    s = str(val).strip()
    if s == "" or "計不" in s:
        return np.nan, np.nan

    m_w = re.search(r"(\d+)", s)
    weight_val = float(m_w.group(1)) if m_w else np.nan

    m_d = re.search(r"\(([-+]?\d+)\)", s)
    diff_val = float(m_d.group(1)) if m_d else np.nan
    return weight_val, diff_val


def _to_numeric_series(df: pd.DataFrame, col: Optional[str]) -> pd.Series:
    if col is None or col not in df.columns:
        return pd.Series([np.nan] * len(df))
    return pd.to_numeric(df[col], errors="coerce")


def _build_feature_df(df: pd.DataFrame, col_map: Dict[str, Optional[str]]) -> pd.DataFrame:
    pop = _to_numeric_series(df, col_map.get("popularity"))
    odds = _to_numeric_series(df, col_map.get("odds"))
    frame = _to_numeric_series(df, col_map.get("frame"))
    weight = _to_numeric_series(df, col_map.get("weight"))

    sex_age_col = col_map.get("sex_age")
    sex_age = df[sex_age_col] if sex_age_col in df.columns else pd.Series([np.nan] * len(df))
    sex_age_parsed = sex_age.map(_parse_sex_age)
    sex = sex_age_parsed.map(lambda x: x[0])
    age = sex_age_parsed.map(lambda x: x[1])

    bw_col = col_map.get("body_weight")
    bw = df[bw_col] if bw_col in df.columns else pd.Series([np.nan] * len(df))
    bw_parsed = bw.map(_parse_body_weight)
    body_weight = bw_parsed.map(lambda x: x[0])
    body_weight_diff = bw_parsed.map(lambda x: x[1])

    feat = pd.DataFrame(
        {
            "popularity": pop,
            "odds": odds,
            "frame": frame,
            "weight": weight,
            "age": age,
            "sex": sex,
            "body_weight": body_weight,
            "body_weight_diff": body_weight_diff,
        }
    )
    feat = feat.fillna(0.0)
    return feat


def _build_training_dataframe(path: Path = TRAIN_XLSX) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"学習用Excelが見つかりません: {path}")

    book = pd.read_excel(path, sheet_name=None, engine="openpyxl")
    rows: List[pd.DataFrame] = []

    for _, df in book.items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue

        c_name = _pick_col(df, COLS_CANDIDATES["horse_name"])
        c_finish = _pick_col(df, COLS_CANDIDATES["finish"])
        c_race_id = _pick_col(df, COLS_CANDIDATES["race_id"])

        if c_name is None or c_finish is None or c_race_id is None:
            continue

        name = df[c_name]
        finish = pd.to_numeric(df[c_finish], errors="coerce")
        race_id = _normalize_rid_series(df[c_race_id])
        race_id_ok = race_id.str.len() == 12

        mask = name.notna() & finish.notna() & race_id_ok
        if mask.sum() == 0:
            continue

        use_df = df.loc[mask].copy()
        finish = finish.loc[mask].reset_index(drop=True)

        col_map = {
            "popularity": _pick_col(use_df, COLS_CANDIDATES["popularity"]),
            "odds": _pick_col(use_df, COLS_CANDIDATES["odds"]),
            "frame": _pick_col(use_df, COLS_CANDIDATES["frame"]),
            "weight": _pick_col(use_df, COLS_CANDIDATES["weight"]),
            "sex_age": _pick_col(use_df, COLS_CANDIDATES["sex_age"]),
            "body_weight": _pick_col(use_df, COLS_CANDIDATES["body_weight"]),
        }

        feat = _build_feature_df(use_df, col_map)
        y = (finish <= 3).astype(int).reset_index(drop=True)

        rid_str = _normalize_rid_series(use_df[c_race_id]).reset_index(drop=True)
        umaban_col = _pick_col(use_df, COLS_CANDIDATES["umaban"])
        umaban = (
            _normalize_umaban_series(use_df[umaban_col]).reset_index(drop=True)
            if umaban_col
            else pd.Series([pd.NA] * len(use_df))
        )

        base = pd.DataFrame(
            {
                "rid_str": rid_str,
                "馬番": umaban,
                "馬名": use_df[c_name].astype(str).reset_index(drop=True),
                "y": y,
            }
        )

        train_df = pd.concat([base, feat.reset_index(drop=True)], axis=1)
        rows.append(train_df)

    if not rows:
        raise RuntimeError("学習データが作成できませんでした（有効な行がありません）")

    df_train = pd.concat(rows, ignore_index=True)
    for col in DL_FEATURE_COLS:
        if col in df_train.columns:
            df_train[col] = pd.to_numeric(df_train[col], errors="coerce").fillna(0.0)

    df_train["y"] = pd.to_numeric(df_train["y"], errors="coerce").fillna(0).astype(int)
    return df_train


class SimpleMLP(nn.Module):
    def __init__(self, input_dim: int) -> None:
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(input_dim, 64),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(64, 32),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(32, 1),
        )

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return self.net(x).squeeze(1)


def _train_model(X: np.ndarray, y: np.ndarray, seed: int = 42) -> Tuple[SimpleMLP, np.ndarray, np.ndarray]:
    rng = np.random.default_rng(seed)
    idx = rng.permutation(len(X))
    split = int(len(X) * 0.8)
    train_idx = idx[:split]
    val_idx = idx[split:]

    X_train = X[train_idx]
    y_train = y[train_idx]
    X_val = X[val_idx]
    y_val = y[val_idx]

    mean = X_train.mean(axis=0)
    std = X_train.std(axis=0)
    std = np.where(std == 0, 1.0, std)

    X_train = (X_train - mean) / std
    X_val = (X_val - mean) / std

    torch.manual_seed(seed)

    train_ds = TensorDataset(torch.from_numpy(X_train), torch.from_numpy(y_train))
    train_loader = DataLoader(train_ds, batch_size=256, shuffle=True)

    model = SimpleMLP(input_dim=X.shape[1])
    optimizer = torch.optim.Adam(model.parameters(), lr=1e-3)
    loss_fn = nn.BCEWithLogitsLoss()

    model.train()
    for epoch in range(20):
        epoch_loss = 0.0
        for xb, yb in train_loader:
            optimizer.zero_grad()
            logits = model(xb)
            loss = loss_fn(logits, yb)
            loss.backward()
            optimizer.step()
            epoch_loss += float(loss.item())

        if (epoch + 1) % 5 == 0:
            model.eval()
            with torch.no_grad():
                val_logits = model(torch.from_numpy(X_val))
                val_loss = loss_fn(val_logits, torch.from_numpy(y_val)).item()
            model.train()
            print(f"[INFO] epoch={epoch+1} train_loss={epoch_loss:.4f} val_loss={val_loss:.4f}")

    return model, mean, std


def _load_now_data(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"予測用Excelが見つかりません: {path}")
    book = pd.read_excel(path, sheet_name=None, engine="openpyxl")
    if NOW_SHEET not in book:
        raise RuntimeError(f"今走シートが見つかりません: sheet={NOW_SHEET} path={path}")
    return book[NOW_SHEET].copy()


def _ensure_rid_umaban(df: pd.DataFrame) -> pd.DataFrame:
    if "rid_str" in df.columns:
        df["rid_str"] = _normalize_rid_series(df["rid_str"])
    else:
        c_race_id = _pick_col(df, COLS_CANDIDATES["race_id"])
        if c_race_id is None:
            raise RuntimeError("rid_str 列が無く、レースID列も見つかりません")
        df["rid_str"] = _normalize_rid_series(df[c_race_id])

    c_umaban = _pick_col(df, COLS_CANDIDATES["umaban"])
    if c_umaban is None:
        raise RuntimeError("馬番列が見つかりません")
    df["馬番"] = _normalize_umaban_series(df[c_umaban])
    return df


def _predict_dl_rank(model: SimpleMLP, mean: np.ndarray, std: np.ndarray, now_df: pd.DataFrame) -> pd.DataFrame:
    """
    今回の修正ポイント
    - dl_rank だけでなく dl_prob も返す
    """
    col_map = {
        "popularity": _pick_col(now_df, COLS_CANDIDATES["popularity"]),
        "odds": _pick_col(now_df, COLS_CANDIDATES["odds"] + ["tansho"]),
        "frame": _pick_col(now_df, COLS_CANDIDATES["frame"]),
        "weight": _pick_col(now_df, COLS_CANDIDATES["weight"]),
        "sex_age": _pick_col(now_df, COLS_CANDIDATES["sex_age"]),
        "body_weight": _pick_col(now_df, COLS_CANDIDATES["body_weight"]),
    }

    feat = _build_feature_df(now_df, col_map)
    X = feat[DL_FEATURE_COLS].to_numpy(dtype=np.float32)
    X = (X - mean) / std

    model.eval()
    with torch.no_grad():
        logits = model(torch.from_numpy(X))
        prob = torch.sigmoid(logits).cpu().numpy()

    out = now_df.copy()
    out = _ensure_rid_umaban(out)
    out["dl_prob"] = prob.astype(float)
    out["dl_rank"] = out.groupby("rid_str")["dl_prob"].rank(ascending=False, method="first").astype("Int64")

    return out[["rid_str", "馬番", "dl_prob", "dl_rank"]]


def _create_dl_rank_df(pred_path: Path) -> pd.DataFrame:
    """
    1回目の with_feat Excel を読み、
    dl_rank / dl_prob を自動作成してDataFrameで返す。
    """
    print("[INFO] DL用の学習データ読み込み中...")
    df_train = _build_training_dataframe(TRAIN_XLSX)
    X = df_train[DL_FEATURE_COLS].to_numpy(dtype=np.float32)
    y = df_train["y"].to_numpy(dtype=np.float32)

    if len(X) == 0:
        raise RuntimeError("学習データが空です")

    print("[INFO] DLモデル学習中...")
    model, mean, std = _train_model(X, y)

    print("[INFO] DL順位を予測中...")
    now_df = _load_now_data(pred_path)
    dl_rank_df = _predict_dl_rank(model, mean, std, now_df)

    print(f"[INFO] DL順位データ作成完了: {len(dl_rank_df)}行")
    return dl_rank_df


# ============================================================
# main
# ============================================================
def main() -> None:
    raceday_str = input("対象レース日付を YYYYMMDD 形式で入力してください（空Enterなら全日対象）: ").strip()
    if raceday_str == "":
        raceday_str = None
    elif not re.fullmatch(r"\d{8}", raceday_str):
        raise ValueError("対象レース日付は YYYYMMDD の8桁で入力してください")

    _require_target_ozzu_csv_for_raceday(raceday_str)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # --------------------------------------------------------
    # 1回目: 元データ -> with_feat
    # --------------------------------------------------------
    src_excel_first, out_excel_first = _resolve_raw_src_out_paths(raceday_str)

    if not os.path.exists(src_excel_first):
        raise FileNotFoundError(f"入力ファイルが見つかりません: {src_excel_first}")

    print("[INFO] ===== 1回目の予想処理を開始します =====")
    run_pipeline(
        SRC_EXCEL=src_excel_first,
        OUT_EXCEL=out_excel_first,
        LEVELS_XL=RACE_LEVEL_XLSX,
        BASE_TIME=BASE_TIME_XLSX,
        ODDS_CSV_PATH=ODDS_CSV,
        RACEDAY=raceday_str,
    )

    actual_first_out = _pick_actual_out_excel(out_excel_first)
    if not os.path.exists(actual_first_out):
        raise FileNotFoundError(f"1回目の出力Excelが見つかりません: {actual_first_out}")

    # --------------------------------------------------------
    # DL順位作成: with_feat -> メモリ上のDL順位データ
    # --------------------------------------------------------
    print("[INFO] ===== DL順位作成を開始します =====")
    dl_rank_df = _create_dl_rank_df(Path(actual_first_out))

    # --------------------------------------------------------
    # 2回目: 1回目のwith_feat + メモリ上のDL順位データ -> 最終 with_feat
    # --------------------------------------------------------
    print("[INFO] ===== 2回目の予想処理を開始します =====")
    run_pipeline(
        SRC_EXCEL=actual_first_out,
        OUT_EXCEL=actual_first_out,
        LEVELS_XL=RACE_LEVEL_XLSX,
        BASE_TIME=BASE_TIME_XLSX,
        ODDS_CSV_PATH=ODDS_CSV,
        RACEDAY=raceday_str,
        DL_RANK_DF=dl_rank_df,
    )

    actual_final_out = _pick_actual_out_excel(actual_first_out)
    if not os.path.exists(actual_final_out):
        raise FileNotFoundError(f"最終出力Excelが見つかりません: {actual_final_out}")

    _fill_tansho_odds_to_bet_sheet(actual_final_out, raceday_str)
    _add_estimated_in3_rate_to_excel(actual_final_out, raceday_str)
    _fill_axis_umaban_to_bet_sheet(actual_final_out)
    append_roi_focus_bet_sheet_to_excel(actual_final_out)
    _format_race_id_column_to_integer(actual_final_out)

    print("[INFO] ===== すべて完了しました =====")
    print(f"[INFO] 最終出力: {actual_final_out}")


if __name__ == "__main__":
    main()
