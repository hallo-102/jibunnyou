# -*- coding: utf-8 -*-
"""最終出力Excelへ、既存予想から独立したレースペース予想シートを追加する。"""

from __future__ import annotations

import gc
import math
import os
import re
import tempfile
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from statistics import pstdev
from typing import Dict, List, Mapping, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# ============================================================
# ペース予想設定値
# ============================================================
NOW_SHEET_NAME = "今走レース情報"
TARGET_SHEET_NAME = "TARGET"
BET_SHEET_NAME = "買い目_レース別1行"
DEFAULT_PACE_SHEET_NAME = "ペース予想"
RECENCY_WEIGHTS: Tuple[float, ...] = (1.00, 0.85, 0.70, 0.55, 0.40)
MAX_PAST_RUNS = 5
MIN_LAP_SAMPLE_COUNT = 3
LATE_START_WORDS: Tuple[str, ...] = ("出遅れ", "出脚鈍い", "アオル", "立遅れ")

COURSE_MATCH_WEIGHTS: Mapping[Tuple[str, str], float] = {
    ("芝", "芝"): 1.00,
    ("ダート", "ダート"): 1.00,
    ("障害", "障害"): 1.00,
    ("芝", "ダート"): 0.25,
    ("ダート", "芝"): 0.25,
}

DISTANCE_DIFF_WEIGHTS: Tuple[Tuple[int, float], ...] = (
    (200, 1.00),
    (400, 0.80),
    (600, 0.55),
    (10_000, 0.30),
)

# 前半3Fの速度指数化に使う緩やかな基準値。最終区分はこの値だけでは決めない。
FIRST_3F_BASELINES: Mapping[Tuple[str, str], float] = {
    ("芝", "短距離"): 34.5,
    ("芝", "マイル"): 35.5,
    ("芝", "中距離"): 36.5,
    ("芝", "長距離"): 37.5,
    ("ダート", "短距離"): 35.0,
    ("ダート", "マイル"): 36.0,
    ("ダート", "中距離"): 36.8,
    ("ダート", "長距離"): 37.8,
    ("障害", "長距離"): 39.0,
}

# 0～100の総合ペーススコアを5区分へ分ける境界値。
# 芝・ダート・距離帯ごとに後から調整できる構造にしている。
PACE_CLASS_THRESHOLDS: Mapping[Tuple[str, str], Tuple[float, float, float, float]] = {
    ("芝", "短距離"): (20.0, 38.0, 59.0, 78.0),
    ("芝", "マイル"): (22.0, 40.0, 61.0, 80.0),
    ("芝", "中距離"): (24.0, 42.0, 63.0, 82.0),
    ("芝", "長距離"): (26.0, 44.0, 65.0, 84.0),
    ("ダート", "短距離"): (20.0, 38.0, 60.0, 79.0),
    ("ダート", "マイル"): (22.0, 40.0, 62.0, 81.0),
    ("ダート", "中距離"): (24.0, 42.0, 64.0, 83.0),
    ("ダート", "長距離"): (26.0, 44.0, 66.0, 85.0),
    ("障害", "長距離"): (28.0, 46.0, 68.0, 86.0),
}

PACE_OUTPUT_COLUMNS: Tuple[str, ...] = (
    "レースID",
    "発走時刻",
    "場所",
    "レース名",
    "コース",
    "馬場",
    "頭数",
    "クラス",
    "予想ペース",
    "ペース記号",
    "予想前半3F",
    "予想後半3F",
    "前後半差",
    "ペース圧力指数",
    "逃げ競合指数",
    "逃げ候補数",
    "先行候補数",
    "本命逃げ候補_馬番",
    "本命逃げ候補_馬名",
    "本命逃げ候補_逃げ指数",
    "対抗逃げ候補_馬番",
    "対抗逃げ候補_馬名",
    "対抗逃げ候補_逃げ指数",
    "逃げ候補一覧",
    "先行候補一覧",
    "単騎逃げ可能性",
    "展開有利脚質",
    "展開不利脚質",
    "予想信頼度",
    "データ充足率",
    "判定理由",
    "注意事項",
)


@dataclass(frozen=True)
class HorsePaceProfile:
    """1頭分の直近走から計算したペース関連プロフィール。"""

    horse_number: object
    horse_name: str
    running_style: str
    escape_index: float
    front_index: float
    early_speed_index: float
    style_stability: float
    first_place_rate: float
    top3_rate: float
    average_position_ratio: Optional[float]
    used_run_count: int
    surface_match_rate: float
    near_distance_rate: float
    passing_coverage_rate: float
    pace_coverage_rate: float
    lap_coverage_rate: float
    late_start_count: int
    large_distance_change_count: int


class PacePredictionError(RuntimeError):
    """ペース予想の作成・検証に失敗した場合の例外。"""


def normalize_column_name(value: object) -> str:
    """列名をNFKC化し、半角・全角空白と記号差を吸収して返す。"""
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).strip().lower()
    return re.sub(r"[\s_\-]+", "", text)


def parse_course(value: object) -> str:
    """コース表記から芝・ダート・障害・不明のいずれかを返す。"""
    text = unicodedata.normalize("NFKC", str(value or "")).strip()
    if "障" in text:
        return "障害"
    if "ダ" in text:
        return "ダート"
    if "芝" in text:
        return "芝"
    return "不明"


def parse_distance(value: object) -> Optional[int]:
    """数値、芝1600、ダ1700などから距離メートルを取得する。"""
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        number = int(round(float(value)))
        return number if 400 <= number <= 5000 else None
    text = unicodedata.normalize("NFKC", str(value))
    matches = re.findall(r"(?<!\d)(\d{3,4})(?:\s*m)?", text, flags=re.IGNORECASE)
    for matched in matches:
        number = int(matched)
        if 400 <= number <= 5000:
            return number
    return None


def parse_pace_value(value: object) -> Tuple[Optional[float], Optional[float]]:
    """前半3F-後半3F形式のペース値を安全に解析する。"""
    if value is None:
        return None, None
    text = unicodedata.normalize("NFKC", str(value)).replace("−", "-").replace("ー", "-")
    matched = re.search(r"(\d{2}(?:\.\d+)?)\s*-\s*(\d{2}(?:\.\d+)?)", text)
    if not matched:
        return None, None
    first, last = float(matched.group(1)), float(matched.group(2))
    if not (25.0 <= first <= 50.0 and 25.0 <= last <= 50.0):
        return None, None
    return first, last


def parse_lap_times(value: object) -> List[float]:
    """ラップタイム文字列から妥当な1Fラップだけを抽出する。"""
    if value is None:
        return []
    text = unicodedata.normalize("NFKC", str(value))
    laps = []
    for token in re.findall(r"\d{1,2}(?:\.\d+)?", text):
        lap = float(token)
        if 8.0 <= lap <= 25.0:
            laps.append(lap)
    return laps


def parse_passing_position(value: object) -> List[int]:
    """通過順から正の整数順位を順番どおり取得する。"""
    if value is None:
        return []
    text = unicodedata.normalize("NFKC", str(value))
    positions = [int(token) for token in re.findall(r"\d+", text)]
    return [position for position in positions if position > 0]


def classify_running_style(
    first_positions: Sequence[int],
    position_ratios: Sequence[float],
    weights: Sequence[float],
) -> str:
    """固定順位と頭数比率を併用し、逃げ・先行・差し・追込へ分類する。"""
    if not first_positions or not position_ratios or not weights:
        return "不明"
    total_weight = sum(weights)
    if total_weight <= 0:
        return "不明"
    first_rate = sum(weight for pos, weight in zip(first_positions, weights) if pos == 1) / total_weight
    top4_rate = sum(weight for pos, weight in zip(first_positions, weights) if pos <= 4) / total_weight
    avg_ratio = sum(ratio * weight for ratio, weight in zip(position_ratios, weights)) / total_weight
    if first_rate >= 0.34 or (first_rate >= 0.20 and avg_ratio <= 0.16):
        return "逃げ"
    if top4_rate >= 0.55 or avg_ratio <= 0.30:
        return "先行"
    if avg_ratio >= 0.75:
        return "追込"
    return "差し"


def _distance_band(distance: Optional[int]) -> str:
    """距離をペース閾値用の4帯へ分ける。"""
    if distance is None:
        return "マイル"
    if distance <= 1400:
        return "短距離"
    if distance <= 1800:
        return "マイル"
    if distance <= 2400:
        return "中距離"
    return "長距離"


def _clip(value: float, lower: float = 0.0, upper: float = 100.0) -> float:
    """数値を指定範囲へ収める。"""
    return max(lower, min(upper, value))


def _to_float(value: object) -> Optional[float]:
    """Excel由来の値を有限のfloatへ変換する。"""
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def _to_positive_int(value: object) -> Optional[int]:
    """頭数や馬番を正の整数へ変換する。"""
    number = _to_float(value)
    if number is None:
        return None
    integer = int(round(number))
    return integer if integer > 0 else None


def _normalize_race_id(value: object) -> Optional[str]:
    """数値・小数・文字列・指数表記のレースIDを12桁文字列へ正規化する。"""
    if value is None:
        return None
    text = unicodedata.normalize("NFKC", str(value)).strip()
    if not text:
        return None
    try:
        race_id = format(Decimal(text).quantize(Decimal("1")), "f")
    except (InvalidOperation, ValueError):
        digits = re.sub(r"\D", "", text)
        race_id = digits
    if race_id.endswith(".0"):
        race_id = race_id[:-2]
    if not race_id.isdigit() or len(race_id) > 12:
        return None
    return race_id.zfill(12)


def _parse_date(value: object) -> datetime:
    """過去走日付を解析し、失敗時は十分古い日付を返す。"""
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return datetime(1900, 1, 1)
    return parsed.to_pydatetime()


def _record_value(record: Mapping[str, object], *candidates: str) -> object:
    """正規化済み辞書から最初に見つかった候補列の値を返す。"""
    for candidate in candidates:
        key = normalize_column_name(candidate)
        if key in record:
            return record[key]
    return None


def _worksheet_records(ws: Worksheet) -> List[Dict[str, object]]:
    """ワークシートを正規化列名のレコード配列へ変換する。"""
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [normalize_column_name(value) for value in next(rows)]
    except StopIteration:
        return []
    records: List[Dict[str, object]] = []
    for values in rows:
        if not any(value is not None for value in values):
            continue
        records.append({header: value for header, value in zip(headers, values) if header})
    return records


def _course_similarity(current_surface: str, past_surface: str) -> float:
    """今回と過去の芝・ダート・障害条件の類似度を返す。"""
    if current_surface == "障害" or past_surface == "障害":
        return 1.00 if current_surface == past_surface else 0.05
    if current_surface == "不明" or past_surface == "不明":
        return 0.55
    return COURSE_MATCH_WEIGHTS.get((current_surface, past_surface), 0.25)


def _distance_similarity(current_distance: Optional[int], past_distance: Optional[int]) -> float:
    """今回と過去の距離差から類似度を返す。"""
    if current_distance is None or past_distance is None:
        return 0.55
    difference = abs(current_distance - past_distance)
    for upper, weight in DISTANCE_DIFF_WEIGHTS:
        if difference <= upper:
            return weight
    return 0.30


def _early_speed_index(first_3f: Optional[float], surface: str, distance: Optional[int]) -> float:
    """前半3Fを条件別基準と比較し、0～100の速度指数へ変換する。"""
    if first_3f is None:
        return 50.0
    baseline = FIRST_3F_BASELINES.get((surface, _distance_band(distance)), 36.0)
    return _clip(50.0 + (baseline - first_3f) * 7.0)


def _extract_early_late_3f(record: Mapping[str, object]) -> Tuple[Optional[float], Optional[float], bool]:
    """ペース列とラップ列を照合し、前後半3Fとラップ有無を返す。"""
    pace_first, pace_last = parse_pace_value(_record_value(record, "ペース"))
    laps = parse_lap_times(_record_value(record, "ラップタイム"))
    lap_first = round(sum(laps[:3]), 3) if len(laps) >= 3 else None
    lap_last = round(sum(laps[-3:]), 3) if len(laps) >= 6 else None

    # 両方が妥当で近い場合だけ平均し、矛盾が大きい場合は明示されたペース列を優先する。
    first = pace_first
    if pace_first is None:
        first = lap_first
    elif lap_first is not None and abs(pace_first - lap_first) <= 2.5:
        first = (pace_first + lap_first) / 2.0

    last = pace_last
    if pace_last is None:
        last = lap_last
    elif lap_last is not None and abs(pace_last - lap_last) <= 2.5:
        last = (pace_last + lap_last) / 2.0
    return first, last, bool(laps)


def calculate_horse_pace_profile(
    horse_number: object,
    horse_name: str,
    past_records: Sequence[Mapping[str, object]],
    current_course: object,
    current_distance: Optional[int],
) -> HorsePaceProfile:
    """1頭の直近5走から脚質と3種類の指数を計算する。"""
    current_surface = parse_course(current_course)
    unique_records: Dict[str, Mapping[str, object]] = {}
    for index, record in enumerate(past_records):
        race_id = _normalize_race_id(_record_value(record, "race_id", "レースID"))
        fallback = f"{_record_value(record, '日付')}|{_record_value(record, '開催')}|{index}"
        unique_records.setdefault(race_id or fallback, record)

    recent_records = sorted(
        unique_records.values(),
        key=lambda row: _parse_date(_record_value(row, "日付", "開催日")),
        reverse=True,
    )[:MAX_PAST_RUNS]

    first_positions: List[int] = []
    position_ratios: List[float] = []
    style_weights: List[float] = []
    escape_scores: List[Tuple[float, float]] = []
    front_scores: List[Tuple[float, float]] = []
    speed_scores: List[Tuple[float, float]] = []
    surface_matches = 0.0
    near_distances = 0.0
    passing_count = 0
    pace_count = 0
    lap_count = 0
    late_start_count = 0
    large_distance_changes = 0

    for recency_index, record in enumerate(recent_records):
        recency_weight = RECENCY_WEIGHTS[recency_index]
        past_course_value = _record_value(record, "距離", "コース")
        past_surface = parse_course(past_course_value)
        past_distance = parse_distance(past_course_value)
        field_size = _to_positive_int(_record_value(record, "頭数", "頭 数"))
        positions = parse_passing_position(_record_value(record, "通過"))
        first_3f, _, has_laps = _extract_early_late_3f(record)
        if first_3f is not None:
            pace_count += 1
        if has_laps:
            lap_count += 1

        surface_weight = _course_similarity(current_surface, past_surface)
        distance_weight = _distance_similarity(current_distance, past_distance)
        similarity_weight = recency_weight * surface_weight * distance_weight
        surface_matches += 1.0 if surface_weight >= 0.99 else 0.0
        near_distances += 1.0 if distance_weight >= 0.80 else 0.0

        distance_adjustment = 0.0
        if current_distance is not None and past_distance is not None:
            distance_change = current_distance - past_distance
            if abs(distance_change) >= 600:
                large_distance_changes += 1
            if distance_change <= -200:
                distance_adjustment = 4.0
            elif distance_change >= 200:
                distance_adjustment = -3.0

        note = unicodedata.normalize("NFKC", str(_record_value(record, "備考") or ""))
        late_start = any(word in note for word in LATE_START_WORDS)
        if late_start:
            late_start_count += 1

        speed_index = _early_speed_index(first_3f, past_surface, past_distance)
        speed_scores.append((speed_index + distance_adjustment, max(similarity_weight, 0.05)))

        if not positions or field_size is None:
            continue
        passing_count += 1
        first_position = positions[0]
        early_positions = positions[: max(1, math.ceil(len(positions) / 2))]
        early_average = sum(early_positions) / len(early_positions)
        first_ratio = _clip(first_position / field_size, 0.0, 1.5)
        average_ratio = _clip(early_average / field_size, 0.0, 1.5)
        position_weight = similarity_weight

        # 出遅れて後方だった走は、本来の脚質を断定する材料として弱く扱う。
        if late_start and first_ratio >= 0.35:
            position_weight *= 0.40

        escape_component = 100.0 if first_position == 1 else _clip(96.0 - first_ratio * 180.0)
        escape_component = 0.75 * escape_component + 0.25 * _clip(100.0 - average_ratio * 135.0)
        front_component = _clip(100.0 - first_ratio * 105.0)
        if first_position <= 4:
            front_component = max(front_component, 72.0 - (first_position - 1) * 8.0)

        first_positions.append(first_position)
        position_ratios.append(first_ratio)
        style_weights.append(max(position_weight, 0.03))
        escape_scores.append((escape_component + distance_adjustment, max(position_weight, 0.03)))
        front_scores.append((front_component + distance_adjustment * 0.5, max(position_weight, 0.03)))

    used_count = len(recent_records)

    def weighted_average(values: Sequence[Tuple[float, float]], default: float) -> float:
        total = sum(weight for _, weight in values)
        if total <= 0:
            return default
        return sum(value * weight for value, weight in values) / total

    total_style_weight = sum(style_weights)
    first_place_rate = (
        sum(weight for pos, weight in zip(first_positions, style_weights) if pos == 1) / total_style_weight
        if total_style_weight > 0
        else 0.0
    )
    top3_rate = (
        sum(weight for pos, weight in zip(first_positions, style_weights) if pos <= 3) / total_style_weight
        if total_style_weight > 0
        else 0.0
    )
    avg_ratio = (
        sum(ratio * weight for ratio, weight in zip(position_ratios, style_weights)) / total_style_weight
        if total_style_weight > 0
        else None
    )
    position_escape = weighted_average(escape_scores, 25.0)
    position_front = weighted_average(front_scores, 30.0)
    early_speed = _clip(weighted_average(speed_scores, 50.0))
    escape_index = _clip(position_escape * 0.58 + first_place_rate * 25.0 + top3_rate * 8.0 + early_speed * 0.09)
    front_index = _clip(position_front * 0.58 + top3_rate * 25.0 + early_speed * 0.17)
    early_speed_index = _clip(early_speed * 0.55 + position_escape * 0.25 + position_front * 0.20)
    style = classify_running_style(first_positions, position_ratios, style_weights)

    if len(position_ratios) >= 2:
        style_stability = _clip(100.0 - pstdev(position_ratios) * 145.0)
    elif len(position_ratios) == 1:
        style_stability = 55.0
    else:
        style_stability = 0.0

    denominator = max(used_count, 1)
    return HorsePaceProfile(
        horse_number=horse_number,
        horse_name=horse_name,
        running_style=style,
        escape_index=round(escape_index, 1),
        front_index=round(front_index, 1),
        early_speed_index=round(early_speed_index, 1),
        style_stability=round(style_stability, 1),
        first_place_rate=round(first_place_rate * 100.0, 1),
        top3_rate=round(top3_rate * 100.0, 1),
        average_position_ratio=None if avg_ratio is None else round(avg_ratio, 3),
        used_run_count=used_count,
        surface_match_rate=round(surface_matches / denominator * 100.0, 1),
        near_distance_rate=round(near_distances / denominator * 100.0, 1),
        passing_coverage_rate=round(passing_count / denominator * 100.0, 1),
        pace_coverage_rate=round(pace_count / denominator * 100.0, 1),
        lap_coverage_rate=round(lap_count / denominator * 100.0, 1),
        late_start_count=late_start_count,
        large_distance_change_count=large_distance_changes,
    )


def _historical_pace_samples(
    all_records: Sequence[Mapping[str, object]],
    current_surface: str,
    current_distance: Optional[int],
) -> List[Tuple[float, float, float]]:
    """過去レース全体のペースをrace_id単位で重複除去して返す。"""
    unique: Dict[str, Tuple[float, float, float]] = {}
    for index, record in enumerate(all_records):
        first, last, _ = _extract_early_late_3f(record)
        if first is None or last is None:
            continue
        race_id = _normalize_race_id(_record_value(record, "race_id", "レースID"))
        fallback = f"{_record_value(record, '日付')}|{_record_value(record, '開催')}|{index}"
        past_course = _record_value(record, "距離", "コース")
        past_surface = parse_course(past_course)
        past_distance = parse_distance(past_course)
        weight = _course_similarity(current_surface, past_surface) * _distance_similarity(
            current_distance, past_distance
        )
        unique.setdefault(race_id or fallback, (first, last, weight))
    return list(unique.values())


def _pace_class_from_score(score: float, surface: str, distance: Optional[int]) -> Tuple[str, str]:
    """条件別閾値を使い、総合スコアを5区分とS/M/Hへ変換する。"""
    thresholds = PACE_CLASS_THRESHOLDS.get(
        (surface, _distance_band(distance)),
        (22.0, 40.0, 62.0, 81.0),
    )
    if score < thresholds[0]:
        return "超スロー", "S"
    if score < thresholds[1]:
        return "スロー", "S"
    if score < thresholds[2]:
        return "ミドル", "M"
    if score < thresholds[3]:
        return "ハイ", "H"
    return "超ハイ", "H"


def _candidate_text(profiles: Sequence[HorsePaceProfile], index_name: str) -> str:
    """候補馬を「3番馬名（指数88.4）」形式で連結する。"""
    parts: List[str] = []
    for profile in profiles:
        number = "?" if profile.horse_number in (None, "") else str(profile.horse_number).replace(".0", "")
        index = profile.escape_index if index_name == "escape" else profile.front_index
        label = "逃げ指数" if index_name == "escape" else "先行指数"
        parts.append(f"{number}番{profile.horse_name}（{label}{index:.1f}）")
    return "、".join(parts)


def calculate_race_pace_prediction(
    race_info: Mapping[str, object],
    profiles: Sequence[HorsePaceProfile],
    all_past_records: Sequence[Mapping[str, object]],
    past_sheet_missing: bool = False,
) -> Dict[str, object]:
    """各馬プロフィールを統合し、1レース分のペース予想を作る。"""
    race_id = _normalize_race_id(_record_value(race_info, "レースID", "rid_str", "race_id")) or ""
    course_value = _record_value(race_info, "コース")
    surface = parse_course(course_value)
    distance = parse_distance(course_value) or parse_distance(_record_value(race_info, "距離"))
    known_profiles = [profile for profile in profiles if profile.used_run_count > 0]
    escape_candidates = sorted(
        [profile for profile in profiles if profile.escape_index >= 50.0],
        key=lambda item: (item.escape_index, item.early_speed_index),
        reverse=True,
    )
    strong_escape_candidates = [profile for profile in escape_candidates if profile.escape_index >= 70.0]
    front_candidates = sorted(
        [profile for profile in profiles if profile.front_index >= 60.0],
        key=lambda item: (item.front_index, item.early_speed_index),
        reverse=True,
    )

    if len(escape_candidates) >= 2:
        top_two = escape_candidates[:2]
        competition_index = _clip(
            18.0
            + sum(profile.escape_index for profile in top_two) * 0.36
            + sum(profile.early_speed_index for profile in top_two) * 0.14
        )
    elif len(escape_candidates) == 1:
        competition_index = _clip(escape_candidates[0].escape_index * 0.43)
    else:
        competition_index = 15.0 if known_profiles else 0.0

    if len(escape_candidates) == 1 and escape_candidates[0].escape_index >= 70.0:
        second_escape = max((profile.escape_index for profile in profiles if profile != escape_candidates[0]), default=0.0)
        solo_escape = "高い" if escape_candidates[0].escape_index - second_escape >= 15.0 else "中程度"
    elif len(escape_candidates) == 1:
        solo_escape = "中程度"
    elif not escape_candidates:
        solo_escape = "不明"
    else:
        solo_escape = "低い"

    pressure = (
        13.0
        + len(strong_escape_candidates) * 14.0
        + max(0, len(escape_candidates) - len(strong_escape_candidates)) * 8.0
        + len(front_candidates) * 4.0
        + competition_index * 0.27
    )
    if solo_escape == "高い":
        pressure -= 16.0
    if not escape_candidates:
        pressure -= 7.0
    pressure = _clip(pressure)

    samples = _historical_pace_samples(all_past_records, surface, distance)
    usable_samples = [sample for sample in samples if sample[2] >= 0.20]
    total_sample_weight = sum(weight for _, _, weight in usable_samples)
    predicted_first: Optional[float] = None
    predicted_last: Optional[float] = None
    historical_score: Optional[float] = None
    if len(usable_samples) >= MIN_LAP_SAMPLE_COUNT and total_sample_weight >= 1.50:
        base_first = sum(first * weight for first, _, weight in usable_samples) / total_sample_weight
        base_last = sum(last * weight for _, last, weight in usable_samples) / total_sample_weight
        # 逃げ・先行圧が高いほど前半を速く、後半をやや遅くする穏やかな補正。
        pressure_adjustment = (pressure - 50.0) / 50.0
        predicted_first = round(base_first - pressure_adjustment * 0.55, 1)
        predicted_last = round(base_last + pressure_adjustment * 0.35, 1)
        historical_difference = base_first - base_last
        historical_score = _clip(50.0 - historical_difference * 11.0)

    if not known_profiles:
        pace_score = 50.0
    elif historical_score is None:
        pace_score = pressure
    else:
        pace_score = pressure * 0.68 + historical_score * 0.32
    pace_label, pace_symbol = _pace_class_from_score(pace_score, surface, distance)

    total_horses = max(len(profiles), 1)
    average_runs = sum(profile.used_run_count for profile in profiles) / total_horses
    run_coverage = _clip(average_runs / MAX_PAST_RUNS * 100.0)
    surface_coverage = sum(profile.surface_match_rate for profile in profiles) / total_horses
    distance_coverage = sum(profile.near_distance_rate for profile in profiles) / total_horses
    passing_coverage = sum(profile.passing_coverage_rate for profile in profiles) / total_horses
    pace_coverage = sum(profile.pace_coverage_rate for profile in profiles) / total_horses
    lap_coverage = sum(profile.lap_coverage_rate for profile in profiles) / total_horses
    data_fill = _clip(
        run_coverage * 0.22
        + surface_coverage * 0.18
        + distance_coverage * 0.18
        + passing_coverage * 0.22
        + pace_coverage * 0.12
        + lap_coverage * 0.08
    )
    average_stability = (
        sum(profile.style_stability for profile in known_profiles) / len(known_profiles)
        if known_profiles
        else 0.0
    )
    if escape_candidates:
        escape_clarity = escape_candidates[0].escape_index
        if len(escape_candidates) >= 2:
            escape_clarity = _clip(55.0 + abs(escape_candidates[0].escape_index - escape_candidates[1].escape_index))
    else:
        escape_clarity = 20.0
    confidence = _clip(data_fill * 0.62 + average_stability * 0.23 + escape_clarity * 0.15)

    top_escape = escape_candidates[0] if escape_candidates else None
    second_escape = escape_candidates[1] if len(escape_candidates) >= 2 else None
    if pace_symbol == "H":
        favorable_style, unfavorable_style = "差し・追込", "逃げ・先行"
    elif pace_symbol == "S":
        favorable_style, unfavorable_style = "逃げ・先行", "差し・追込"
    else:
        favorable_style, unfavorable_style = "先行・差し", "極端な追込"

    reasons: List[str] = []
    if len(strong_escape_candidates) >= 2:
        reasons.append(f"逃げ指数70以上が{len(strong_escape_candidates)}頭")
    elif solo_escape == "高い" and top_escape is not None:
        reasons.append(f"明確な逃げ候補が{top_escape.horse_name}1頭")
    elif not escape_candidates:
        reasons.append("明確な逃げ候補が不在")
    else:
        reasons.append(f"逃げ候補が{len(escape_candidates)}頭")
    reasons.append(f"先行候補が{len(front_candidates)}頭")
    reasons.append(f"ペース圧力指数{pressure:.1f}")

    notes: List[str] = []
    new_horse_count = sum(profile.used_run_count == 0 for profile in profiles)
    if past_sheet_missing:
        notes.append("過去走シートなし")
    if new_horse_count:
        notes.append(f"過去走なし{new_horse_count}頭")
    if predicted_first is None or predicted_last is None:
        notes.append("有効なラップデータ不足")
    late_start_total = sum(profile.late_start_count for profile in profiles)
    if late_start_total:
        notes.append(f"出遅れ等の履歴{late_start_total}件")
    large_change_total = sum(profile.large_distance_change_count for profile in profiles)
    if large_change_total:
        notes.append(f"大幅距離変更履歴{large_change_total}件")
    if not escape_candidates:
        notes.append("押し出され逃げの可能性があり信頼度低下")

    field_size = _to_positive_int(_record_value(race_info, "頭数")) or len(profiles)
    difference = (
        round(predicted_first - predicted_last, 1)
        if predicted_first is not None and predicted_last is not None
        else None
    )
    return {
        "レースID": race_id,
        "発走時刻": _record_value(race_info, "発走時刻"),
        "場所": _record_value(race_info, "場所", "競馬場"),
        "レース名": _record_value(race_info, "レース名"),
        "コース": course_value,
        "馬場": _record_value(race_info, "馬場"),
        "頭数": field_size,
        "クラス": _record_value(race_info, "クラス"),
        "予想ペース": pace_label,
        "ペース記号": pace_symbol,
        "予想前半3F": predicted_first,
        "予想後半3F": predicted_last,
        "前後半差": difference,
        "ペース圧力指数": round(pressure, 1),
        "逃げ競合指数": round(competition_index, 1),
        "逃げ候補数": len(escape_candidates),
        "先行候補数": len(front_candidates),
        "本命逃げ候補_馬番": None if top_escape is None else top_escape.horse_number,
        "本命逃げ候補_馬名": None if top_escape is None else top_escape.horse_name,
        "本命逃げ候補_逃げ指数": None if top_escape is None else top_escape.escape_index,
        "対抗逃げ候補_馬番": None if second_escape is None else second_escape.horse_number,
        "対抗逃げ候補_馬名": None if second_escape is None else second_escape.horse_name,
        "対抗逃げ候補_逃げ指数": None if second_escape is None else second_escape.escape_index,
        "逃げ候補一覧": _candidate_text(escape_candidates, "escape"),
        "先行候補一覧": _candidate_text(front_candidates, "front"),
        "単騎逃げ可能性": solo_escape,
        "展開有利脚質": favorable_style,
        "展開不利脚質": unfavorable_style,
        "予想信頼度": round(confidence, 1),
        "データ充足率": round(data_fill, 1),
        "判定理由": "、".join(reasons) + "と判断",
        "注意事項": "、".join(notes),
    }


def build_pace_prediction_dataframe(excel_path: str) -> pd.DataFrame:
    """Excel内の今走情報と12桁過去走シートからレース単位DataFrameを作る。"""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if NOW_SHEET_NAME not in workbook.sheetnames:
            raise PacePredictionError(f"'{NOW_SHEET_NAME}'シートがありません")
        current_records = _worksheet_records(workbook[NOW_SHEET_NAME])
        grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
        for record in current_records:
            race_id = _normalize_race_id(_record_value(record, "レースID", "rid_str", "race_id"))
            if race_id:
                grouped[race_id].append(record)
        if not grouped:
            raise PacePredictionError("今走レース情報から12桁レースIDを取得できません")

        print(f"[INFO] ペース予想対象レース数: {len(grouped)}")
        output_rows: List[Dict[str, object]] = []
        for race_id, race_rows in grouped.items():
            race_info = race_rows[0]
            past_sheet_missing = race_id not in workbook.sheetnames
            past_records = [] if past_sheet_missing else _worksheet_records(workbook[race_id])
            histories_by_horse: Dict[str, List[Dict[str, object]]] = defaultdict(list)
            for past_record in past_records:
                name = str(_record_value(past_record, "馬名") or "").strip()
                if name:
                    histories_by_horse[name].append(past_record)

            # 同一馬が今走情報に重複していても、馬名と馬番の組み合わせで1頭にまとめる。
            current_horses: Dict[Tuple[str, str], Mapping[str, object]] = {}
            for race_row in race_rows:
                horse_name = str(_record_value(race_row, "馬名") or "").strip()
                horse_number = _record_value(race_row, "馬番")
                key = (horse_name, str(horse_number))
                current_horses.setdefault(key, race_row)

            current_course = _record_value(race_info, "コース")
            current_distance = parse_distance(current_course) or parse_distance(_record_value(race_info, "距離"))
            profiles = [
                calculate_horse_pace_profile(
                    horse_number=_record_value(horse_row, "馬番"),
                    horse_name=horse_name,
                    past_records=histories_by_horse.get(horse_name, []),
                    current_course=current_course,
                    current_distance=current_distance,
                )
                for (horse_name, _), horse_row in current_horses.items()
            ]
            prediction = calculate_race_pace_prediction(
                race_info=race_info,
                profiles=profiles,
                all_past_records=past_records,
                past_sheet_missing=past_sheet_missing,
            )
            if prediction["予想前半3F"] is None:
                print(f"[WARN] レースID={race_id} 有効なラップデータが不足しています")
            print(
                f"[INFO] レースID={race_id} ペース={prediction['予想ペース']} "
                f"信頼度={prediction['予想信頼度']:.1f}"
            )
            output_rows.append(prediction)
    finally:
        workbook.close()

    return pd.DataFrame(output_rows, columns=list(PACE_OUTPUT_COLUMNS))


def _target_snapshot(ws: Optional[Worksheet]) -> Optional[Tuple[int, Tuple[str, ...], Tuple[Tuple[object, ...], ...]]]:
    """TARGETの行数と識別・score・rank列を非変更検証用に保存する。"""
    if ws is None:
        return None
    row_iterator = ws.iter_rows(values_only=True)
    try:
        headers = [normalize_column_name(value) for value in next(row_iterator)]
    except StopIteration:
        raise PacePredictionError("TARGETシートが空です")
    required_candidates = (
        ("ridstr", "レースid", "raceid"),
        ("馬番",),
        ("score",),
        ("rank",),
    )
    selected_indexes: List[int] = []
    selected_names: List[str] = []
    for candidates in required_candidates:
        found = next((headers.index(candidate) for candidate in candidates if candidate in headers), None)
        if found is None:
            raise PacePredictionError(f"TARGETの検証列がありません: {candidates[0]}")
        selected_indexes.append(found)
        selected_names.append(headers[found])
    rows = tuple(
        tuple(values[column_index] if column_index < len(values) else None for column_index in selected_indexes)
        for values in row_iterator
    )
    return len(rows), tuple(selected_names), rows


def _worksheet_value_snapshot(ws: Optional[Worksheet]) -> Optional[Tuple[Tuple[object, ...], ...]]:
    """指定シートのセル値を、非変更検証用の行タプルとして保存する。"""
    if ws is None:
        return None
    return tuple(tuple(values) for values in ws.iter_rows(values_only=True))


def _write_pace_sheet(workbook: object, pace_df: pd.DataFrame, sheet_name: str) -> None:
    """ペース予想DataFrameを指定ブックへ書式付きで出力する。"""
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for column_index, header in enumerate(PACE_OUTPUT_COLUMNS, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_index, row in enumerate(pace_df.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            worksheet.cell(row=row_index, column=column_index, value=value)

    worksheet.auto_filter.ref = f"A1:{get_column_letter(len(PACE_OUTPUT_COLUMNS))}{worksheet.max_row}"
    header_index = {name: index + 1 for index, name in enumerate(PACE_OUTPUT_COLUMNS)}
    long_columns = {"逃げ候補一覧", "先行候補一覧", "判定理由", "注意事項"}
    one_decimal_columns = {
        "予想前半3F",
        "予想後半3F",
        "前後半差",
        "ペース圧力指数",
        "逃げ競合指数",
        "本命逃げ候補_逃げ指数",
        "対抗逃げ候補_逃げ指数",
        "予想信頼度",
        "データ充足率",
    }
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top")
        for column_name in long_columns:
            worksheet.cell(row=row[0].row, column=header_index[column_name]).alignment = Alignment(
                vertical="top", wrap_text=True
            )
        worksheet.cell(row=row[0].row, column=header_index["レースID"]).number_format = "@"
        for column_name in one_decimal_columns:
            worksheet.cell(row=row[0].row, column=header_index[column_name]).number_format = "0.0"

        symbol = worksheet.cell(row=row[0].row, column=header_index["ペース記号"]).value
        fill_color = {"S": "DDEBF7", "M": "FFF2CC", "H": "FCE4D6"}.get(str(symbol))
        if fill_color:
            fill = PatternFill("solid", fgColor=fill_color)
            worksheet.cell(row=row[0].row, column=header_index["予想ペース"]).fill = fill
            worksheet.cell(row=row[0].row, column=header_index["ペース記号"]).fill = fill

    for column_index, column_name in enumerate(PACE_OUTPUT_COLUMNS, start=1):
        values = [str(worksheet.cell(row=row, column=column_index).value or "") for row in range(1, worksheet.max_row + 1)]
        maximum = max((len(value) for value in values), default=len(column_name))
        if column_name in long_columns:
            width = min(max(maximum * 0.85, 22.0), 48.0)
        else:
            width = min(max(maximum * 1.15 + 2.0, 9.0), 24.0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width
    worksheet.row_dimensions[1].height = 42


def _verify_pace_dataframe(pace_df: pd.DataFrame, expected_race_count: int) -> None:
    """ペース予想データの必須条件を検証し、結果をログへ出す。"""
    print("[verify] ===== ペース予想の自動検証 =====")
    checks = [
        ("今走レース数とペース予想行数が一致", len(pace_df) == expected_race_count),
        ("レースID重複なし", not pace_df["レースID"].duplicated().any()),
        ("レースID欠損なし", pace_df["レースID"].astype(str).str.fullmatch(r"\d{12}").all()),
        ("予想ペース欠損なし", pace_df["予想ペース"].notna().all()),
        ("ペース記号がS/M/Hのみ", pace_df["ペース記号"].isin(["S", "M", "H"]).all()),
        (
            "各指数が0～100以内",
            all(
                pd.to_numeric(pace_df[column], errors="coerce").dropna().between(0, 100).all()
                for column in (
                    "ペース圧力指数",
                    "逃げ競合指数",
                    "本命逃げ候補_逃げ指数",
                    "対抗逃げ候補_逃げ指数",
                    "データ充足率",
                )
            ),
        ),
        (
            "信頼度が0～100以内",
            pd.to_numeric(pace_df["予想信頼度"], errors="coerce").between(0, 100).all(),
        ),
    ]
    failed = []
    for label, passed in checks:
        print(f"[verify] {label}: {'OK' if passed else 'NG'}")
        if not passed:
            failed.append(label)
    if failed:
        raise PacePredictionError(f"ペース予想の検証に失敗しました: {failed}")


def _verify_candidate_pairs(pace_df: pd.DataFrame, excel_path: str) -> None:
    """本命・対抗逃げ候補の馬番と馬名が今走情報の同一馬を指すか検証する。"""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        current_pairs = set()
        for record in _worksheet_records(workbook[NOW_SHEET_NAME]):
            race_id = _normalize_race_id(_record_value(record, "レースID", "rid_str", "race_id"))
            horse_number = _to_positive_int(_record_value(record, "馬番"))
            horse_name = str(_record_value(record, "馬名") or "").strip()
            if race_id and horse_number is not None and horse_name:
                current_pairs.add((race_id, horse_number, horse_name))
    finally:
        workbook.close()

    mismatches = []
    for _, row in pace_df.iterrows():
        race_id = str(row["レースID"])
        for prefix in ("本命逃げ候補", "対抗逃げ候補"):
            horse_name_value = row[f"{prefix}_馬名"]
            horse_number_value = row[f"{prefix}_馬番"]
            if pd.isna(horse_name_value) and pd.isna(horse_number_value):
                continue
            horse_number = _to_positive_int(horse_number_value)
            horse_name = str(horse_name_value or "").strip()
            if horse_number is None or (race_id, horse_number, horse_name) not in current_pairs:
                mismatches.append((race_id, horse_number, horse_name))
    passed = not mismatches
    print(f"[verify] 逃げ候補の馬番と馬名が今走情報と一致: {'OK' if passed else 'NG'}")
    if not passed:
        raise PacePredictionError(f"逃げ候補の馬番・馬名対応が不正です: {mismatches[:5]}")


def _verify_target_unchanged(
    before: Optional[Tuple[int, Tuple[str, ...], Tuple[Tuple[object, ...], ...]]],
    after: Optional[Tuple[int, Tuple[str, ...], Tuple[Tuple[object, ...], ...]]],
) -> None:
    """TARGETの行数・rid_str・馬番・score・rankが変わっていないことを検証する。"""
    if before is None and after is None:
        print("[verify] TARGETシートなしのため非変更検証: SKIP")
        return
    if before is None or after is None:
        raise PacePredictionError("ペース予想追加前後でTARGETシートの有無が変わりました")
    row_count_ok = before[0] == after[0]
    all_rows_ok = before == after
    print(f"[verify] 既存のTARGET行数に変化なし: {'OK' if row_count_ok else 'NG'}")
    print(f"[verify] 既存のscoreに変化なし: {'OK' if all_rows_ok else 'NG'}")
    print(f"[verify] 既存のrankに変化なし: {'OK' if all_rows_ok else 'NG'}")
    if not row_count_ok or not all_rows_ok:
        raise PacePredictionError("TARGETのrid_str・馬番・score・rankのいずれかが変化しました")


def _verify_sheet_unchanged(
    sheet_name: str,
    before: Optional[Tuple[Tuple[object, ...], ...]],
    after: Optional[Tuple[Tuple[object, ...], ...]],
) -> None:
    """既存シートのセル値が追加前後で変わっていないことを検証する。"""
    if before is None and after is None:
        print(f"[verify] 既存の{sheet_name}シートなし: SKIP")
        return
    unchanged = before == after
    print(f"[verify] 既存の{sheet_name}内容に変化なし: {'OK' if unchanged else 'NG'}")
    if not unchanged:
        raise PacePredictionError(f"既存の{sheet_name}シート内容が変化しました")


def _count_current_races(excel_path: str) -> int:
    """今走レース情報に含まれるユニークな12桁レースID数を返す。"""
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if NOW_SHEET_NAME not in workbook.sheetnames:
            raise PacePredictionError(f"'{NOW_SHEET_NAME}'シートがありません")
        race_ids = {
            _normalize_race_id(_record_value(record, "レースID", "rid_str", "race_id"))
            for record in _worksheet_records(workbook[NOW_SHEET_NAME])
        }
        return len({race_id for race_id in race_ids if race_id})
    finally:
        workbook.close()


def append_pace_prediction_sheet_to_excel(
    excel_path: str,
    sheet_name: str = DEFAULT_PACE_SHEET_NAME,
) -> int:
    """
    最終出力Excel内の今走情報と過去走シートを使用して、
    レース単位のペース予想を作成し、新規シートへ出力する。

    戻り値:
        出力したレース数。失敗時は既存Excelを変更せず0を返す。
    """
    print("[INFO] ===== ペース予想処理を開始します =====")
    path = Path(excel_path)
    temporary_path: Optional[Path] = None
    try:
        if not path.exists():
            raise FileNotFoundError(f"最終出力Excelが見つかりません: {path}")
        if path.suffix.lower() not in {".xlsx", ".xlsm"}:
            raise PacePredictionError(f"未対応のExcel形式です: {path.suffix}")

        pace_df = build_pace_prediction_dataframe(str(path))
        expected_race_count = _count_current_races(str(path))
        _verify_pace_dataframe(pace_df, expected_race_count)
        _verify_candidate_pairs(pace_df, str(path))

        keep_vba = path.suffix.lower() == ".xlsm"
        workbook = load_workbook(path, data_only=False, keep_vba=keep_vba)
        try:
            before_target = _target_snapshot(
                workbook[TARGET_SHEET_NAME] if TARGET_SHEET_NAME in workbook.sheetnames else None
            )
            before_now = _worksheet_value_snapshot(workbook[NOW_SHEET_NAME])
            before_bet = _worksheet_value_snapshot(
                workbook[BET_SHEET_NAME] if BET_SHEET_NAME in workbook.sheetnames else None
            )
            _write_pace_sheet(workbook, pace_df, sheet_name)
            file_handle, temporary_name = tempfile.mkstemp(
                prefix=f".{path.stem}_pace_",
                suffix=path.suffix,
                dir=path.parent,
            )
            os.close(file_handle)
            temporary_path = Path(temporary_name)
            workbook.save(temporary_path)
        finally:
            workbook.close()

        # 通常モードで開いた大規模ブックのセルを解放してから、読取専用検証を行う。
        del workbook
        gc.collect()

        verification_workbook = load_workbook(temporary_path, read_only=True, data_only=False, keep_vba=keep_vba)
        try:
            if sheet_name not in verification_workbook.sheetnames:
                raise PacePredictionError(f"'{sheet_name}'シートが保存されていません")
            saved_sheet = verification_workbook[sheet_name]
            saved_row_count = max(saved_sheet.max_row - 1, 0)
            if saved_row_count != expected_race_count:
                raise PacePredictionError(
                    f"保存後の行数が一致しません: expected={expected_race_count}, actual={saved_row_count}"
                )
            after_target = _target_snapshot(
                verification_workbook[TARGET_SHEET_NAME]
                if TARGET_SHEET_NAME in verification_workbook.sheetnames
                else None
            )
            after_now = _worksheet_value_snapshot(verification_workbook[NOW_SHEET_NAME])
            after_bet = _worksheet_value_snapshot(
                verification_workbook[BET_SHEET_NAME]
                if BET_SHEET_NAME in verification_workbook.sheetnames
                else None
            )
        finally:
            verification_workbook.close()
        del verification_workbook
        gc.collect()
        _verify_target_unchanged(before_target, after_target)
        _verify_sheet_unchanged(NOW_SHEET_NAME, before_now, after_now)
        _verify_sheet_unchanged(BET_SHEET_NAME, before_bet, after_bet)

        # 検証済み一時ファイルだけを最終ファイルへ置換し、途中失敗時の破損を防止する。
        os.replace(temporary_path, path)
        temporary_path = None
        print(f"[INFO] ペース予想シートを作成しました: {len(pace_df)}レース")
        print("[INFO] ===== ペース予想処理が完了しました =====")
        return len(pace_df)
    except PermissionError as exc:
        print(f"[ERROR] ペース予想シートを書き込めません。Excelが開いていないか確認してください: {exc}")
    except Exception as exc:
        print(f"[ERROR] ペース予想シートの作成に失敗しました: {type(exc).__name__}: {exc}")
    finally:
        if temporary_path is not None and temporary_path.exists():
            try:
                temporary_path.unlink()
            except OSError:
                pass
    print("[INFO] ===== ペース予想処理を終了します =====")
    return 0


__all__ = [
    "append_pace_prediction_sheet_to_excel",
    "build_pace_prediction_dataframe",
    "calculate_horse_pace_profile",
    "calculate_race_pace_prediction",
    "classify_running_style",
    "normalize_column_name",
    "parse_course",
    "parse_distance",
    "parse_lap_times",
    "parse_pace_value",
    "parse_passing_position",
]
